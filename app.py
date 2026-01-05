import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.io as pio
from datetime import datetime, timedelta
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import numpy as np
import time
from supabase import create_client, Client
import sys
from collections import defaultdict
import json
import traceback
import hashlib

# Page configuration
st.set_page_config(
    page_title="Dashboard Pengadaan BULOG",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Set default Plotly template to light theme
pio.templates.default = "plotly_white"

# Custom CSS - Light Blue Theme
st.markdown("""
<style>
            
    .st-emotion-cache-13k62yr {
        background-color: #f5f7fa;
    }
    /* Main Container */
    .main {
        background: linear-gradient(135deg, #f5f7fa 0%, #e8f4f8 100%);
        padding: 2rem;
    }
            
    .stAppToolbar{
        background: linear-gradient(135deg, #1f497d 0%, #4bacc6 100%);
    }
    
    /* Progress Bar */
    .stProgress > div > div > div > div {
        background: linear-gradient(90deg, #1f497d 0%, #4bacc6 100%);
        border-radius: 5px;
    }
    .stProgress > div > div > div {
        background-color: #e0e0e0; /* Light grey for the empty part */
        border-radius: 5px;
    }
    /* Style for the text of the progress bar */
    .stProgress label {
        color: #1f497d;
        font-weight: bold;
    }
    
    /* Header Title */
    .dashboard-header {
        background: linear-gradient(135deg, #1f497d 0%, #4bacc6 100%);
        padding: 2rem;
        border-radius: 15px;
        text-align: center;
        margin-bottom: 2rem;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
    }

    .dashboard-title {
        color: white;
        font-size: 2.5rem;
        font-weight: 700;
        margin: 0;
        text-shadow: 2px 2px 4px rgba(0, 0, 0, 0.2);
    }

    .dashboard-subtitle {
        color: #e8f4f8;
        font-size: 1.1rem;
        margin-top: 0.5rem;
    }

    /* Metrics Cards */
    div[data-testid="stMetric"] {
        background: white;
        padding: 1rem;
        border-radius: 10px;
        border-left: 4px solid #4bacc6;
        box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
    }

    div[data-testid="stMetric"] label {
        color: #000000 !important;
    }

    div[data-testid="stMetricValue"] {
        font-size: 1.8rem;
        color: #000000 !important;
        font-weight: 700;
    }

    div[data-testid="stMetricDelta"] {
        color: #000000 !important;
    }

    /* Sidebar */
    section[data-testid="stSidebar"] {
        background: linear-gradient(180deg, #1f497d 0%, #2c5f8d 100%);
    }


    /* Chart Container */
    .chart-container {
        background: white;
        padding: 2rem;
        border-radius: 15px;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
        margin: 2rem 0;
    }

    .chart-title {
        color: #1f497d;
        font-size: 1.5rem;
        font-weight: 700;
        margin-bottom: 1rem;
        border-bottom: 3px solid #4bacc6;
        padding-bottom: 0.5rem;
    }

    /* Table Container */
    .table-container {
        padding: 2rem;
        border-radius: 15px;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
        margin: 2rem 0;
    }

    /* Hide Streamlit branding */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}

    /* Better spacing */
    .block-container {
        padding-top: 1rem;
        padding-bottom: 1rem;
    }
    tbody{
            color: #1f497d;}
    stAppToolbar st-emotion-cache-14vh5up e1o8oa9v2 {
        background-color: #f5f7fa;
            }
</style>
""", unsafe_allow_html=True)

# Konfigurasi Supabase
try:
    SUPABASE_URL = st.secrets["supabase"]["project_url"]
    SUPABASE_KEY = st.secrets["supabase"]["api_key"]
except:
    st.error("ERROR: Tidak dapat membaca secrets.toml")
    st.info("Pastikan file .streamlit/secrets.toml sudah dikonfigurasi dengan benar")
    st.stop()

# Inisialisasi Supabase client
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# ===== HELPER FUNCTIONS UNTUK MIGRASI DATA =====
def clean_value(value):
    """Bersihkan nilai untuk database (handle NaN, None, empty string)"""
    if pd.isna(value) or value == '' or value == 'nan' or value == 'None':
        return None
    if isinstance(value, str):
        value_stripped = value.strip()
        if value_stripped == '' or value_stripped.lower() in ('nan', 'none', 'null'):
            return None
        return value_stripped
    return value

def convert_to_decimal(value):
    """Convert value ke Decimal/float dengan presisi yang tepat"""
    from decimal import Decimal
    if pd.isna(value):
        return None
    try:
        if isinstance(value, str):
            return Decimal(value)
        else:
            return Decimal(str(value))
    except:
        return None

def convert_to_int(value):
    """Convert value ke integer"""
    if pd.isna(value):
        return None
    try:
        return int(float(value))
    except:
        return None

def convert_to_date(value):
    """Convert value ke date string (YYYY-MM-DD)"""
    if pd.isna(value):
        return None
    try:
        if isinstance(value, str):
            dt = pd.to_datetime(value)
            return dt.strftime('%Y-%m-%d')
        else:
            return value.strftime('%Y-%m-%d') if hasattr(value, 'strftime') else str(value)
    except:
        return None

def prepare_realisasi_for_db(df, kanwil_map, kancab_map):
    """Prepare dataframe realisasi untuk insert ke database"""
    data_to_insert = []

    for idx, row in df.iterrows():
        # Map nama ke ID
        kanwil_name = clean_value(row.get('kanwil'))
        kancab_name = clean_value(row.get('Entitas'))

        kanwil_id = kanwil_map.get(kanwil_name) if kanwil_name else None
        kancab_id = kancab_map.get(kancab_name) if kancab_name else None

        # Prepare record dengan handling tipe data yang tepat
        record = {
            "kanwil_id": kanwil_id,
            "kancab_id": kancab_id,
            "lokasi_persediaan": clean_value(row.get('Lokasi Persediaan')),
            "id_pemasok": convert_to_int(row.get('No. ID Pemasok')),
            "nama_pemasok": clean_value(row.get('Nama Pemasok')),
            "tanggal_po": convert_to_date(row.get('Tanggal PO')),
            "nomor_po": clean_value(row.get('Nomor PO')),
            "produk": clean_value(row.get('Produk')),
            "no_jurnal": clean_value(row.get('No Jurnal')),
            "no_in_out": clean_value(row.get('Nomor IN / OUT')),
            "tanggal_penerimaan": convert_to_date(row.get('Tanggal Penerimaan')),
            "komoditi": clean_value(row.get('Komoditi')),
            "spesifikasi": clean_value(row.get('spesifikasi')),
            "tahun_stok": convert_to_int(row.get('Tahun Stok')),
            "tanggal_kirim_keuangan": convert_to_date(row.get('Tanggal Kirim Keuangan')),
            "jenis_transaksi": clean_value(row.get('Jenis Transaksi')),
            "akun_analitik": clean_value(row.get('Akun Analitik')),
            "jenis_pengadaan": clean_value(row.get('Jenis Pengadaan')),
            "satuan": clean_value(row.get('Satuan')),
            "uom_po": clean_value(row.get('uom_po')),
            "kuantum_po_kg": float(convert_to_decimal(row.get('Kuantum PO (Kg)'))) if convert_to_decimal(row.get('Kuantum PO (Kg)')) else None,
            "qty_in_out": float(convert_to_decimal(row.get('In / Out'))) if convert_to_decimal(row.get('In / Out')) else None,
            "harga_include_ppn": float(convert_to_decimal(row.get('Harga Include ppn'))) if convert_to_decimal(row.get('Harga Include ppn')) else None,
            "nominal_realisasi_incl_ppn": float(convert_to_decimal(row.get('Nominal Realisasi Incl ppn'))) if convert_to_decimal(row.get('Nominal Realisasi Incl ppn')) else None,
            "status": clean_value(row.get('Status'))
        }

        data_to_insert.append(record)

    return data_to_insert

def prepare_target_kanwil_for_db(df, kanwil_map):
    """Prepare dataframe target_kanwil untuk insert ke database"""
    data_to_insert = []

    for idx, row in df.iterrows():
        kanwil_name = clean_value(row.get('kanwil'))
        if not kanwil_name or kanwil_name not in kanwil_map:
            continue

        record = {
            "kanwil_id": kanwil_map[kanwil_name],
            "target_setara_beras": clean_value(row.get('Target Setara Beras')),
            "date": datetime.now().date().isoformat()
        }
        data_to_insert.append(record)

    return data_to_insert

def prepare_target_kancab_for_db(df, kancab_map):
    """Prepare dataframe target_kancab untuk insert ke database"""
    data_to_insert = []

    for idx, row in df.iterrows():
        kancab_name = clean_value(row.get('kancab'))
        if not kancab_name or kancab_name not in kancab_map:
            continue

        record = {
            "kancab_id": kancab_map[kancab_name],
            "target_setara_beras": clean_value(row.get('Target Setara Beras')),
            "date": datetime.now().date().isoformat()
        }
        data_to_insert.append(record)

    return data_to_insert

# ===== FUNGSI LOAD DATA DARI DATABASE =====

def load_all_realisasi_from_db_with_progress(supabase):
    """
    Load SEMUA data dari tabel realisasi per 1000 baris dengan progress bar
    Menambahkan nama_kanwil dan nama_kancab dari join
    """
    import time
    start_time = time.time()

    st.markdown("---")
    st.markdown('<h4 style="color: #1f497d;">📥 Loading Data dari Database</h4>', unsafe_allow_html=True)

    # Step 1: Get total count
    print("=" * 80)
    print("LOADING ALL DATA FROM DATABASE - START")
    print("=" * 80)

    st.info("🔄 Step 1: Menghitung total data di database...")
    count_result = supabase.table('realisasi').select('*', count='exact').execute()
    total_records = count_result.count
    st.success(f"✅ Total data di database: **{total_records:,}** records")

    print(f"[STEP 1] Total records in database: {total_records:,}")

    if total_records == 0:
        st.warning("⚠️ Tabel realisasi masih kosong")
        print("[WARNING] Table is empty")
        return pd.DataFrame(), {}, {}

    # Step 2: Load kanwil and kancab mappings
    st.info("🔄 Step 2: Loading mapping Kanwil & Kancab...")
    kanwil_map = {}  # nama_kanwil -> kanwil_id (untuk prepare function)
    kancab_map = {}  # nama_kancab -> kancab_id (untuk prepare function)
    kanwil_id_to_name = {}  # kanwil_id -> nama_kanwil (untuk add columns to df)
    kancab_id_to_name = {}  # kancab_id -> nama_kancab (untuk add columns to df)

    kanwil_result = supabase.table('kanwil').select('*').execute()
    for kw in kanwil_result.data:
        kanwil_map[kw['nama_kanwil']] = kw['kanwil_id']  # nama -> id
        kanwil_id_to_name[kw['kanwil_id']] = kw['nama_kanwil']  # id -> nama

    kancab_result = supabase.table('kancab').select('*').execute()
    for kc in kancab_result.data:
        kancab_map[kc['nama_kancab']] = kc['kancab_id']  # nama -> id
        kancab_id_to_name[kc['kancab_id']] = kc['nama_kancab']  # id -> nama

    st.success(f"✅ Loaded {len(kanwil_map)} Kanwil & {len(kancab_map)} Kancab mappings")
    print(f"[STEP 2] Loaded {len(kanwil_map)} Kanwil & {len(kancab_map)} Kancab mappings")

    # Step 3: Load data per 1000 rows
    st.info("🔄 Step 3: Loading data dari database per 1000 baris...")
    batch_size = 1000
    total_batches = (total_records + batch_size - 1) // batch_size

    print(f"[STEP 3] Will load {total_batches} batches of {batch_size} records each")
    print("-" * 80)

    all_data = []
    progress_bar = st.progress(0, f"Loading batch 1/{total_batches}...")

    # Create placeholder for real-time updates
    status_placeholder = st.empty()

    for batch_num in range(total_batches):
        batch_start = time.time()
        offset = batch_num * batch_size

        # Fetch batch
        result = supabase.table('realisasi')\
            .select('*')\
            .range(offset, offset + batch_size - 1)\
            .execute()

        # Add to list
        batch_count = len(result.data)
        all_data.extend(result.data)

        batch_elapsed = time.time() - batch_start

        # Update progress
        progress = ((batch_num + 1) / total_batches)
        progress_bar.progress(
            progress,
            f"Loading batch {batch_num + 1}/{total_batches} - {len(all_data):,}/{total_records:,} records"
        )

        # Print to console
        print(f"[BATCH {batch_num + 1}/{total_batches}] Loaded {batch_count:,} records (offset {offset:,}) - "
              f"Total so far: {len(all_data):,}/{total_records:,} ({progress*100:.1f}%) - "
              f"Time: {batch_elapsed:.2f}s")

        # Update status
        status_placeholder.info(f"📊 Loaded {len(all_data):,} / {total_records:,} records ({progress*100:.1f}%)")

    progress_bar.empty()
    status_placeholder.empty()

    print("-" * 80)
    print(f"[STEP 3 COMPLETE] Total records loaded: {len(all_data):,}")

    # Step 4: Convert to DataFrame and add nama columns
    st.info("🔄 Step 4: Converting ke DataFrame dan menambahkan nama_kanwil & nama_kancab...")
    print("[STEP 4] Converting to DataFrame...")

    df = pd.DataFrame(all_data)
    print(f"[STEP 4] DataFrame created with shape: {df.shape}")

    # Add nama_kanwil and nama_kancab columns (using id -> nama mapping)
    df['nama_kanwil'] = df['kanwil_id'].map(kanwil_id_to_name)
    df['nama_kancab'] = df['kancab_id'].map(kancab_id_to_name)

    print(f"[STEP 4] Added nama_kanwil and nama_kancab columns")
    print(f"[STEP 4] Final DataFrame columns: {df.columns.tolist()}")

    total_elapsed = time.time() - start_time
    st.success(f"✅ Berhasil load {len(df):,} records dengan nama_kanwil & nama_kancab dalam {total_elapsed:.2f} detik")

    print("=" * 80)
    print(f"LOADING COMPLETE - Total time: {total_elapsed:.2f} seconds ({total_elapsed/60:.2f} minutes)")
    print(f"Average: {len(df)/total_elapsed:.0f} records/second")
    print("=" * 80)

    # Debug: Show sample
    with st.expander("👁️ Preview Data dari Database (5 rows pertama)", expanded=False):
        st.dataframe(df.head(), use_container_width=True)
        st.write(f"Total Columns: {len(df.columns)}")
        st.write(f"Columns: {df.columns.tolist()}")

    return df, kanwil_map, kancab_map


def find_unique_records(df_db, df_new, kanwil_map, kancab_map):
    """
    Compare df_new dengan df_db dan return unique records
    Unique berdasarkan key fields yang penting
    """
    st.markdown("---")
    st.markdown('<h4 style="color: #1f497d;">🔍 Analisis Data Unik</h4>', unsafe_allow_html=True)

    st.info("🔄 Preparing data baru untuk comparison...")

    # Prepare df_new: add kanwil_id and kancab_id
    df_new_prepared = df_new.copy()

    # kanwil_map and kancab_map dari load function sudah format: nama -> id
    # Langsung map ke IDs
    df_new_prepared['kanwil_id'] = df_new_prepared['kanwil'].map(kanwil_map)
    df_new_prepared['kancab_id'] = df_new_prepared['Entitas'].map(kancab_map)

    # Define key fields for uniqueness check (sesuai prompt.txt)
    key_fields = [
        'nomor_po', 'no_in_out', 'tanggal_penerimaan',
        'komoditi', 'spesifikasi', 'kanwil_id', 'kancab_id'
    ]

    # Map Excel columns to DB columns for new data
    df_new_keys = df_new_prepared.copy()
    df_new_keys['nomor_po'] = df_new_prepared['Nomor PO'].apply(clean_value)
    df_new_keys['no_in_out'] = df_new_prepared['Nomor IN / OUT'].apply(clean_value)
    df_new_keys['tanggal_penerimaan'] = df_new_prepared['Tanggal Penerimaan'].apply(convert_to_date)
    df_new_keys['komoditi'] = df_new_prepared['Komoditi'].apply(clean_value)
    df_new_keys['spesifikasi'] = df_new_prepared['spesifikasi'].apply(clean_value)

    print(f"[DEBUG] NEW data cleaning complete - sample tanggal_penerimaan: {df_new_keys['tanggal_penerimaan'].iloc[0]}")
    print(f"[DEBUG] NEW data sample nomor_po: {df_new_keys['nomor_po'].iloc[0]}")

    st.info("🔄 Creating hash untuk comparison...")

    # Clean data dari database juga dengan cara yang sama
    df_db_cleaned = df_db.copy()
    if len(df_db) > 0:
        # Clean database fields to match format data baru
        df_db_cleaned['nomor_po'] = df_db['nomor_po'].apply(clean_value)
        df_db_cleaned['no_in_out'] = df_db['no_in_out'].apply(clean_value)

        # Normalize tanggal_penerimaan - convert to string format YYYY-MM-DD
        def normalize_date(val):
            if pd.isna(val) or val is None:
                return None
            # If already string, clean it
            if isinstance(val, str):
                return convert_to_date(val)
            # If date/datetime object, convert to string
            else:
                try:
                    return val.strftime('%Y-%m-%d') if hasattr(val, 'strftime') else str(val)
                except:
                    return str(val)

        df_db_cleaned['tanggal_penerimaan'] = df_db['tanggal_penerimaan'].apply(normalize_date)
        df_db_cleaned['komoditi'] = df_db['komoditi'].apply(clean_value)
        df_db_cleaned['spesifikasi'] = df_db['spesifikasi'].apply(clean_value)
        # kanwil_id dan kancab_id sudah integer, tidak perlu clean

        print(f"[DEBUG] DB cleaning complete - sample tanggal_penerimaan: {df_db_cleaned['tanggal_penerimaan'].iloc[0]}")

    # Create hash for comparison - normalize to string dan handle None
    def create_hash(row):
        values = []
        for field in key_fields:
            if field in row.index:
                val = row[field]
                # Convert to string, handle None/NaN
                if pd.isna(val) or val is None:
                    values.append('NULL')
                else:
                    values.append(str(val).strip().lower())
            else:
                values.append('NULL')
        return hash(tuple(values))

    # Get existing hashes from DB
    if len(df_db) > 0:
        existing_hashes = set(df_db_cleaned[key_fields].apply(create_hash, axis=1))
        print(f"[COMPARISON] Created {len(existing_hashes):,} unique hashes from database")
    else:
        existing_hashes = set()

    # Get new hashes
    new_hashes = df_new_keys[key_fields].apply(create_hash, axis=1)
    print(f"[COMPARISON] Created {len(new_hashes):,} hashes from new data")

    # Debug: Show sample data untuk comparison
    if len(df_db) > 0:
        print("\n[DEBUG] Sample comparison:")
        print("=" * 80)
        print("DB Record 1:")
        db_vals = []
        for field in key_fields:
            val = df_db_cleaned.iloc[0][field] if field in df_db_cleaned.columns else 'MISSING'
            print(f"  {field}: '{val}' (type: {type(val).__name__})")
            # Show what it becomes in hash
            if pd.isna(val) or val is None:
                db_vals.append('NULL')
            else:
                db_vals.append(str(val).strip().lower())
        print(f"  Hash values: {db_vals}")
        print(f"  Hash: {hash(tuple(db_vals))}")

        print("\nNEW Record 1:")
        new_vals = []
        for field in key_fields:
            val = df_new_keys.iloc[0][field] if field in df_new_keys.columns else 'MISSING'
            print(f"  {field}: '{val}' (type: {type(val).__name__})")
            # Show what it becomes in hash
            if pd.isna(val) or val is None:
                new_vals.append('NULL')
            else:
                new_vals.append(str(val).strip().lower())
        print(f"  Hash values: {new_vals}")
        print(f"  Hash: {hash(tuple(new_vals))}")

        print(f"\nHash match: {hash(tuple(db_vals)) == hash(tuple(new_vals))}")
        print("=" * 80)

    # Find unique
    unique_mask = ~new_hashes.isin(existing_hashes)
    df_unique = df_new[unique_mask].copy()

    num_unique = len(df_unique)
    num_duplicate = len(df_new) - num_unique

    print(f"\n[COMPARISON RESULT]")
    print(f"  Total new data: {len(df_new):,}")
    print(f"  Unique records: {num_unique:,}")
    print(f"  Duplicate records: {num_duplicate:,}")
    print(f"  Duplicate percentage: {(num_duplicate/len(df_new)*100):.1f}%")

    st.success(f"✅ Analisis selesai!")

    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("📊 Total Data Baru", f"{len(df_new):,}")
    with col2:
        st.metric("✅ Data Unik", f"{num_unique:,}")
    with col3:
        st.metric("⚠️ Data Duplikat", f"{num_duplicate:,}")

    # Debug expander untuk melihat sample comparison
    if len(df_db) > 0:
        with st.expander("🔍 Debug: Sample Comparison Data", expanded=False):
            st.write("**Key fields used for uniqueness:**")
            st.code(", ".join(key_fields))

            col_a, col_b = st.columns(2)
            with col_a:
                st.write("**Sample dari Database (cleaned):**")
                st.dataframe(df_db_cleaned[key_fields].head(3), use_container_width=True)

            with col_b:
                st.write("**Sample dari File Baru (cleaned):**")
                st.dataframe(df_new_keys[key_fields].head(3), use_container_width=True)

    return df_unique, num_unique, num_duplicate


def truncate_table_with_reset(supabase, table_name):
    """
    TRUNCATE table dan reset sequence ID menggunakan reset_table_sequence RPC
    Generic function untuk semua tabel
    """
    try:
        add_log(f"🗑️ Mereset tabel {table_name} dan sequence (TRUNCATE)...", "warning")
        st.warning(f"🗑️ Mereset tabel {table_name} dan sequence (TRUNCATE)...")
        # Use reset_table_sequence RPC function
        supabase.rpc(
            "reset_table_sequence",
            {"p_table_name": table_name}
        ).execute()
        add_log(f"✅ Tabel {table_name} telah di-truncate dan sequence di-reset", "success")
        st.success(f"✅ Tabel {table_name} telah di-truncate dan sequence di-reset")
        return True
    except Exception as e:
        add_log(f"❌ Error saat reset {table_name}: {e}", "error")
        st.error(f"❌ Error saat reset: {e}")
        add_log("🔄 Mencoba alternatif method (manual delete)...", "warning")
        st.warning("Mencoba alternatif method (manual delete)...")
        try:
            # Manual delete all rows
            supabase.table(table_name).delete().neq('id', 0).execute()
            add_log(f"✅ Tabel {table_name} berhasil dikosongkan menggunakan delete (sequence mungkin perlu reset manual)", "warning")
            st.success(f"✅ Tabel {table_name} berhasil dikosongkan menggunakan delete")
            st.warning("⚠️ Perhatian: Sequence ID mungkin perlu direset manual")
            return True
        except Exception as e2:
            add_log(f"❌ Error pada alternatif method untuk {table_name}: {e2}", "error")
            st.error(f"❌ Error pada alternatif method: {e2}")
            return False

def truncate_table_realisasi(supabase):
    """
    TRUNCATE table realisasi dan reset sequence ID
    Wrapper untuk backward compatibility
    """
    return truncate_table_with_reset(supabase, "realisasi")


# ===== FUNGSI ALGORITMA NEW COMPARISON (dari new_comparison_algorithm.py) =====

def generate_row_hash(record):
    """
    Generate SHA256 hash from record data for duplicate detection.
    Excludes auto-generated fields like id and created_at.
    """
    hash_data = {
        'kanwil_id': record.get('kanwil_id'),
        'kancab_id': record.get('kancab_id'),
        'lokasi_persediaan': record.get('lokasi_persediaan'),
        'id_pemasok': record.get('id_pemasok'),
        'nama_pemasok': record.get('nama_pemasok'),
        'tanggal_po': record.get('tanggal_po'),
        'nomor_po': record.get('nomor_po'),
        'produk': record.get('produk'),
        'no_jurnal': record.get('no_jurnal'),
        'no_in_out': record.get('no_in_out'),
        'tanggal_penerimaan': record.get('tanggal_penerimaan'),
        'komoditi': record.get('komoditi'),
        'spesifikasi': record.get('spesifikasi'),
        'tahun_stok': record.get('tahun_stok'),
        'tanggal_kirim_keuangan': record.get('tanggal_kirim_keuangan'),
        'jenis_transaksi': record.get('jenis_transaksi'),
        'akun_analitik': record.get('akun_analitik'),
        'jenis_pengadaan': record.get('jenis_pengadaan'),
        'satuan': record.get('satuan'),
        'uom_po': record.get('uom_po'),
        'kuantum_po_kg': record.get('kuantum_po_kg'),
        'qty_in_out': record.get('qty_in_out'),
        'harga_include_ppn': record.get('harga_include_ppn'),
        'nominal_realisasi_incl_ppn': record.get('nominal_realisasi_incl_ppn'),
        'status': record.get('status'),
    }

    json_string = json.dumps(hash_data, sort_keys=True, default=str)
    return hashlib.sha256(json_string.encode()).hexdigest()


def add_log(message, level="info"):
    """Add log message to session state for persistent logging"""
    if 'process_logs' not in st.session_state:
        st.session_state.process_logs = []

    import datetime
    timestamp = datetime.datetime.now().strftime("%H:%M:%S")
    log_entry = {
        'timestamp': timestamp,
        'level': level,
        'message': message
    }
    st.session_state.process_logs.append(log_entry)


def migrate_to_realisasi_compare_streamlit(supabase, df, kanwil_mapping, kancab_mapping, kancab_column='Entitas'):
    """
    Migrate data from DataFrame to realisasi_compare table (Streamlit version).
    Returns: total_inserted, skipped_kanwil, skipped_kancab

    Parameters:
    - kancab_column: Nama kolom kancab di Excel ('Entitas' untuk Realisasi, 'kancab' untuk Target Kancab)
    """
    add_log("📥 Starting Migration to realisasi_compare...", "info")
    st.info("📥 Starting Migration to realisasi_compare...")

    # Clear table first
    add_log("🗑️ Clearing realisasi_compare table...", "warning")
    st.warning("🗑️  Clearing realisasi_compare table...")
    try:
        truncate_table_with_reset(supabase, "realisasi_compare")
        add_log("✅ realisasi_compare table cleared successfully", "success")
    except Exception as e:
        add_log(f"❌ Error clearing table: {e}", "error")
        st.error(f"❌ Error clearing table: {e}")

    # Migrate data
    add_log(f"📥 Migrating {len(df):,} records to realisasi_compare...", "info")
    st.info("📥 Migrating data to realisasi_compare...")
    realisasi_compare_data = []
    batch_size = 1000
    total_inserted = 0
    skipped_kanwil = 0
    skipped_kancab = 0

    progress_bar = st.progress(0, "Processing records...")

    for idx, row in df.iterrows():
        try:
            kanwil_name = str(row['kanwil']) if pd.notna(row['kanwil']) else None
            kancab_name = str(row[kancab_column]) if pd.notna(row.get(kancab_column)) else None

            kanwil_id = kanwil_mapping.get(kanwil_name)
            kancab_id = kancab_mapping.get((kanwil_name, kancab_name))

            if not kanwil_id:
                skipped_kanwil += 1
            if not kancab_id and kancab_name:
                skipped_kancab += 1

            tanggal_po = pd.to_datetime(row['Tanggal PO']).date() if pd.notna(row['Tanggal PO']) else None
            tanggal_penerimaan = pd.to_datetime(row['Tanggal Penerimaan']).date() if pd.notna(row['Tanggal Penerimaan']) else None
            tanggal_kirim = pd.to_datetime(row['Tanggal Kirim Keuangan']).date() if pd.notna(row['Tanggal Kirim Keuangan']) else None

            record = {
                'kanwil_id': kanwil_id,
                'kancab_id': kancab_id,
                'lokasi_persediaan': str(row['Lokasi Persediaan']) if pd.notna(row['Lokasi Persediaan']) else None,
                'id_pemasok': int(row['No. ID Pemasok']) if pd.notna(row['No. ID Pemasok']) else None,
                'nama_pemasok': str(row['Nama Pemasok']) if pd.notna(row['Nama Pemasok']) else None,
                'tanggal_po': tanggal_po.isoformat() if tanggal_po else None,
                'nomor_po': str(row['Nomor PO']) if pd.notna(row['Nomor PO']) else None,
                'produk': str(row['Produk']) if pd.notna(row['Produk']) else None,
                'no_jurnal': str(row['No Jurnal']) if pd.notna(row['No Jurnal']) else None,
                'no_in_out': str(row['Nomor IN / OUT']) if pd.notna(row['Nomor IN / OUT']) else None,
                'tanggal_penerimaan': tanggal_penerimaan.isoformat() if tanggal_penerimaan else None,
                'komoditi': str(row['Komoditi']) if pd.notna(row['Komoditi']) else None,
                'spesifikasi': str(row['spesifikasi']) if pd.notna(row['spesifikasi']) else None,
                'tahun_stok': int(row['Tahun Stok']) if pd.notna(row['Tahun Stok']) else None,
                'tanggal_kirim_keuangan': tanggal_kirim.isoformat() if tanggal_kirim else None,
                'jenis_transaksi': str(row['Jenis Transaksi']) if pd.notna(row['Jenis Transaksi']) else None,
                'akun_analitik': str(row['Akun Analitik']) if pd.notna(row['Akun Analitik']) else None,
                'jenis_pengadaan': str(row['Jenis Pengadaan']) if pd.notna(row['Jenis Pengadaan']) else None,
                'satuan': str(row['Satuan']) if pd.notna(row['Satuan']) else None,
                'uom_po': str(row['uom_po']) if pd.notna(row['uom_po']) else None,
                'kuantum_po_kg': float(row['Kuantum PO (Kg)']) if pd.notna(row['Kuantum PO (Kg)']) else None,
                'qty_in_out': float(row['In / Out']) if pd.notna(row['In / Out']) else None,
                'harga_include_ppn': float(row['Harga Include ppn']) if pd.notna(row['Harga Include ppn']) else None,
                'nominal_realisasi_incl_ppn': float(row['Nominal Realisasi Incl ppn']) if pd.notna(row['Nominal Realisasi Incl ppn']) else None,
                'status': str(row['Status']) if pd.notna(row['Status']) else None,
            }

            record['row_hash'] = generate_row_hash(record)
            realisasi_compare_data.append(record)

            if len(realisasi_compare_data) >= batch_size:
                supabase.table('realisasi_compare').insert(realisasi_compare_data).execute()
                total_inserted += len(realisasi_compare_data)
                progress = (idx + 1) / len(df) * 100
                add_log(f"✅ Batch inserted: {total_inserted:,} records ({progress:.1f}%)", "success")
                progress_bar.progress(int(progress) / 100, f"Inserted {total_inserted:,} records...")
                realisasi_compare_data = []

        except Exception as e:
            add_log(f"⚠️ Error at row {idx}: {e}", "warning")
            st.warning(f"⚠️  Error at row {idx}: {e}")
            continue

    if realisasi_compare_data:
        supabase.table('realisasi_compare').insert(realisasi_compare_data).execute()
        total_inserted += len(realisasi_compare_data)
        add_log(f"✅ Final batch inserted: {total_inserted:,} total records", "success")

    progress_bar.progress(100, "✅ Migration completed")
    progress_bar.empty()

    add_log(f"📊 Migration Summary - Inserted: {total_inserted:,}, Skipped Kanwil: {skipped_kanwil:,}, Skipped Kancab: {skipped_kancab:,}", "success")
    st.success(f"""
    📊 Migration Summary:
    - Total records inserted: **{total_inserted:,}**
    - Skipped (kanwil not found): **{skipped_kanwil:,}**
    - Skipped (kancab not found): **{skipped_kancab:,}**
    """)

    return total_inserted, skipped_kanwil, skipped_kancab


def migrate_to_realisasi_direct_streamlit(supabase, df, kanwil_mapping, kancab_mapping, kancab_column='Entitas'):
    """
    Migrate data from DataFrame directly to realisasi table (REPLACE MODE - Streamlit version).
    Returns: total_inserted, skipped_kanwil, skipped_kancab

    Parameters:
    - kancab_column: Nama kolom kancab di Excel ('Entitas' untuk Realisasi, 'kancab' untuk Target Kancab)
    """
    add_log("📥 Starting Direct Migration to realisasi (REPLACE MODE)...", "warning")
    st.warning("📥 Starting Direct Migration to realisasi (REPLACE MODE)...")

    # Reset realisasi table
    add_log("🗑️ Resetting realisasi table and sequence...", "warning")
    st.info("🗑️  Resetting realisasi table...")
    if not truncate_table_with_reset(supabase, "realisasi"):
        add_log("❌ Failed to reset realisasi table. Aborting.", "error")
        st.error("❌ Failed to reset realisasi table. Aborting.")
        return 0, 0, 0
    add_log("✅ realisasi table reset successfully", "success")

    # Migrate data
    add_log(f"📥 Migrating {len(df):,} records to realisasi...", "info")
    st.info("📥 Migrating data to realisasi...")
    realisasi_data = []
    batch_size = 1000
    total_inserted = 0
    skipped_kanwil = 0
    skipped_kancab = 0

    progress_bar = st.progress(0, "Processing records...")

    for idx, row in df.iterrows():
        try:
            kanwil_name = str(row['kanwil']) if pd.notna(row['kanwil']) else None
            kancab_name = str(row[kancab_column]) if pd.notna(row.get(kancab_column)) else None

            kanwil_id = kanwil_mapping.get(kanwil_name)
            kancab_id = kancab_mapping.get((kanwil_name, kancab_name))

            if not kanwil_id:
                skipped_kanwil += 1
            if not kancab_id and kancab_name:
                skipped_kancab += 1

            tanggal_po = pd.to_datetime(row['Tanggal PO']).date() if pd.notna(row['Tanggal PO']) else None
            tanggal_penerimaan = pd.to_datetime(row['Tanggal Penerimaan']).date() if pd.notna(row['Tanggal Penerimaan']) else None
            tanggal_kirim = pd.to_datetime(row['Tanggal Kirim Keuangan']).date() if pd.notna(row['Tanggal Kirim Keuangan']) else None

            record = {
                'kanwil_id': kanwil_id,
                'kancab_id': kancab_id,
                'lokasi_persediaan': str(row['Lokasi Persediaan']) if pd.notna(row['Lokasi Persediaan']) else None,
                'id_pemasok': int(row['No. ID Pemasok']) if pd.notna(row['No. ID Pemasok']) else None,
                'nama_pemasok': str(row['Nama Pemasok']) if pd.notna(row['Nama Pemasok']) else None,
                'tanggal_po': tanggal_po.isoformat() if tanggal_po else None,
                'nomor_po': str(row['Nomor PO']) if pd.notna(row['Nomor PO']) else None,
                'produk': str(row['Produk']) if pd.notna(row['Produk']) else None,
                'no_jurnal': str(row['No Jurnal']) if pd.notna(row['No Jurnal']) else None,
                'no_in_out': str(row['Nomor IN / OUT']) if pd.notna(row['Nomor IN / OUT']) else None,
                'tanggal_penerimaan': tanggal_penerimaan.isoformat() if tanggal_penerimaan else None,
                'komoditi': str(row['Komoditi']) if pd.notna(row['Komoditi']) else None,
                'spesifikasi': str(row['spesifikasi']) if pd.notna(row['spesifikasi']) else None,
                'tahun_stok': int(row['Tahun Stok']) if pd.notna(row['Tahun Stok']) else None,
                'tanggal_kirim_keuangan': tanggal_kirim.isoformat() if tanggal_kirim else None,
                'jenis_transaksi': str(row['Jenis Transaksi']) if pd.notna(row['Jenis Transaksi']) else None,
                'akun_analitik': str(row['Akun Analitik']) if pd.notna(row['Akun Analitik']) else None,
                'jenis_pengadaan': str(row['Jenis Pengadaan']) if pd.notna(row['Jenis Pengadaan']) else None,
                'satuan': str(row['Satuan']) if pd.notna(row['Satuan']) else None,
                'uom_po': str(row['uom_po']) if pd.notna(row['uom_po']) else None,
                'kuantum_po_kg': float(row['Kuantum PO (Kg)']) if pd.notna(row['Kuantum PO (Kg)']) else None,
                'qty_in_out': float(row['In / Out']) if pd.notna(row['In / Out']) else None,
                'harga_include_ppn': float(row['Harga Include ppn']) if pd.notna(row['Harga Include ppn']) else None,
                'nominal_realisasi_incl_ppn': float(row['Nominal Realisasi Incl ppn']) if pd.notna(row['Nominal Realisasi Incl ppn']) else None,
                'status': str(row['Status']) if pd.notna(row['Status']) else None,
            }

            record['row_hash'] = generate_row_hash(record)
            realisasi_data.append(record)

            if len(realisasi_data) >= batch_size:
                supabase.table('realisasi').insert(realisasi_data).execute()
                total_inserted += len(realisasi_data)
                progress = (idx + 1) / len(df) * 100
                add_log(f"✅ Batch inserted to realisasi: {total_inserted:,} records ({progress:.1f}%)", "success")
                progress_bar.progress(int(progress) / 100, f"Inserted {total_inserted:,} records...")
                realisasi_data = []

        except Exception as e:
            add_log(f"⚠️ Error at row {idx}: {e}", "warning")
            st.warning(f"⚠️  Error at row {idx}: {e}")
            continue

    if realisasi_data:
        supabase.table('realisasi').insert(realisasi_data).execute()
        total_inserted += len(realisasi_data)
        add_log(f"✅ Final batch inserted to realisasi: {total_inserted:,} total records", "success")

    progress_bar.progress(100, "✅ Migration completed")
    progress_bar.empty()

    add_log(f"📊 REPLACE MODE Summary - Inserted: {total_inserted:,}, Skipped Kanwil: {skipped_kanwil:,}, Skipped Kancab: {skipped_kancab:,}", "success")
    st.success(f"""
    📊 Migration Summary (REPLACE MODE):
    - Total records inserted: **{total_inserted:,}**
    - Skipped (kanwil not found): **{skipped_kanwil:,}**
    - Skipped (kancab not found): **{skipped_kancab:,}**
    """)

    return total_inserted, skipped_kanwil, skipped_kancab


def fetch_with_retry_streamlit(supabase, last_id, limit, max_retries=3, retry_delay=2, retries=0):
    """
    Fetch data from RPC with automatic retry on timeout (Streamlit version).
    Implements exponential backoff retry strategy.
    """
    try:
        res = supabase.rpc(
            "get_realisasi_compare_not_exists_page",
            {"p_last_id": last_id, "p_limit": limit}
        ).execute()
        return res.data
    except Exception as e:
        if retries < max_retries:
            delay = retry_delay * (2 ** retries)  # Exponential backoff
            add_log(f"⚠️ Error pada last_id={last_id}, retry ke-{retries + 1}/{max_retries} - waiting {delay}s", "warning")
            st.warning(f"⚠️  Error pada last_id={last_id}, retry ke-{retries + 1}/{max_retries}")
            st.info(f"⏳ Waiting {delay} seconds before retry...")
            time.sleep(delay)
            return fetch_with_retry_streamlit(supabase, last_id, limit, max_retries, retry_delay, retries + 1)
        else:
            add_log(f"❌ Gagal setelah {max_retries} retry pada last_id={last_id}: {e}", "error")
            st.error(f"❌ Gagal setelah {max_retries} retry pada last_id={last_id}: {e}")
            raise e


def process_comparison_with_rpc_streamlit(supabase):
    """
    Process comparison using RPC get_realisasi_compare_not_exists_page (Streamlit version).
    Returns list of IDs that don't exist in realisasi table.
    """
    add_log("🔍 Step 2: Starting RPC comparison process...", "info")
    st.info("🔍 Step 2: Comparing data with database using RPC...")

    # Configuration
    max_retries = 3
    retry_delay = 2
    limit = 1000

    # Get total rows
    try:
        result = supabase.table("realisasi_compare").select("id", count="exact").limit(1).execute()
        total_rows = result.count
        add_log(f"📊 Total rows in realisasi_compare: {total_rows:,}", "info")
    except Exception as e:
        add_log(f"❌ Error getting total rows: {e}", "error")
        st.error(f"Error getting total rows: {e}")
        return []

    # Get min ID
    try:
        res = supabase.table("realisasi_compare").select("id").order("id", desc=False).limit(1).execute()
        min_id = res.data[0]["id"] if res.data else 0
        add_log(f"🔢 Starting from ID: {min_id}", "info")
    except Exception as e:
        add_log(f"❌ Error getting min id: {e}", "error")
        st.error(f"Error getting min id: {e}")
        return []

    if total_rows is None or total_rows == 0:
        add_log("📊 No data in realisasi_compare to process", "info")
        st.info("📊 No data in realisasi_compare to process")
        return []

    st.info(f"📊 Total rows in realisasi_compare: **{total_rows:,}**")
    st.info(f"🔢 Starting from ID: **{min_id}**")

    last_id = min_id
    processed_count = 0
    all_results = []

    progress_bar = st.progress(0, "Comparing data...")

    while True:
        try:
            # Call RPC with retry mechanism
            data = fetch_with_retry_streamlit(supabase, last_id, limit, max_retries, retry_delay)

            if data and len(data) > 0:
                all_results.extend(data)
                processed_count += len(data)
                last_id = max([row['realisasi_compare_id'] for row in data])

                progress = (processed_count / total_rows) * 100 if total_rows > 0 else 0
                add_log(f"✅ RPC batch: {len(data)} records found - Total unique: {len(all_results):,} ({progress:.1f}%)", "success")
                progress_bar.progress(min(int(progress) / 100, 0.99), f"Progress: {processed_count:,}/{total_rows:,} ({progress:.1f}%)")
            else:
                # Empty result
                if last_id >= (min_id + total_rows):
                    add_log(f"✅ Reached end of data at last_id={last_id}", "success")
                    st.info(f"✅ Reached end of data at last_id={last_id}")
                    break

                # Try advancing last_id
                last_id += limit
                add_log(f"🔍 Empty batch, advancing to last_id={last_id}", "info")

                # Safety check to avoid infinite loop
                if last_id > (min_id + total_rows + limit * 10):
                    add_log("⚠️ Exceeded maximum range, stopping iteration", "warning")
                    st.warning("⚠️  Exceeded maximum range, stopping iteration")
                    break

        except Exception as e:
            add_log(f"❌ Fatal error that couldn't be retried: {e}", "error")
            st.error(f"❌ Fatal error that couldn't be retried: {e}")
            with st.expander("🔍 Error Details"):
                st.code(str(e))
            break

        time.sleep(0.2)  # Rate limiting

    progress_bar.progress(100, "✅ Comparison completed")
    progress_bar.empty()

    add_log(f"📊 RPC Comparison Complete - Found {len(all_results):,} unique records to migrate", "success")
    st.success(f"✅ Found **{len(all_results):,}** unique records to migrate")

    return all_results


def migrate_from_compare_to_realisasi_streamlit(supabase, comparison_results):
    """
    Migrate data from realisasi_compare to realisasi based on comparison results (Streamlit version).
    Returns: migrated_count
    """
    if not comparison_results:
        add_log("ℹ️ No data to migrate", "info")
        st.info("ℹ️  No data to migrate")
        return 0

    add_log(f"📥 Step 3: Starting migration of {len(comparison_results):,} records to realisasi...", "info")
    st.info(f"📥 Step 3: Migrating {len(comparison_results):,} records to realisasi...")

    ids_to_migrate = [r['realisasi_compare_id'] for r in comparison_results]
    batch_size = 1000
    migrated_count = 0
    total_batches = (len(ids_to_migrate) + batch_size - 1) // batch_size

    progress_bar = st.progress(0, "Migrating data...")

    for i in range(0, len(ids_to_migrate), batch_size):
        batch_ids = ids_to_migrate[i:i + batch_size]
        batch_num = i // batch_size + 1

        try:
            # Fetch from realisasi_compare
            response = supabase.table("realisasi_compare").select("*").in_("id", batch_ids).execute()
            data_to_insert = response.data

            if not data_to_insert:
                continue

            # Remove 'id' column
            for row in data_to_insert:
                if 'id' in row:
                    del row['id']

            # Insert to realisasi
            supabase.table("realisasi").insert(data_to_insert).execute()
            migrated_count += len(data_to_insert)

            progress = (migrated_count / len(ids_to_migrate)) * 100
            add_log(f"✅ Migration batch {batch_num}/{total_batches}: {len(data_to_insert)} records - Total: {migrated_count:,}/{len(ids_to_migrate):,} ({progress:.1f}%)", "success")
            progress_bar.progress(int(progress) / 100, f"Migrated: {migrated_count:,}/{len(ids_to_migrate):,} ({progress:.1f}%)")

            time.sleep(0.2)

        except Exception as e:
            add_log(f"❌ Error on batch {batch_num}: {e}", "error")
            st.error(f"❌ Error on batch {batch_num}: {e}")
            continue

    progress_bar.progress(100, "✅ Migration completed")
    progress_bar.empty()

    add_log(f"📊 Migration Complete - Successfully migrated {migrated_count:,} records to realisasi", "success")
    st.success(f"✅ Successfully migrated **{migrated_count:,}** records to realisasi")

    return migrated_count


# ===== FUNGSI ALGORITMA UNTUK TARGET_KANWIL =====

def generate_target_kanwil_hash(record):
    """Generate SHA256 hash from target_kanwil record data for duplicate detection."""
    hash_data = {
        'kanwil_id': record.get('kanwil_id'),
        'target_setara_beras': record.get('target_setara_beras'),
    }
    json_string = json.dumps(hash_data, sort_keys=True, default=str)
    return hashlib.sha256(json_string.encode()).hexdigest()


def migrate_to_target_kanwil_compare_streamlit(supabase, df, kanwil_mapping):
    """
    Migrate data from DataFrame to target_kanwil_compare table (Streamlit version).
    Returns: total_inserted, skipped_kanwil
    """
    add_log("📥 Starting Migration to target_kanwil_compare...", "info")
    st.info("📥 Starting Migration to target_kanwil_compare...")

    # Clear table first
    add_log("🗑️ Clearing target_kanwil_compare table...", "warning")
    st.warning("🗑️  Clearing target_kanwil_compare table...")
    try:
        truncate_table_with_reset(supabase, "target_kanwil_compare")
        add_log("✅ target_kanwil_compare table cleared successfully", "success")
    except Exception as e:
        add_log(f"❌ Error clearing table: {e}", "error")
        st.error(f"❌ Error clearing table: {e}")

    # Migrate data
    add_log(f"📥 Migrating {len(df):,} records to target_kanwil_compare...", "info")
    st.info("📥 Migrating data to target_kanwil_compare...")
    target_kanwil_compare_data = []
    batch_size = 1000
    total_inserted = 0
    skipped_kanwil = 0

    progress_bar = st.progress(0, "Processing records...")

    for idx, row in df.iterrows():
        try:
            kanwil_name = str(row['kanwil']) if pd.notna(row['kanwil']) else None
            kanwil_id = kanwil_mapping.get(kanwil_name)

            if not kanwil_id:
                skipped_kanwil += 1
                continue

            # Use current date for database (but not for hash comparison)
            import datetime
            target_date = datetime.datetime.now().date()

            record = {
                'kanwil_id': kanwil_id,
                'target_setara_beras': str(row['Target Setara Beras']) if pd.notna(row.get('Target Setara Beras')) else None,
                'date': target_date.isoformat(),
            }

            record['row_hash'] = generate_target_kanwil_hash(record)
            target_kanwil_compare_data.append(record)

            if len(target_kanwil_compare_data) >= batch_size:
                supabase.table('target_kanwil_compare').insert(target_kanwil_compare_data).execute()
                total_inserted += len(target_kanwil_compare_data)
                progress = (idx + 1) / len(df) * 100
                add_log(f"✅ Batch inserted: {total_inserted:,} records ({progress:.1f}%)", "success")
                progress_bar.progress(int(progress) / 100, f"Inserted {total_inserted:,} records...")
                target_kanwil_compare_data = []

        except Exception as e:
            add_log(f"⚠️ Error at row {idx}: {e}", "warning")
            st.warning(f"⚠️  Error at row {idx}: {e}")
            continue

    if target_kanwil_compare_data:
        supabase.table('target_kanwil_compare').insert(target_kanwil_compare_data).execute()
        total_inserted += len(target_kanwil_compare_data)
        add_log(f"✅ Final batch inserted: {total_inserted:,} total records", "success")

    progress_bar.progress(100, "✅ Migration completed")
    progress_bar.empty()

    add_log(f"📊 Migration Summary - Inserted: {total_inserted:,}, Skipped Kanwil: {skipped_kanwil:,}", "success")
    st.success(f"""
    📊 Migration Summary:
    - Total records inserted: **{total_inserted:,}**
    - Skipped (kanwil not found): **{skipped_kanwil:,}**
    """)

    return total_inserted, skipped_kanwil


def migrate_to_target_kanwil_direct_streamlit(supabase, df, kanwil_mapping):
    """
    Migrate data from DataFrame directly to target_kanwil table (REPLACE MODE - Streamlit version).
    Returns: total_inserted, skipped_kanwil
    """
    add_log("📥 Starting Direct Migration to target_kanwil (REPLACE MODE)...", "warning")
    st.warning("📥 Starting Direct Migration to target_kanwil (REPLACE MODE)...")

    # Reset target_kanwil table
    add_log("🗑️ Resetting target_kanwil table and sequence...", "warning")
    st.info("🗑️  Resetting target_kanwil table...")
    if not truncate_table_with_reset(supabase, "target_kanwil"):
        add_log("❌ Failed to reset target_kanwil table. Aborting.", "error")
        st.error("❌ Failed to reset target_kanwil table. Aborting.")
        return 0, 0
    add_log("✅ target_kanwil table reset successfully", "success")

    # Migrate data
    add_log(f"📥 Migrating {len(df):,} records to target_kanwil...", "info")
    st.info("📥 Migrating data to target_kanwil...")
    target_kanwil_data = []
    batch_size = 1000
    total_inserted = 0
    skipped_kanwil = 0

    progress_bar = st.progress(0, "Processing records...")

    for idx, row in df.iterrows():
        try:
            kanwil_name = str(row['kanwil']) if pd.notna(row['kanwil']) else None
            kanwil_id = kanwil_mapping.get(kanwil_name)

            if not kanwil_id:
                skipped_kanwil += 1
                continue

            # Use current date for database
            import datetime
            target_date = datetime.datetime.now().date()

            record = {
                'kanwil_id': kanwil_id,
                'target_setara_beras': str(row['Target Setara Beras']) if pd.notna(row.get('Target Setara Beras')) else None,
                'date': target_date.isoformat(),
            }

            target_kanwil_data.append(record)

            if len(target_kanwil_data) >= batch_size:
                supabase.table('target_kanwil').insert(target_kanwil_data).execute()
                total_inserted += len(target_kanwil_data)
                progress = (idx + 1) / len(df) * 100
                add_log(f"✅ Batch inserted to target_kanwil: {total_inserted:,} records ({progress:.1f}%)", "success")
                progress_bar.progress(int(progress) / 100, f"Inserted {total_inserted:,} records...")
                target_kanwil_data = []

        except Exception as e:
            add_log(f"⚠️ Error at row {idx}: {e}", "warning")
            st.warning(f"⚠️  Error at row {idx}: {e}")
            continue

    if target_kanwil_data:
        supabase.table('target_kanwil').insert(target_kanwil_data).execute()
        total_inserted += len(target_kanwil_data)
        add_log(f"✅ Final batch inserted to target_kanwil: {total_inserted:,} total records", "success")

    progress_bar.progress(100, "✅ Migration completed")
    progress_bar.empty()

    add_log(f"📊 REPLACE MODE Summary - Inserted: {total_inserted:,}, Skipped Kanwil: {skipped_kanwil:,}", "success")
    st.success(f"""
    📊 Migration Summary (REPLACE MODE):
    - Total records inserted: **{total_inserted:,}**
    - Skipped (kanwil not found): **{skipped_kanwil:,}**
    """)

    return total_inserted, skipped_kanwil


# ===== FUNGSI ALGORITMA UNTUK TARGET_KANCAB =====

def generate_target_kancab_hash(record):
    """Generate SHA256 hash from target_kancab record data for duplicate detection."""
    hash_data = {
        'kancab_id': record.get('kancab_id'),
        'target_setara_beras': record.get('target_setara_beras'),
    }
    json_string = json.dumps(hash_data, sort_keys=True, default=str)
    return hashlib.sha256(json_string.encode()).hexdigest()


def migrate_to_target_kancab_compare_streamlit(supabase, df, kancab_mapping):
    """
    Migrate data from DataFrame to target_kancab_compare table (Streamlit version).
    Returns: total_inserted, skipped_kancab
    """
    add_log("📥 Starting Migration to target_kancab_compare...", "info")
    st.info("📥 Starting Migration to target_kancab_compare...")

    # Clear table first
    add_log("🗑️ Clearing target_kancab_compare table...", "warning")
    st.warning("🗑️  Clearing target_kancab_compare table...")
    try:
        truncate_table_with_reset(supabase, "target_kancab_compare")
        add_log("✅ target_kancab_compare table cleared successfully", "success")
    except Exception as e:
        add_log(f"❌ Error clearing table: {e}", "error")
        st.error(f"❌ Error clearing table: {e}")

    # Migrate data
    add_log(f"📥 Migrating {len(df):,} records to target_kancab_compare...", "info")
    st.info("📥 Migrating data to target_kancab_compare...")
    target_kancab_compare_data = []
    batch_size = 1000
    total_inserted = 0
    skipped_kancab = 0

    progress_bar = st.progress(0, "Processing records...")

    for idx, row in df.iterrows():
        try:
            kancab_name = str(row['kancab']) if pd.notna(row['kancab']) else None
            kancab_id = kancab_mapping.get(kancab_name)

            if not kancab_id:
                skipped_kancab += 1
                continue

            # Use current date for database (but not for hash comparison)
            import datetime
            target_date = datetime.datetime.now().date()

            record = {
                'kancab_id': kancab_id,
                'target_setara_beras': str(row['Target Setara Beras']) if pd.notna(row.get('Target Setara Beras')) else None,
                'date': target_date.isoformat(),
            }

            record['row_hash'] = generate_target_kancab_hash(record)
            target_kancab_compare_data.append(record)

            if len(target_kancab_compare_data) >= batch_size:
                supabase.table('target_kancab_compare').insert(target_kancab_compare_data).execute()
                total_inserted += len(target_kancab_compare_data)
                progress = (idx + 1) / len(df) * 100
                add_log(f"✅ Batch inserted: {total_inserted:,} records ({progress:.1f}%)", "success")
                progress_bar.progress(int(progress) / 100, f"Inserted {total_inserted:,} records...")
                target_kancab_compare_data = []

        except Exception as e:
            add_log(f"⚠️ Error at row {idx}: {e}", "warning")
            st.warning(f"⚠️  Error at row {idx}: {e}")
            continue

    if target_kancab_compare_data:
        supabase.table('target_kancab_compare').insert(target_kancab_compare_data).execute()
        total_inserted += len(target_kancab_compare_data)
        add_log(f"✅ Final batch inserted: {total_inserted:,} total records", "success")

    progress_bar.progress(100, "✅ Migration completed")
    progress_bar.empty()

    add_log(f"📊 Migration Summary - Inserted: {total_inserted:,}, Skipped Kancab: {skipped_kancab:,}", "success")
    st.success(f"""
    📊 Migration Summary:
    - Total records inserted: **{total_inserted:,}**
    - Skipped (kancab not found): **{skipped_kancab:,}**
    """)

    return total_inserted, skipped_kancab


def migrate_to_target_kancab_direct_streamlit(supabase, df, kancab_mapping):
    """
    Migrate data from DataFrame directly to target_kancab table (REPLACE MODE - Streamlit version).
    Returns: total_inserted, skipped_kancab
    """
    add_log("📥 Starting Direct Migration to target_kancab (REPLACE MODE)...", "warning")
    st.warning("📥 Starting Direct Migration to target_kancab (REPLACE MODE)...")

    # Reset target_kancab table
    add_log("🗑️ Resetting target_kancab table and sequence...", "warning")
    st.info("🗑️  Resetting target_kancab table...")
    if not truncate_table_with_reset(supabase, "target_kancab"):
        add_log("❌ Failed to reset target_kancab table. Aborting.", "error")
        st.error("❌ Failed to reset target_kancab table. Aborting.")
        return 0, 0
    add_log("✅ target_kancab table reset successfully", "success")

    # Migrate data
    add_log(f"📥 Migrating {len(df):,} records to target_kancab...", "info")
    st.info("📥 Migrating data to target_kancab...")
    target_kancab_data = []
    batch_size = 1000
    total_inserted = 0
    skipped_kancab = 0

    progress_bar = st.progress(0, "Processing records...")

    for idx, row in df.iterrows():
        try:
            kancab_name = str(row['kancab']) if pd.notna(row['kancab']) else None
            kancab_id = kancab_mapping.get(kancab_name)

            if not kancab_id:
                skipped_kancab += 1
                continue

            # Use current date for database
            import datetime
            target_date = datetime.datetime.now().date()

            record = {
                'kancab_id': kancab_id,
                'target_setara_beras': str(row['Target Setara Beras']) if pd.notna(row.get('Target Setara Beras')) else None,
                'date': target_date.isoformat(),
            }

            target_kancab_data.append(record)

            if len(target_kancab_data) >= batch_size:
                supabase.table('target_kancab').insert(target_kancab_data).execute()
                total_inserted += len(target_kancab_data)
                progress = (idx + 1) / len(df) * 100
                add_log(f"✅ Batch inserted to target_kancab: {total_inserted:,} records ({progress:.1f}%)", "success")
                progress_bar.progress(int(progress) / 100, f"Inserted {total_inserted:,} records...")
                target_kancab_data = []

        except Exception as e:
            add_log(f"⚠️ Error at row {idx}: {e}", "warning")
            st.warning(f"⚠️  Error at row {idx}: {e}")
            continue

    if target_kancab_data:
        supabase.table('target_kancab').insert(target_kancab_data).execute()
        total_inserted += len(target_kancab_data)
        add_log(f"✅ Final batch inserted to target_kancab: {total_inserted:,} total records", "success")

    progress_bar.progress(100, "✅ Migration completed")
    progress_bar.empty()

    add_log(f"📊 REPLACE MODE Summary - Inserted: {total_inserted:,}, Skipped Kancab: {skipped_kancab:,}", "success")
    st.success(f"""
    📊 Migration Summary (REPLACE MODE):
    - Total records inserted: **{total_inserted:,}**
    - Skipped (kancab not found): **{skipped_kancab:,}**
    """)

    return total_inserted, skipped_kancab


# ===== RPC COMPARISON FUNCTIONS FOR TARGET_KANWIL =====

def fetch_with_retry_target_kanwil_streamlit(supabase, last_id, limit, max_retries=3, retry_delay=2, retries=0):
    """
    Fetch data from RPC with automatic retry on timeout (target_kanwil version).
    Implements exponential backoff retry strategy.
    """
    try:
        res = supabase.rpc(
            "get_target_kanwil_compare_not_exists_page",
            {"p_last_id": last_id, "p_limit": limit}
        ).execute()
        return res.data
    except Exception as e:
        if retries < max_retries:
            delay = retry_delay * (2 ** retries)  # Exponential backoff
            add_log(f"⚠️ Error pada last_id={last_id}, retry ke-{retries + 1}/{max_retries} - waiting {delay}s", "warning")
            st.warning(f"⚠️  Error pada last_id={last_id}, retry ke-{retries + 1}/{max_retries}")
            st.info(f"⏳ Waiting {delay} seconds before retry...")
            time.sleep(delay)
            return fetch_with_retry_target_kanwil_streamlit(supabase, last_id, limit, max_retries, retry_delay, retries + 1)
        else:
            add_log(f"❌ Gagal setelah {max_retries} retry pada last_id={last_id}: {e}", "error")
            st.error(f"❌ Gagal setelah {max_retries} retry pada last_id={last_id}: {e}")
            raise e


def process_comparison_target_kanwil_with_rpc_streamlit(supabase):
    """
    Process comparison using RPC get_target_kanwil_compare_not_exists_page (Streamlit version).
    Returns list of IDs that don't exist in target_kanwil table.
    """
    add_log("🔍 Step 2: Starting RPC comparison process for target_kanwil...", "info")
    st.info("🔍 Step 2: Comparing data with database using RPC...")

    # Configuration
    max_retries = 3
    retry_delay = 2
    limit = 1000

    # Get total rows
    try:
        result = supabase.table("target_kanwil_compare").select("id", count="exact").limit(1).execute()
        total_rows = result.count
        add_log(f"📊 Total rows in target_kanwil_compare: {total_rows:,}", "info")
    except Exception as e:
        add_log(f"❌ Error getting total rows: {e}", "error")
        st.error(f"Error getting total rows: {e}")
        return []

    # Get min ID
    try:
        res = supabase.table("target_kanwil_compare").select("id").order("id", desc=False).limit(1).execute()
        min_id = res.data[0]["id"] if res.data else 0
        add_log(f"🔢 Starting from ID: {min_id}", "info")
    except Exception as e:
        add_log(f"❌ Error getting min id: {e}", "error")
        st.error(f"Error getting min id: {e}")
        return []

    if total_rows is None or total_rows == 0:
        add_log("📊 No data in target_kanwil_compare to process", "info")
        st.info("📊 No data in target_kanwil_compare to process")
        return []

    st.info(f"📊 Total rows in target_kanwil_compare: **{total_rows:,}**")
    st.info(f"🔢 Starting from ID: **{min_id}**")

    last_id = min_id - 1
    processed_count = 0
    all_results = []

    progress_bar = st.progress(0, "Comparing data...")

    while True:
        try:
            # Call RPC with retry mechanism
            data = fetch_with_retry_target_kanwil_streamlit(supabase, last_id, limit, max_retries, retry_delay)

            if data and len(data) > 0:
                all_results.extend(data)
                processed_count += len(data)
                last_id = max([row['target_kanwil_compare_id'] for row in data])

                progress = (processed_count / total_rows) * 100 if total_rows > 0 else 0
                add_log(f"✅ RPC batch: {len(data)} records found - Total unique: {len(all_results):,} ({progress:.1f}%)", "success")
                progress_bar.progress(min(int(progress) / 100, 0.99), f"Progress: {processed_count:,}/{total_rows:,} ({progress:.1f}%)")
            else:
                # Empty result - finished
                add_log(f"✅ Reached end of data at last_id={last_id}", "success")
                break

        except Exception as e:
            add_log(f"❌ Fatal error that couldn't be retried: {e}", "error")
            st.error(f"❌ Fatal error that couldn't be retried: {e}")
            with st.expander("🔍 Error Details"):
                st.code(str(e))
            break

        time.sleep(0.2)  # Rate limiting

    progress_bar.progress(100, "✅ Comparison completed")
    progress_bar.empty()

    add_log(f"📊 RPC Comparison Complete - Found {len(all_results):,} unique records to migrate", "success")
    st.success(f"✅ Found **{len(all_results):,}** unique records to migrate")

    return all_results


def migrate_from_target_kanwil_compare_to_target_kanwil_streamlit(supabase, comparison_results):
    """
    Migrate data from target_kanwil_compare to target_kanwil based on comparison results (Streamlit version).
    Returns: migrated_count
    """
    if not comparison_results:
        add_log("ℹ️ No data to migrate", "info")
        st.info("ℹ️  No data to migrate")
        return 0

    add_log(f"📥 Step 3: Starting migration of {len(comparison_results):,} records to target_kanwil...", "info")
    st.info(f"📥 Step 3: Migrating {len(comparison_results):,} records to target_kanwil...")

    ids_to_migrate = [r['target_kanwil_compare_id'] for r in comparison_results]
    batch_size = 1000
    migrated_count = 0
    total_batches = (len(ids_to_migrate) + batch_size - 1) // batch_size

    progress_bar = st.progress(0, "Migrating data...")

    for i in range(0, len(ids_to_migrate), batch_size):
        batch_ids = ids_to_migrate[i:i + batch_size]
        batch_num = i // batch_size + 1

        try:
            # Fetch from target_kanwil_compare
            response = supabase.table("target_kanwil_compare").select("*").in_("id", batch_ids).execute()
            data_to_insert = response.data

            if not data_to_insert:
                continue

            # Remove 'id', 'created_at', 'row_hash' columns
            for row in data_to_insert:
                if 'id' in row:
                    del row['id']
                if 'created_at' in row:
                    del row['created_at']
                if 'row_hash' in row:
                    del row['row_hash']

            # Insert to target_kanwil
            supabase.table("target_kanwil").insert(data_to_insert).execute()
            migrated_count += len(data_to_insert)

            progress = (migrated_count / len(ids_to_migrate)) * 100
            add_log(f"✅ Migration batch {batch_num}/{total_batches}: {len(data_to_insert)} records - Total: {migrated_count:,}/{len(ids_to_migrate):,} ({progress:.1f}%)", "success")
            progress_bar.progress(int(progress) / 100, f"Migrated: {migrated_count:,}/{len(ids_to_migrate):,} ({progress:.1f}%)")

            time.sleep(0.2)

        except Exception as e:
            add_log(f"❌ Error on batch {batch_num}: {e}", "error")
            st.error(f"❌ Error on batch {batch_num}: {e}")
            continue

    progress_bar.progress(100, "✅ Migration completed")
    progress_bar.empty()

    add_log(f"📊 Migration Complete - Successfully migrated {migrated_count:,} records to target_kanwil", "success")
    st.success(f"✅ Successfully migrated **{migrated_count:,}** records to target_kanwil")

    return migrated_count


# ===== RPC COMPARISON FUNCTIONS FOR TARGET_KANCAB =====

def fetch_with_retry_target_kancab_streamlit(supabase, last_id, limit, max_retries=3, retry_delay=2, retries=0):
    """
    Fetch data from RPC with automatic retry on timeout (target_kancab version).
    Implements exponential backoff retry strategy.
    """
    try:
        res = supabase.rpc(
            "get_target_kancab_compare_not_exists_page",
            {"p_last_id": last_id, "p_limit": limit}
        ).execute()
        return res.data
    except Exception as e:
        if retries < max_retries:
            delay = retry_delay * (2 ** retries)  # Exponential backoff
            add_log(f"⚠️ Error pada last_id={last_id}, retry ke-{retries + 1}/{max_retries} - waiting {delay}s", "warning")
            st.warning(f"⚠️  Error pada last_id={last_id}, retry ke-{retries + 1}/{max_retries}")
            st.info(f"⏳ Waiting {delay} seconds before retry...")
            time.sleep(delay)
            return fetch_with_retry_target_kancab_streamlit(supabase, last_id, limit, max_retries, retry_delay, retries + 1)
        else:
            add_log(f"❌ Gagal setelah {max_retries} retry pada last_id={last_id}: {e}", "error")
            st.error(f"❌ Gagal setelah {max_retries} retry pada last_id={last_id}: {e}")
            raise e


def process_comparison_target_kancab_with_rpc_streamlit(supabase):
    """
    Process comparison using RPC get_target_kancab_compare_not_exists_page (Streamlit version).
    Returns list of IDs that don't exist in target_kancab table.
    """
    add_log("🔍 Step 2: Starting RPC comparison process for target_kancab...", "info")
    st.info("🔍 Step 2: Comparing data with database using RPC...")

    # Configuration
    max_retries = 3
    retry_delay = 2
    limit = 1000

    # Get total rows
    try:
        result = supabase.table("target_kancab_compare").select("id", count="exact").limit(1).execute()
        total_rows = result.count
        add_log(f"📊 Total rows in target_kancab_compare: {total_rows:,}", "info")
    except Exception as e:
        add_log(f"❌ Error getting total rows: {e}", "error")
        st.error(f"Error getting total rows: {e}")
        return []

    # Get min ID
    try:
        res = supabase.table("target_kancab_compare").select("id").order("id", desc=False).limit(1).execute()
        min_id = res.data[0]["id"] if res.data else 0
        add_log(f"🔢 Starting from ID: {min_id}", "info")
    except Exception as e:
        add_log(f"❌ Error getting min id: {e}", "error")
        st.error(f"Error getting min id: {e}")
        return []

    if total_rows is None or total_rows == 0:
        add_log("📊 No data in target_kancab_compare to process", "info")
        st.info("📊 No data in target_kancab_compare to process")
        return []

    st.info(f"📊 Total rows in target_kancab_compare: **{total_rows:,}**")
    st.info(f"🔢 Starting from ID: **{min_id}**")

    last_id = min_id - 1
    processed_count = 0
    all_results = []

    progress_bar = st.progress(0, "Comparing data...")

    while True:
        try:
            # Call RPC with retry mechanism
            data = fetch_with_retry_target_kancab_streamlit(supabase, last_id, limit, max_retries, retry_delay)

            if data and len(data) > 0:
                all_results.extend(data)
                processed_count += len(data)
                last_id = max([row['target_kancab_compare_id'] for row in data])

                progress = (processed_count / total_rows) * 100 if total_rows > 0 else 0
                add_log(f"✅ RPC batch: {len(data)} records found - Total unique: {len(all_results):,} ({progress:.1f}%)", "success")
                progress_bar.progress(min(int(progress) / 100, 0.99), f"Progress: {processed_count:,}/{total_rows:,} ({progress:.1f}%)")
            else:
                # Empty result - finished
                add_log(f"✅ Reached end of data at last_id={last_id}", "success")
                break

        except Exception as e:
            add_log(f"❌ Fatal error that couldn't be retried: {e}", "error")
            st.error(f"❌ Fatal error that couldn't be retried: {e}")
            with st.expander("🔍 Error Details"):
                st.code(str(e))
            break

        time.sleep(0.2)  # Rate limiting

    progress_bar.progress(100, "✅ Comparison completed")
    progress_bar.empty()

    add_log(f"📊 RPC Comparison Complete - Found {len(all_results):,} unique records to migrate", "success")
    st.success(f"✅ Found **{len(all_results):,}** unique records to migrate")

    return all_results


def migrate_from_target_kancab_compare_to_target_kancab_streamlit(supabase, comparison_results):
    """
    Migrate data from target_kancab_compare to target_kancab based on comparison results (Streamlit version).
    Returns: migrated_count
    """
    if not comparison_results:
        add_log("ℹ️ No data to migrate", "info")
        st.info("ℹ️  No data to migrate")
        return 0

    add_log(f"📥 Step 3: Starting migration of {len(comparison_results):,} records to target_kancab...", "info")
    st.info(f"📥 Step 3: Migrating {len(comparison_results):,} records to target_kancab...")

    ids_to_migrate = [r['target_kancab_compare_id'] for r in comparison_results]
    batch_size = 1000
    migrated_count = 0
    total_batches = (len(ids_to_migrate) + batch_size - 1) // batch_size

    progress_bar = st.progress(0, "Migrating data...")

    for i in range(0, len(ids_to_migrate), batch_size):
        batch_ids = ids_to_migrate[i:i + batch_size]
        batch_num = i // batch_size + 1

        try:
            # Fetch from target_kancab_compare
            response = supabase.table("target_kancab_compare").select("*").in_("id", batch_ids).execute()
            data_to_insert = response.data

            if not data_to_insert:
                continue

            # Remove 'id', 'created_at', 'row_hash' columns
            for row in data_to_insert:
                if 'id' in row:
                    del row['id']
                if 'created_at' in row:
                    del row['created_at']
                if 'row_hash' in row:
                    del row['row_hash']

            # Insert to target_kancab
            supabase.table("target_kancab").insert(data_to_insert).execute()
            migrated_count += len(data_to_insert)

            progress = (migrated_count / len(ids_to_migrate)) * 100
            add_log(f"✅ Migration batch {batch_num}/{total_batches}: {len(data_to_insert)} records - Total: {migrated_count:,}/{len(ids_to_migrate):,} ({progress:.1f}%)", "success")
            progress_bar.progress(int(progress) / 100, f"Migrated: {migrated_count:,}/{len(ids_to_migrate):,} ({progress:.1f}%)")

            time.sleep(0.2)

        except Exception as e:
            add_log(f"❌ Error on batch {batch_num}: {e}", "error")
            st.error(f"❌ Error on batch {batch_num}: {e}")
            continue

    progress_bar.progress(100, "✅ Migration completed")
    progress_bar.empty()

    add_log(f"📊 Migration Complete - Successfully migrated {migrated_count:,} records to target_kancab", "success")
    st.success(f"✅ Successfully migrated **{migrated_count:,}** records to target_kancab")

    return migrated_count


# ===== FUNGSI RPC SUPABASE =====

# Cache helper using session_state
def get_cached_data(cache_key, fetch_function, *args, ttl_seconds=300, **kwargs):
    """
    Helper function untuk caching data menggunakan session_state

    Parameters:
    - cache_key: Unique key untuk cache
    - fetch_function: Function untuk fetch data jika cache miss
    - ttl_seconds: Time to live dalam detik (default 5 menit)
    - *args, **kwargs: Arguments untuk fetch_function

    Returns:
    Data dari cache atau hasil fetch_function
    """
    # Initialize cache dictionary jika belum ada
    if 'data_cache' not in st.session_state:
        st.session_state.data_cache = {}

    current_time = time.time()

    # Check if cache exists and is still valid
    if cache_key in st.session_state.data_cache:
        cached_item = st.session_state.data_cache[cache_key]
        if current_time - cached_item['timestamp'] < ttl_seconds:
            return cached_item['data']

    # Cache miss or expired - fetch new data
    data = fetch_function(*args, **kwargs)

    # Store in cache
    st.session_state.data_cache[cache_key] = {
        'data': data,
        'timestamp': current_time
    }

    return data

def handle_rpc_error(e, function_name):
    """
    Menangani error dari RPC call dengan pesan yang sesuai

    Parameters:
    - e: Exception object
    - function_name: Nama fungsi yang error
    """
    error_str = str(e)

    # Cek apakah ini timeout error
    if 'statement timeout' in error_str.lower() or '57014' in error_str:
        st.markdown('<p style="color:#B8860B; font-size:16px; font-weight:bold;">⏱️ Request Timeout</p>', unsafe_allow_html=True)
        st.markdown('<p style="color:#B8860B; font-size:14px;">⚠️ Database membutuhkan waktu terlalu lama untuk memproses data. Refresh halaman</p>', unsafe_allow_html=True)

        # Tampilkan detail error untuk debugging
        with st.expander("🔍 Detail Error (untuk debugging)"):
            st.code(error_str)

        # Tambahkan tombol refresh
        col1, col2 = st.columns(2)
        with col1:
            if st.button("🔄 Refresh Halaman", key=f"refresh_{function_name}", use_container_width=True):
                st.rerun()
        with col2:
            if st.button("⬅️ Kembali ke Dashboard", key=f"back_{function_name}", use_container_width=True):
                st.session_state.clear()
                st.rerun()
    else:
        st.markdown(f'<p style="color:#B8860B; font-size:16px; font-weight:bold;">❌ Error pada {function_name}</p>', unsafe_allow_html=True)
        st.code(error_str)

        if st.button("🔄 Coba Lagi", key=f"retry_{function_name}"):
            st.rerun()

def get_metric_card_data(p_nama_kanwil, p_akun_analitik, p_start_date, p_end_date, p_today):
    """
    Mengambil data untuk metric card menggunakan RPC get_realisasi_setara_beras

    Parameters:
    - p_nama_kanwil: Value filter Kanwil (list atau None untuk semua)
    - p_akun_analitik: Value filter akun analitik
    - p_start_date: value start date filter periode
    - p_end_date: value end date filter periode
    - p_today: value end date filter periode (untuk hari ini)

    Returns:
    Dict dengan keys: total_setara_beras_rentang, total_setara_beras_hari_ini, target_setara_beras, sisa_target
    """
    try:
        response = supabase.rpc(
            "get_realisasi_setara_beras",
            {
                "p_nama_kanwil": p_nama_kanwil,
                "p_akun_analitik": p_akun_analitik,
                "p_start_date": p_start_date,
                "p_end_date": p_end_date,
                "p_today": p_today
            }
        ).execute()

        if response.data and len(response.data) > 0:
            return response.data[0]
        else:
            return {
                'total_setara_beras_rentang': 0.0,
                'total_setara_beras_hari_ini': 0.0,
                'target_setara_beras': 0.0,
                'sisa_target': 0.0
            }
    except Exception as e:
        handle_rpc_error(e, "get_metric_card_data")
        return {
            'total_setara_beras_rentang': 0.0,
            'total_setara_beras_hari_ini': 0.0,
            'target_setara_beras': 0.0,
            'sisa_target': 0.0
        }

def get_tabel_realisasi_kanwil(p_akun_analitik, p_start_date, p_end_date):
    """
    Mengambil data untuk Tabel Realisasi per-Kanwil menggunakan RPC get_overview_setara_beras_all_kanwil

    Parameters:
    - p_akun_analitik: Value filter akun analitik
    - p_start_date: value start date filter periode
    - p_end_date: value end date filter periode

    Returns:
    DataFrame dengan kolom: kanwil, target_setara_beras, beras, gkg, gkp, setara_beras, capaian_persen
    """
    try:
        response = supabase.rpc(
            "get_overview_setara_beras_all_kanwil",
            {
                "p_akun_analitik": p_akun_analitik,
                "p_start_date": p_start_date,
                "p_end_date": p_end_date
            }
        ).execute()

        df = pd.DataFrame(response.data)
        if not df.empty:
            df["capaian_persen"] = df["capaian_persen"].round(1)
        return df
    except Exception as e:
        handle_rpc_error(e, "get_tabel_realisasi_kanwil")
        return pd.DataFrame()

def get_tabel_realisasi_kancab(p_nama_kanwil, p_akun_analitik, p_start_date, p_end_date):
    """
    Mengambil data untuk Tabel Realisasi per-Kancab menggunakan RPC get_overview_setara_beras_kancab

    Parameters:
    - p_nama_kanwil: Value filter Kanwil
    - p_akun_analitik: Value filter akun analitik
    - p_start_date: value start date filter periode
    - p_end_date: value end date filter periode

    Returns:
    DataFrame dengan kolom: kancab, target_setara_beras, beras_a, gkg_b, gkp_c, setara_beras_d, capaian_persen
    """
    try:
        # Debug: print parameter yang dikirim
        print(f"DEBUG get_tabel_realisasi_kancab - p_nama_kanwil: {p_nama_kanwil}")
        print(f"DEBUG get_tabel_realisasi_kancab - p_akun_analitik: {p_akun_analitik}")

        response = supabase.rpc(
            "get_overview_setara_beras_kancab",
            {
                "p_nama_kanwil": p_nama_kanwil,
                "p_akun_analitik": p_akun_analitik,
                "p_start_date": p_start_date,
                "p_end_date": p_end_date
            }
        ).execute()

        df = pd.DataFrame(response.data)
        print(f"DEBUG get_tabel_realisasi_kancab - Rows returned: {len(df)}")
        if not df.empty:
            print(f"DEBUG get_tabel_realisasi_kancab - Sample kancab: {df['kancab'].head(3).tolist() if 'kancab' in df.columns else 'No kancab column'}")
            # Print all columns to see what data structure we're getting
            print(f"DEBUG get_tabel_realisasi_kancab - Columns: {df.columns.tolist()}")
            # If there's a kanwil-related column, print it to verify filtering
            if 'nama_kanwil' in df.columns:
                print(f"DEBUG get_tabel_realisasi_kancab - Unique kanwil in result: {df['nama_kanwil'].unique().tolist()}")

        return df
    except Exception as e:
        handle_rpc_error(e, "get_tabel_realisasi_kancab")
        return pd.DataFrame()

def get_tren_realisasi_kanwil(p_nama_kanwil, p_akun_analitik, p_start_date, p_end_date):
    """
    Mengambil data untuk Tren Realisasi Kanwil menggunakan RPC get_realisasi_harian_setara_beras

    Parameters:
    - p_nama_kanwil: Value filter Kanwil
    - p_akun_analitik: Value filter akun analitik
    - p_start_date: value start date filter periode
    - p_end_date: value end date filter periode

    Returns:
    DataFrame dengan kolom: tanggal, nama_kanwil, beras, gkg, gkp, setara_beras
    """
    try:
        response = supabase.rpc(
            "get_realisasi_harian_setara_beras",
            {
                "p_nama_kanwil": p_nama_kanwil,
                "p_akun_analitik": p_akun_analitik,
                "p_start_date": p_start_date,
                "p_end_date": p_end_date
            }
        ).execute()

        df = pd.DataFrame(response.data)
        return df
    except Exception as e:
        handle_rpc_error(e, "get_tren_realisasi_kanwil")
        return pd.DataFrame()

def get_realisasi_7_hari_terakhir(p_nama_kanwil, p_akun_analitik, p_end_date):
    """
    Mengambil data untuk Realisasi 7 Hari Terakhir
    Menggunakan RPC get_realisasi_harian_setara_beras dengan periode 7 hari ke belakang

    Parameters:
    - p_nama_kanwil: Value filter Kanwil
    - p_akun_analitik: Value filter akun analitik
    - p_end_date: value end date filter periode

    Returns:
    DataFrame dengan kolom: tanggal, nama_kanwil, beras, gkg, gkp, setara_beras
    """
    try:
        # Hitung 7 hari ke belakang
        if isinstance(p_end_date, str):
            end_date_obj = datetime.strptime(p_end_date, '%Y-%m-%d').date()
        else:
            end_date_obj = p_end_date

        start_date_7days = end_date_obj - timedelta(days=6)  # 7 hari termasuk hari ini

        response = supabase.rpc(
            "get_realisasi_harian_setara_beras",
            {
                "p_nama_kanwil": p_nama_kanwil,
                "p_akun_analitik": p_akun_analitik,
                "p_start_date": start_date_7days.strftime('%Y-%m-%d'),
                "p_end_date": p_end_date if isinstance(p_end_date, str) else p_end_date.strftime('%Y-%m-%d')
            }
        ).execute()

        df = pd.DataFrame(response.data)
        return df
    except Exception as e:
        handle_rpc_error(e, "get_realisasi_7_hari_terakhir")
        return pd.DataFrame()

def create_summary_table_from_rpc(p_akun_analitik, p_start_date, p_end_date):
    """
    Create summary table with Kanwil Sentra Produksi and Kanwil Lainnya
    Menggunakan data dari RPC get_overview_setara_beras_all_kanwil
    Shows: No, Kanwil, Target Setara Beras, Realisasi (Beras, GKG, GKP, Setara Beras, Capaian %)
    """
    # List kanwil sentra produksi
    kanwil_sentra = [
        '08001 - KANTOR WILAYAH LAMPUNG',
        '21001 - KANTOR WILAYAH SULSEL SULBAR',
        '20001 - KANTOR WILAYAH SULTRA',
        '12001 - KANTOR WILAYAH DI YOGYAKARTA',
        '09001 - KANTOR WILAYAH DKI JAKARTA BANTEN',
        '23001 - KANTOR WILAYAH N.T.B',
        '13001 - KANTOR WILAYAH JATIM',
        '10001 - KANTOR WILAYAH JABAR',
        '01001 - KANTOR WILAYAH ACEH',
        '06001 - KANTOR WILAYAH SUMSEL',
        '11001 - KANTOR WILAYAH JATENG'
    ]

    # List kanwil lainnya
    kanwil_lainnya = [
        '15001 - KANTOR WILAYAH KALTIM KALTARA',
        '25001 - KANTOR WILAYAH MALUKU MALUT',
        '26001 - KANTOR WILAYAH PAPUA PABAR',
        '02001 - KANTOR WILAYAH SUMUT',
        '04001 - KANTOR WILAYAH SUMBAR',
        '17001 - KANTOR WILAYAH KALTENG',
        '16001 - KANTOR WILAYAH KALSEL',
        '14001 - KANTOR WILAYAH KALBAR',
        '18001 - KANTOR WILAYAH SULUT GORONTALO',
        '05001 - KANTOR WILAYAH JAMBI',
        '19001 - KANTOR WILAYAH SULTENG',
        '24001 - KANTOR WILAYAH N.T.T',
        '03001 - KANTOR WILAYAH RIAU DAN KEPRI',
        '22001 - KANTOR WILAYAH BALI',
        '07001 - KANTOR WILAYAH BENGKULU'
    ]

    # Ambil data dari RPC
    df_all_kanwil = get_tabel_realisasi_kanwil(p_akun_analitik, p_start_date, p_end_date)

    if df_all_kanwil.empty:
        return [], []

    # Build data for Kanwil Sentra Produksi
    data_sentra = []
    for kanwil in kanwil_sentra:
        # Cari data kanwil di DataFrame
        # Format di database: "16 - 15001 - KANTOR WILAYAH KALTIM KALTARA"
        # Format di aplikasi: "15001 - KANTOR WILAYAH KALTIM KALTARA"
        # Kita cari dengan menggunakan kode kanwil (misal: "08001")
        kode_kanwil = kanwil.split(' - ')[0]  # Ambil kode seperti "08001"
        kanwil_data = df_all_kanwil[df_all_kanwil['kanwil'].str.contains(f" {kode_kanwil} ", na=False)]

        if not kanwil_data.empty:
            row = kanwil_data.iloc[0]
            data_sentra.append({
                'Kanwil': kanwil,
                'Target Setara Beras': row['target_setara_beras'],
                'Beras (a)': row['beras'],
                'GKG (b)': row['gkg'],
                'GKP (c)': row['gkp'],
                'Setara Beras (d)': row['setara_beras'],
                'Capaian (%)': row['capaian_persen']
            })
        else:
            # Jika tidak ada data, buat row dengan nilai 0
            data_sentra.append({
                'Kanwil': kanwil,
                'Target Setara Beras': 0,
                'Beras (a)': 0,
                'GKG (b)': 0,
                'GKP (c)': 0,
                'Setara Beras (d)': 0,
                'Capaian (%)': 0
            })

    # Sort by Capaian (%) descending
    data_sentra = sorted(data_sentra, key=lambda x: x['Capaian (%)'], reverse=True)
    # Re-assign No after sorting
    for idx, row in enumerate(data_sentra, 1):
        row['No'] = idx

    # Build data for Kanwil Lainnya
    data_lainnya = []
    for kanwil in kanwil_lainnya:
        # Cari data kanwil di DataFrame dengan kode kanwil
        kode_kanwil = kanwil.split(' - ')[0]  # Ambil kode seperti "15001"
        kanwil_data = df_all_kanwil[df_all_kanwil['kanwil'].str.contains(f" {kode_kanwil} ", na=False)]

        if not kanwil_data.empty:
            row = kanwil_data.iloc[0]
            data_lainnya.append({
                'Kanwil': kanwil,
                'Target Setara Beras': row['target_setara_beras'],
                'Beras (a)': row['beras'],
                'GKG (b)': row['gkg'],
                'GKP (c)': row['gkp'],
                'Setara Beras (d)': row['setara_beras'],
                'Capaian (%)': row['capaian_persen']
            })
        else:
            # Jika tidak ada data, buat row dengan nilai 0
            data_lainnya.append({
                'Kanwil': kanwil,
                'Target Setara Beras': 0,
                'Beras (a)': 0,
                'GKG (b)': 0,
                'GKP (c)': 0,
                'Setara Beras (d)': 0,
                'Capaian (%)': 0
            })

    # Sort by Capaian (%) descending
    data_lainnya = sorted(data_lainnya, key=lambda x: x['Capaian (%)'], reverse=True)
    # Re-assign No after sorting
    for idx, row in enumerate(data_lainnya, 1):
        row['No'] = idx

    return data_sentra, data_lainnya

def create_kancab_table_from_rpc(p_nama_kanwil, p_akun_analitik, p_start_date, p_end_date):
    """
    Create kancab table menggunakan data dari RPC get_overview_setara_beras_kancab

    Parameters:
    - p_nama_kanwil: Value filter Kanwil
    - p_akun_analitik: Value filter akun analitik
    - p_start_date: value start date filter periode
    - p_end_date: value end date filter periode

    Returns:
    DataFrame dengan kolom: NO, Kancab, Target Setara Beras, Beras (a), GKG (b), GKP (c), Setara Beras (d), Capaian (%)
    """
    # PENTING: Jika p_nama_kanwil adalah None, berarti user tidak memilih kanwil spesifik
    # Dalam kasus ini, jangan tampilkan data kancab (karena akan bingung dari kanwil mana)
    if p_nama_kanwil is None:
        print(f"DEBUG create_kancab_table_from_rpc - p_nama_kanwil is None, tidak menampilkan data kancab")
        return pd.DataFrame()

    # Ambil data dari RPC
    df = get_tabel_realisasi_kancab(p_nama_kanwil, p_akun_analitik, p_start_date, p_end_date)

    # Jika tidak ada data yang dikembalikan, return DataFrame kosong
    if df.empty:
        print(f"DEBUG create_kancab_table_from_rpc - Data kosong untuk kanwil: {p_nama_kanwil}")
        return pd.DataFrame()

    # Rename kolom untuk sesuai dengan format yang diharapkan
    result_df = df.rename(columns={
        'kancab': 'Kancab',
        'target_setara_beras': 'Target Setara Beras',
        'beras_a': 'Beras (a)',
        'gkg_b': 'GKG (b)',
        'gkp_c': 'GKP (c)',
        'setara_beras_d': 'Setara Beras (d)',
        'capaian_persen': 'Capaian (%)'
    })

    # Add NO column
    result_df.insert(0, 'NO', range(1, len(result_df) + 1))

    # Calculate TOTAL row
    total_target = result_df['Target Setara Beras'].sum()
    total_beras = result_df['Beras (a)'].sum()
    total_gkg = result_df['GKG (b)'].sum()
    total_gkp = result_df['GKP (c)'].sum()
    total_setara = result_df['Setara Beras (d)'].sum()

    # Calculate total capaian
    if total_target > 0:
        total_capaian = (total_setara / total_target) * 100
    else:
        total_capaian = None

    total_row = {
        'NO': '',
        'Kancab': 'TOTAL KANWIL',
        'Target Setara Beras': total_target,
        'Beras (a)': total_beras,
        'GKG (b)': total_gkg,
        'GKP (c)': total_gkp,
        'Setara Beras (d)': total_setara,
        'Capaian (%)': total_capaian
    }
    result_df = pd.concat([result_df, pd.DataFrame([total_row])], ignore_index=True)

    return result_df

def create_line_chart_from_rpc(p_nama_kanwil, p_akun_analitik, p_start_date, p_end_date, start_date, end_date):
    """
    Create line chart menggunakan data dari RPC get_realisasi_harian_setara_beras
    Menampilkan 3 lines: BERAS, GKG, GKP dalam Ton

    Parameters:
    - p_nama_kanwil: Value filter Kanwil
    - p_akun_analitik: Value filter akun analitik
    - p_start_date: value start date filter periode (string)
    - p_end_date: value end date filter periode (string)
    - start_date: Start date object for display
    - end_date: End date object for display

    Returns:
    Plotly figure object
    """
    # Ambil data dari RPC
    df = get_tren_realisasi_kanwil(p_nama_kanwil, p_akun_analitik, p_start_date, p_end_date)

    if df.empty:
        # Return empty chart if no data with light theme
        fig = go.Figure()
        fig.update_layout(
            title="Tidak ada data untuk periode yang dipilih",
            xaxis_title="Tanggal",
            yaxis_title="Realisasi (Ton)",
            template='plotly_white',
            plot_bgcolor='white',
            paper_bgcolor='white',
            font=dict(color='#1f497d', size=12)
        )
        return fig

    # Convert tanggal to datetime
    df['tanggal'] = pd.to_datetime(df['tanggal'])

    # Create complete date range
    date_range = pd.date_range(start=start_date, end=end_date, freq='D')
    complete_dates = pd.DataFrame({'tanggal': date_range})

    # Merge dengan data untuk mengisi gap dengan 0
    df_beras = df[['tanggal', 'beras']].copy()
    df_gkg = df[['tanggal', 'gkg']].copy()
    df_gkp = df[['tanggal', 'gkp']].copy()

    # Merge dengan complete dates
    df_beras = complete_dates.merge(df_beras, on='tanggal', how='left').fillna(0)
    df_gkg = complete_dates.merge(df_gkg, on='tanggal', how='left').fillna(0)
    df_gkp = complete_dates.merge(df_gkp, on='tanggal', how='left').fillna(0)

    # Create figure with subplots
    from plotly.subplots import make_subplots

    fig = make_subplots(specs=[[{"secondary_y": False}]])

    # Add BERAS bar chart
    fig.add_trace(go.Bar(
        x=df_beras['tanggal'],
        y=df_beras['beras'],
        name='BERAS',
        marker=dict(
            color='rgba(31, 73, 125, 0.5)',
            line=dict(color='#1f497d', width=1)
        ),
        hovertemplate='<b>BERAS</b><br>Tanggal: %{x|%d %b %Y}<br>Realisasi: %{y:,.2f} Ton<extra></extra>'
    ))

    # Add GKP bar chart
    fig.add_trace(go.Bar(
        x=df_gkp['tanggal'],
        y=df_gkp['gkp'],
        name='GKP',
        marker=dict(
            color='rgba(75, 172, 198, 0.5)',
            line=dict(color='#4bacc6', width=1)
        ),
        hovertemplate='<b>GKP</b><br>Tanggal: %{x|%d %b %Y}<br>Realisasi: %{y:,.2f} Ton<extra></extra>'
    ))

    # Add GKG bar chart
    fig.add_trace(go.Bar(
        x=df_gkg['tanggal'],
        y=df_gkg['gkg'],
        name='GKG',
        marker=dict(
            color='rgba(157, 195, 230, 0.5)',
            line=dict(color='#9dc3e6', width=1)
        ),
        hovertemplate='<b>GKG</b><br>Tanggal: %{x|%d %b %Y}<br>Realisasi: %{y:,.2f} Ton<extra></extra>'
    ))

    # Filter data untuk line trace - hanya tampilkan jika value > 0
    df_beras_line = df_beras[df_beras['beras'] > 0].copy()
    df_gkp_line = df_gkp[df_gkp['gkp'] > 0].copy()
    df_gkg_line = df_gkg[df_gkg['gkg'] > 0].copy()

    # Add BERAS line trace (overlay) - hanya tampil jika value > 0
    fig.add_trace(go.Scatter(
        x=df_beras_line['tanggal'],
        y=df_beras_line['beras'],
        mode='lines',
        name='BERAS Trend',
        line=dict(color='#1f497d', width=3, shape='linear'),
        showlegend=False,
        hoverinfo='skip',
        connectgaps=False  # Jangan hubungkan gap (value 0)
    ))

    # Add GKP line trace (overlay) - hanya tampil jika value > 0
    fig.add_trace(go.Scatter(
        x=df_gkp_line['tanggal'],
        y=df_gkp_line['gkp'],
        mode='lines',
        name='GKP Trend',
        line=dict(color='#4bacc6', width=3, shape='linear'),
        showlegend=False,
        hoverinfo='skip',
        connectgaps=False  # Jangan hubungkan gap (value 0)
    ))

    # Add GKG line trace (overlay) - hanya tampil jika value > 0
    fig.add_trace(go.Scatter(
        x=df_gkg_line['tanggal'],
        y=df_gkg_line['gkg'],
        mode='lines',
        name='GKG Trend',
        line=dict(color='#9dc3e6', width=3, shape='linear'),
        showlegend=False,
        hoverinfo='skip',
        connectgaps=False  # Jangan hubungkan gap (value 0)
    ))

    # Update layout - Light theme
    fig.update_layout(
        template='plotly_white',  # Use light template
        xaxis_title="Tanggal",
        yaxis_title="Realisasi (Ton)",
        hovermode='x unified',
        barmode='group',
        bargap=0.3,
        bargroupgap=0.1,
        plot_bgcolor='white',
        paper_bgcolor='white',
        font=dict(color='#1f497d', size=12),
        legend=dict(
            orientation="h",
            yanchor="bottom",
            y=1.02,
            xanchor="right",
            x=1,
            bgcolor='rgba(255,255,255,0.8)',
            bordercolor='#1f497d',
            borderwidth=1
        ),
        margin=dict(l=50, r=50, t=50, b=50),
        dragmode='pan'  # Default cursor mode adalah Pan (geser/drag)
    )

    # Update axes - dengan tick labels berwarna dark
    fig.update_xaxes(
        showgrid=True,
        gridwidth=1,
        gridcolor='#e0e0e0',
        showline=True,
        linewidth=1,
        linecolor='#1f497d',
        title_font=dict(color='#1f497d'),
        tickfont=dict(color='#1f497d', size=10)  # Tick labels berwarna dark
    )

    fig.update_yaxes(
        showgrid=True,
        gridwidth=1,
        gridcolor='#e0e0e0',
        showline=True,
        linewidth=1,
        linecolor='#1f497d',
        title_font=dict(color='#1f497d'),
        tickfont=dict(color='#1f497d', size=10)  # Tick labels berwarna dark
    )

    return fig

def create_bar_chart_7days_from_rpc(p_nama_kanwil, p_akun_analitik, p_end_date, end_date):
    """
    Create bar chart 7 hari terakhir menggunakan data dari RPC

    Parameters:
    - p_nama_kanwil: Value filter Kanwil
    - p_akun_analitik: Value filter akun analitik
    - p_end_date: value end date filter periode (string)
    - end_date: End date object for display

    Returns:
    Plotly figure object
    """
    # Hitung 7 hari ke belakang terlebih dahulu
    if isinstance(end_date, str):
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
    else:
        end_date_obj = end_date

    start_date_7days = end_date_obj - timedelta(days=6)

    # Create complete date range untuk 7 hari (selalu ada, bahkan jika data kosong)
    date_range = pd.date_range(start=start_date_7days, end=end_date_obj, freq='D')
    complete_dates = pd.DataFrame({'tanggal': date_range})

    # Ambil data dari RPC
    df = get_realisasi_7_hari_terakhir(p_nama_kanwil, p_akun_analitik, p_end_date)

    # Jika ada data, convert tanggal dan merge
    if not df.empty:
        df['tanggal'] = pd.to_datetime(df['tanggal'])

        # Merge untuk setiap komoditi
        df_beras = complete_dates.merge(df[['tanggal', 'beras']], on='tanggal', how='left').fillna(0)
        df_gkg = complete_dates.merge(df[['tanggal', 'gkg']], on='tanggal', how='left').fillna(0)
        df_gkp = complete_dates.merge(df[['tanggal', 'gkp']], on='tanggal', how='left').fillna(0)
    else:
        # Jika tidak ada data, buat DataFrame dengan nilai 0 untuk semua tanggal
        df_beras = complete_dates.copy()
        df_beras['beras'] = 0
        df_gkg = complete_dates.copy()
        df_gkg['gkg'] = 0
        df_gkp = complete_dates.copy()
        df_gkp['gkp'] = 0

    # Create figure
    fig = go.Figure()

    # Add BERAS bar
    fig.add_trace(go.Bar(
        x=df_beras['tanggal'],
        y=df_beras['beras'],
        name='BERAS',
        marker_color='#1f497d',
        hovertemplate='<b>BERAS</b><br>Tanggal: %{x|%d %b %Y}<br>Realisasi: %{y:,.2f} Ton<extra></extra>'
    ))

    # Add GKP bar
    fig.add_trace(go.Bar(
        x=df_gkp['tanggal'],
        y=df_gkp['gkp'],
        name='GKP',
        marker_color='#4bacc6',
        hovertemplate='<b>GKP</b><br>Tanggal: %{x|%d %b %Y}<br>Realisasi: %{y:,.2f} Ton<extra></extra>'
    ))

    # Add GKG bar
    fig.add_trace(go.Bar(
        x=df_gkg['tanggal'],
        y=df_gkg['gkg'],
        name='GKG',
        marker_color='#9dc3e6',
        hovertemplate='<b>GKG</b><br>Tanggal: %{x|%d %b %Y}<br>Realisasi: %{y:,.2f} Ton<extra></extra>'
    ))

    # Update layout - Light theme
    fig.update_layout(
        template='plotly_white',  # Use light template
        xaxis_title="Tanggal",
        yaxis_title="Realisasi (Ton)",
        barmode='group',
        hovermode='x unified',
        plot_bgcolor='white',
        paper_bgcolor='white',
        font=dict(color='#1f497d', size=12),
        legend=dict(
            orientation="h",
            yanchor="bottom",
            y=1.02,
            xanchor="right",
            x=1,
            bgcolor='rgba(255,255,255,0.8)',
            bordercolor='#1f497d',
            borderwidth=1
        ),
        margin=dict(l=50, r=50, t=50, b=50),
        dragmode='pan'  # Default cursor mode adalah Pan (geser/drag)
    )

    # Update axes - dengan tick labels berwarna dark
    fig.update_xaxes(
        showgrid=True,
        gridwidth=1,
        gridcolor='#e0e0e0',
        showline=True,
        linewidth=1,
        linecolor='#1f497d',
        tickformat='%d-%m-%Y',  # Format: 29-12-2025
        title_font=dict(color='#1f497d'),
        tickfont=dict(color='#1f497d', size=10),  # Tick labels berwarna dark
        tickangle=-45  # Miringkan label agar tidak overlap
    )

    fig.update_yaxes(
        showgrid=True,
        gridwidth=1,
        gridcolor='#e0e0e0',
        showline=True,
        linewidth=1,
        linecolor='#1f497d',
        title_font=dict(color='#1f497d'),
        tickfont=dict(color='#1f497d', size=10)  # Tick labels berwarna dark
    )

    return fig

def calculate_setara_beras(df):
    """
    Calculate Setara Beras from a dataframe
    Formula: d = a + 0.635*b + 0.53375*c
    where:
    a = BERAS (BERAS PREMIUM + BERAS MEDIUM)
    b = GKG (GABAH with spesifikasi containing 'GKG')
    c = GKP (GABAH with spesifikasi containing 'GKP')
    Returns value in Ton
    """
    # a) BERAS (BERAS MEDIUM + BERAS PREMIUM) - convert to Ton
    beras = df[
        df['Komoditi'].isin(['BERAS MEDIUM', 'BERAS PREMIUM'])
    ]['In / Out'].sum() / 1000

    # b) GKG - Gabah dengan spesifikasi GKG - convert to Ton
    gkg = df[
        (df['Komoditi'] == 'GABAH') &
        (df['spesifikasi'].astype(str).str.contains('GKG', case=False, na=False))
    ]['In / Out'].sum() / 1000

    # c) GKP - Gabah dengan spesifikasi GKP - convert to Ton
    gkp = df[
        (df['Komoditi'] == 'GABAH') &
        (df['spesifikasi'].astype(str).str.contains('GKP', case=False, na=False))
    ]['In / Out'].sum() / 1000

    # d) Setara Beras = a + 0.635*b + 0.53375*c
    setara_beras = beras + (0.635 * gkg) + (0.53375 * gkp)

    return setara_beras


def render_summary_table_html(data_sentra, data_lainnya, start_date, end_date):
    """Render summary table with Kanwil Sentra Produksi and Kanwil Lainnya in HTML"""

    def format_value(val, decimal=2):
        if pd.isna(val) or val is None or val == 0:
            return '-'
        return f'{val:,.{decimal}f}'

    # Build rows for Kanwil Sentra Produksi
    rows_sentra = ""
    total_sentra = {
        'Target Setara Beras': 0,
        'Beras (a)': 0,
        'GKG (b)': 0,
        'GKP (c)': 0,
        'Setara Beras (d)': 0
    }

    for row in data_sentra:
        rows_sentra += '<tr>'
        rows_sentra += f'<td>{row["No"]}</td>'
        rows_sentra += f'<td style="text-align: left; padding-left: 8px;">{row["Kanwil"]}</td>'
        rows_sentra += f'<td>{format_value(row["Target Setara Beras"])}</td>'
        rows_sentra += f'<td>{format_value(row["Beras (a)"])}</td>'
        rows_sentra += f'<td>{format_value(row["GKG (b)"])}</td>'
        rows_sentra += f'<td>{format_value(row["GKP (c)"])}</td>'
        rows_sentra += f'<td>{format_value(row["Setara Beras (d)"])}</td>'
        rows_sentra += f'<td>{format_value(row["Capaian (%)"], 1)}%</td>'
        rows_sentra += '</tr>'

        # Accumulate totals
        total_sentra['Target Setara Beras'] += row['Target Setara Beras']
        total_sentra['Beras (a)'] += row['Beras (a)']
        total_sentra['GKG (b)'] += row['GKG (b)']
        total_sentra['GKP (c)'] += row['GKP (c)']
        total_sentra['Setara Beras (d)'] += row['Setara Beras (d)']

    # Calculate capaian sentra
    capaian_sentra = (total_sentra['Setara Beras (d)'] / total_sentra['Target Setara Beras'] * 100) if total_sentra['Target Setara Beras'] > 0 else 0

    # Build rows for Kanwil Lainnya
    rows_lainnya = ""
    total_lainnya = {
        'Target Setara Beras': 0,
        'Beras (a)': 0,
        'GKG (b)': 0,
        'GKP (c)': 0,
        'Setara Beras (d)': 0
    }

    for row in data_lainnya:
        rows_lainnya += '<tr>'
        rows_lainnya += f'<td>{row["No"]}</td>'
        rows_lainnya += f'<td style="text-align: left; padding-left: 8px;">{row["Kanwil"]}</td>'
        rows_lainnya += f'<td>{format_value(row["Target Setara Beras"])}</td>'
        rows_lainnya += f'<td>{format_value(row["Beras (a)"])}</td>'
        rows_lainnya += f'<td>{format_value(row["GKG (b)"])}</td>'
        rows_lainnya += f'<td>{format_value(row["GKP (c)"])}</td>'
        rows_lainnya += f'<td>{format_value(row["Setara Beras (d)"])}</td>'
        rows_lainnya += f'<td>{format_value(row["Capaian (%)"], 1)}%</td>'
        rows_lainnya += '</tr>'

        # Accumulate totals
        total_lainnya['Target Setara Beras'] += row['Target Setara Beras']
        total_lainnya['Beras (a)'] += row['Beras (a)']
        total_lainnya['GKG (b)'] += row['GKG (b)']
        total_lainnya['GKP (c)'] += row['GKP (c)']
        total_lainnya['Setara Beras (d)'] += row['Setara Beras (d)']

    # Calculate capaian lainnya
    capaian_lainnya = (total_lainnya['Setara Beras (d)'] / total_lainnya['Target Setara Beras'] * 100) if total_lainnya['Target Setara Beras'] > 0 else 0

    # Calculate grand total
    total_seindo = {
        'Target Setara Beras': total_sentra['Target Setara Beras'] + total_lainnya['Target Setara Beras'],
        'Beras (a)': total_sentra['Beras (a)'] + total_lainnya['Beras (a)'],
        'GKG (b)': total_sentra['GKG (b)'] + total_lainnya['GKG (b)'],
        'GKP (c)': total_sentra['GKP (c)'] + total_lainnya['GKP (c)'],
        'Setara Beras (d)': total_sentra['Setara Beras (d)'] + total_lainnya['Setara Beras (d)']
    }
    capaian_seindo = (total_seindo['Setara Beras (d)'] / total_seindo['Target Setara Beras'] * 100) if total_seindo['Target Setara Beras'] > 0 else 0

    html = f'''
<table border="1" cellspacing="0" cellpadding="8" style="border-collapse: collapse; font-size: 12px; width: 100%; font-family: Arial;">
    <thead>
        <tr style="background: rgb(31, 73, 125); color: white; font-weight: bold;">
            <th rowspan="2">No.</th>
            <th rowspan="2">Kanwil</th>
            <th rowspan="2">Target Setara Beras</th>
            <th colspan="5">Realisasi Periode: {start_date.strftime("%d %b %Y")} - {end_date.strftime("%d %b %Y")}</th>
        </tr>
        <tr style="background: rgb(75, 172, 198); color: white; font-weight: bold;">
            <th>Beras (a)</th>
            <th>GKG (b)</th>
            <th>GKP (c)</th>
            <th>Setara Beras (d)</th>
            <th>Capaian (%)</th>
        </tr>
    </thead>
    <tbody>
        <!-- BAGIAN A: Kanwil Sentra Produksi -->
        <tr style="background:#fdc128; font-weight:bold;">
            <td colspan="8">a) Kanwil Sentra Produksi</td>
        </tr>
        {rows_sentra}
        <!-- SUBTOTAL A -->
        <tr style="background:#ffe599; font-weight:bold;">
            <td colspan="2">Total Kanwil Sentra Produksi</td>
            <td>{format_value(total_sentra['Target Setara Beras'])}</td>
            <td>{format_value(total_sentra['Beras (a)'])}</td>
            <td>{format_value(total_sentra['GKG (b)'])}</td>
            <td>{format_value(total_sentra['GKP (c)'])}</td>
            <td>{format_value(total_sentra['Setara Beras (d)'])}</td>
            <td>{format_value(capaian_sentra, 1)}%</td>
        </tr>

        <!-- BAGIAN B: Kanwil Lainnya -->
        <tr style="background:#fdc128; font-weight:bold;">
            <td colspan="8">b) Kanwil Lainnya</td>
        </tr>
        {rows_lainnya}
        <!-- SUBTOTAL B -->
        <tr style="background:#ffe599; font-weight:bold;">
            <td colspan="2">Total Kanwil Lainnya</td>
            <td>{format_value(total_lainnya['Target Setara Beras'])}</td>
            <td>{format_value(total_lainnya['Beras (a)'])}</td>
            <td>{format_value(total_lainnya['GKG (b)'])}</td>
            <td>{format_value(total_lainnya['GKP (c)'])}</td>
            <td>{format_value(total_lainnya['Setara Beras (d)'])}</td>
            <td>{format_value(capaian_lainnya, 1)}%</td>
        </tr>

        <!-- TOTAL SE-INDO -->
        <tr style="background: rgb(31, 73, 125); color: white; font-weight: bold;">
            <td colspan="2">TOTAL SE-INDO</td>
            <td>{format_value(total_seindo['Target Setara Beras'])}</td>
            <td>{format_value(total_seindo['Beras (a)'])}</td>
            <td>{format_value(total_seindo['GKG (b)'])}</td>
            <td>{format_value(total_seindo['GKP (c)'])}</td>
            <td>{format_value(total_seindo['Setara Beras (d)'])}</td>
            <td>{format_value(capaian_seindo, 1)}%</td>
        </tr>
    </tbody>
</table>
'''
    return html, total_sentra, total_lainnya, total_seindo, capaian_sentra, capaian_lainnya, capaian_seindo

def render_kancab_table_html(df, start_date, end_date):
    """Render Kancab table in HTML with new structure"""
    if df.empty:
        return "<p>Tidak ada data</p>"

    def format_value(val, decimal=2, is_percent=False):
        if pd.isna(val) or val is None:
            return '-'
        if val == 0:
            return '-'
        if is_percent:
            return f'{val:,.1f}%'
        return f'{val:,.{decimal}f}'

    # Build rows
    rows_html = ""
    for _, row in df.iterrows():
        is_total = row['Kancab'] == 'TOTAL KANWIL'
        if is_total:
            row_style = 'background: #ffe599; color: black; font-weight: bold;'
        else:
            row_style = 'color: black;'

        rows_html += f'<tr style="{row_style}">'
        rows_html += f'<td>{row["NO"] if row["NO"] != "" else ""}</td>'
        rows_html += f'<td style="text-align: left; padding-left: 8px;">{row["Kancab"]}</td>'
        rows_html += f'<td>{format_value(row["Target Setara Beras"])}</td>'
        rows_html += f'<td>{format_value(row["Beras (a)"])}</td>'
        rows_html += f'<td>{format_value(row["GKG (b)"])}</td>'
        rows_html += f'<td>{format_value(row["GKP (c)"])}</td>'
        rows_html += f'<td>{format_value(row["Setara Beras (d)"])}</td>'
        rows_html += f'<td>{format_value(row["Capaian (%)"], is_percent=True)}</td>'
        rows_html += '</tr>'

    html = f'''
<table border="1" cellspacing="0" cellpadding="8" style="border-collapse: collapse; font-size: 12px; width: 100%; font-family: Arial;">
    <thead>
        <tr style="background: rgb(31, 73, 125); color: white; font-weight: bold;">
            <th rowspan="2">No.</th>
            <th rowspan="2">Kancab</th>
            <th rowspan="2">Target Setara Beras</th>
            <th colspan="5">Realisasi S. d. {end_date.strftime("%d %b %Y")}</th>
        </tr>
        <tr style="background: rgb(75, 172, 198); color: white; font-weight: bold;">
            <th>Beras (a)</th>
            <th>GKG (b)</th>
            <th>GKP (c)</th>
            <th>Setara Beras (d)</th>
            <th>Capaian (%)</th>
        </tr>
    </thead>
    <tbody>
        {rows_html}
    </tbody>
</table>
'''
    return html

def create_kancab_excel_export(df, end_date):
    """Create Excel file for Kancab table with same styling as HTML"""
    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Write empty dataframe to create workbook
        pd.DataFrame().to_excel(writer, sheet_name='Kancab', index=False)

        # Get workbook and worksheet
        workbook = writer.book
        worksheet = workbook['Kancab']

        # Define styles
        header_fill_dark = PatternFill(start_color='1f497d', end_color='1f497d', fill_type='solid')
        header_fill_light = PatternFill(start_color='4bacc6', end_color='4bacc6', fill_type='solid')
        total_fill = PatternFill(start_color='b6d7a8', end_color='b6d7a8', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        bold_font = Font(bold=True, size=10)
        normal_font = Font(size=10)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write header row 1
        worksheet.cell(row=1, column=1, value='No.').fill = header_fill_dark
        worksheet.cell(row=1, column=1).font = header_font
        worksheet.cell(row=1, column=1).alignment = center_align
        worksheet.cell(row=1, column=1).border = thin_border
        worksheet.merge_cells('A1:A2')

        worksheet.cell(row=1, column=2, value='Kancab').fill = header_fill_dark
        worksheet.cell(row=1, column=2).font = header_font
        worksheet.cell(row=1, column=2).alignment = center_align
        worksheet.cell(row=1, column=2).border = thin_border
        worksheet.merge_cells('B1:B2')

        worksheet.cell(row=1, column=3, value='Target Setara Beras').fill = header_fill_dark
        worksheet.cell(row=1, column=3).font = header_font
        worksheet.cell(row=1, column=3).alignment = center_align
        worksheet.cell(row=1, column=3).border = thin_border
        worksheet.merge_cells('C1:C2')

        worksheet.cell(row=1, column=4, value=f'Realisasi S. d. {end_date.strftime("%d %b %Y")}').fill = header_fill_dark
        worksheet.cell(row=1, column=4).font = header_font
        worksheet.cell(row=1, column=4).alignment = center_align
        worksheet.cell(row=1, column=4).border = thin_border
        worksheet.merge_cells('D1:H1')

        # Write header row 2
        headers_row2 = ['Beras (a)', 'GKG (b)', 'GKP (c)', 'Setara Beras (d)', 'Capaian (%)']
        for col_idx, header in enumerate(headers_row2, start=4):
            cell = worksheet.cell(row=2, column=col_idx, value=header)
            cell.fill = header_fill_light
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        # Write data rows
        current_row = 3
        for _, row in df.iterrows():
            is_total = row['Kancab'] == 'TOTAL KANWIL'

            worksheet.cell(row=current_row, column=1, value=row['NO'] if row['NO'] != '' else '')
            worksheet.cell(row=current_row, column=1).alignment = center_align
            worksheet.cell(row=current_row, column=1).border = thin_border

            worksheet.cell(row=current_row, column=2, value=row['Kancab'])
            worksheet.cell(row=current_row, column=2).alignment = left_align
            worksheet.cell(row=current_row, column=2).border = thin_border

            worksheet.cell(row=current_row, column=3, value=row['Target Setara Beras'] if pd.notna(row['Target Setara Beras']) else None)
            worksheet.cell(row=current_row, column=3).number_format = '#,##0.00'
            worksheet.cell(row=current_row, column=3).alignment = center_align
            worksheet.cell(row=current_row, column=3).border = thin_border

            worksheet.cell(row=current_row, column=4, value=row['Beras (a)'] if pd.notna(row['Beras (a)']) and row['Beras (a)'] > 0 else None)
            worksheet.cell(row=current_row, column=4).number_format = '#,##0.00'
            worksheet.cell(row=current_row, column=4).alignment = center_align
            worksheet.cell(row=current_row, column=4).border = thin_border

            worksheet.cell(row=current_row, column=5, value=row['GKG (b)'] if pd.notna(row['GKG (b)']) and row['GKG (b)'] > 0 else None)
            worksheet.cell(row=current_row, column=5).number_format = '#,##0.00'
            worksheet.cell(row=current_row, column=5).alignment = center_align
            worksheet.cell(row=current_row, column=5).border = thin_border

            worksheet.cell(row=current_row, column=6, value=row['GKP (c)'] if pd.notna(row['GKP (c)']) and row['GKP (c)'] > 0 else None)
            worksheet.cell(row=current_row, column=6).number_format = '#,##0.00'
            worksheet.cell(row=current_row, column=6).alignment = center_align
            worksheet.cell(row=current_row, column=6).border = thin_border

            worksheet.cell(row=current_row, column=7, value=row['Setara Beras (d)'] if pd.notna(row['Setara Beras (d)']) and row['Setara Beras (d)'] > 0 else None)
            worksheet.cell(row=current_row, column=7).number_format = '#,##0.00'
            worksheet.cell(row=current_row, column=7).alignment = center_align
            worksheet.cell(row=current_row, column=7).border = thin_border

            worksheet.cell(row=current_row, column=8, value=row['Capaian (%)'] if pd.notna(row['Capaian (%)']) else None)
            worksheet.cell(row=current_row, column=8).number_format = '0.0"%"'
            worksheet.cell(row=current_row, column=8).alignment = center_align
            worksheet.cell(row=current_row, column=8).border = thin_border

            # Apply styling
            if is_total:
                for col_idx in range(1, 9):
                    worksheet.cell(row=current_row, column=col_idx).fill = total_fill
                    worksheet.cell(row=current_row, column=col_idx).font = bold_font
            else:
                for col_idx in range(1, 9):
                    worksheet.cell(row=current_row, column=col_idx).font = normal_font

            current_row += 1

        # Set column widths
        worksheet.column_dimensions['A'].width = 5
        worksheet.column_dimensions['B'].width = 40
        worksheet.column_dimensions['C'].width = 18
        worksheet.column_dimensions['D'].width = 12
        worksheet.column_dimensions['E'].width = 12
        worksheet.column_dimensions['F'].width = 12
        worksheet.column_dimensions['G'].width = 18
        worksheet.column_dimensions['H'].width = 14

    output.seek(0)
    return output

def create_complex_table(df_filtered, selected_date, selected_kanwil_list):
    """
    Create complex table with multi-row header
    Columns: NO, KANWIL, TARGET (BERAS, GABAH),
             REALISASI SD KEMARIN, REALISASI HARI INI, REALISASI SD TGL,
             CAPAIAN (% BERAS, % GABAH)
    """
    # Tanggal kemarin
    tanggal_kemarin = selected_date - timedelta(days=1)
    tanggal_hari_ini = selected_date

    # Use the passed list of ALL selected kanwil
    # This ensures we show all selected kanwil even if they don't have BERAS or GABAH
    kanwil_list = sorted([k for k in selected_kanwil_list if pd.notna(k)])

    result_data = []

    for kanwil in kanwil_list:
        df_kanwil = df_filtered[df_filtered['kanwil'] == kanwil]

        # TARGET BERAS (from Kuantum PO) - BERAS MEDIUM + BERAS PREMIUM
        target_beras = df_kanwil[
            df_kanwil['Komoditi'].isin(['BERAS MEDIUM', 'BERAS PREMIUM'])
        ]['Kuantum PO (Kg)'].sum() / 1000  # to Ton

        # TARGET GABAH (from Kuantum PO)
        target_gabah = df_kanwil[
            df_kanwil['Komoditi'] == 'GABAH'
        ]['Kuantum PO (Kg)'].sum() / 1000  # to Ton

        # Check if has data for BERAS or GABAH
        has_beras = target_beras > 0
        has_gabah = target_gabah > 0

        # REALISASI SD KEMARIN - BERAS
        real_sd_kemarin_beras = df_kanwil[
            (df_kanwil['Tanggal Penerimaan'].dt.date < tanggal_kemarin) &
            (df_kanwil['Komoditi'].isin(['BERAS MEDIUM', 'BERAS PREMIUM']))
        ]['In / Out'].sum() / 1000 if has_beras else None  # to Ton

        # REALISASI SD KEMARIN - GABAH
        real_sd_kemarin_gabah = df_kanwil[
            (df_kanwil['Tanggal Penerimaan'].dt.date < tanggal_kemarin) &
            (df_kanwil['Komoditi'] == 'GABAH')
        ]['In / Out'].sum() / 1000 if has_gabah else None  # to Ton

        # REALISASI HARI INI - BERAS (only on selected date)
        real_hari_ini_beras = df_kanwil[
            (df_kanwil['Tanggal Penerimaan'].dt.date == tanggal_hari_ini) &
            (df_kanwil['Komoditi'].isin(['BERAS MEDIUM', 'BERAS PREMIUM']))
        ]['In / Out'].sum() / 1000 if has_beras else None  # to Ton

        # REALISASI HARI INI - GABAH (only on selected date)
        real_hari_ini_gabah = df_kanwil[
            (df_kanwil['Tanggal Penerimaan'].dt.date == tanggal_hari_ini) &
            (df_kanwil['Komoditi'] == 'GABAH')
        ]['In / Out'].sum() / 1000 if has_gabah else None  # to Ton

        # REALISASI SD TGL - BERAS (until selected date)
        real_sd_tgl_beras = df_kanwil[
            (df_kanwil['Tanggal Penerimaan'].dt.date <= tanggal_hari_ini) &
            (df_kanwil['Komoditi'].isin(['BERAS MEDIUM', 'BERAS PREMIUM']))
        ]['In / Out'].sum() / 1000 if has_beras else None  # to Ton

        # REALISASI SD TGL - GABAH (until selected date)
        real_sd_tgl_gabah = df_kanwil[
            (df_kanwil['Tanggal Penerimaan'].dt.date <= tanggal_hari_ini) &
            (df_kanwil['Komoditi'] == 'GABAH')
        ]['In / Out'].sum() / 1000 if has_gabah else None  # to Ton

        # CAPAIAN %
        capaian_beras = (real_sd_tgl_beras / target_beras * 100) if (has_beras and target_beras > 0) else None
        capaian_gabah = (real_sd_tgl_gabah / target_gabah * 100) if (has_gabah and target_gabah > 0) else None

        result_data.append({
            'Kanwil': kanwil,
            'Target Beras': target_beras if has_beras else None,
            'Target Gabah': target_gabah if has_gabah else None,
            'Real SD Kemarin Beras': real_sd_kemarin_beras,
            'Real SD Kemarin Gabah': real_sd_kemarin_gabah,
            'Real Hari Ini Beras': real_hari_ini_beras,
            'Real Hari Ini Gabah': real_hari_ini_gabah,
            'Real SD Tgl Beras': real_sd_tgl_beras,
            'Real SD Tgl Gabah': real_sd_tgl_gabah,
            'Capaian Beras %': capaian_beras,
            'Capaian Gabah %': capaian_gabah
        })

    result_df = pd.DataFrame(result_data)

    # Sort by Target Beras descending (None values will be at the bottom)
    result_df = result_df.sort_values('Target Beras', ascending=False, na_position='last').reset_index(drop=True)

    # Add NO column
    result_df.insert(0, 'NO', range(1, len(result_df) + 1))

    # Calculate TOTAL row (sum only non-None values)
    total_target_beras = result_df['Target Beras'].dropna().sum()
    total_target_gabah = result_df['Target Gabah'].dropna().sum()
    total_real_sd_kemarin_beras = result_df['Real SD Kemarin Beras'].dropna().sum()
    total_real_sd_kemarin_gabah = result_df['Real SD Kemarin Gabah'].dropna().sum()
    total_real_hari_ini_beras = result_df['Real Hari Ini Beras'].dropna().sum()
    total_real_hari_ini_gabah = result_df['Real Hari Ini Gabah'].dropna().sum()
    total_real_sd_tgl_beras = result_df['Real SD Tgl Beras'].dropna().sum()
    total_real_sd_tgl_gabah = result_df['Real SD Tgl Gabah'].dropna().sum()

    total_row = {
        'NO': '',
        'Kanwil': 'TOTAL',
        'Target Beras': total_target_beras if total_target_beras > 0 else None,
        'Target Gabah': total_target_gabah if total_target_gabah > 0 else None,
        'Real SD Kemarin Beras': total_real_sd_kemarin_beras if total_real_sd_kemarin_beras > 0 else None,
        'Real SD Kemarin Gabah': total_real_sd_kemarin_gabah if total_real_sd_kemarin_gabah > 0 else None,
        'Real Hari Ini Beras': total_real_hari_ini_beras if total_real_hari_ini_beras > 0 else None,
        'Real Hari Ini Gabah': total_real_hari_ini_gabah if total_real_hari_ini_gabah > 0 else None,
        'Real SD Tgl Beras': total_real_sd_tgl_beras if total_real_sd_tgl_beras > 0 else None,
        'Real SD Tgl Gabah': total_real_sd_tgl_gabah if total_real_sd_tgl_gabah > 0 else None,
        'Capaian Beras %': (total_real_sd_tgl_beras / total_target_beras * 100) if total_target_beras > 0 else None,
        'Capaian Gabah %': (total_real_sd_tgl_gabah / total_target_gabah * 100) if total_target_gabah > 0 else None
    }

    result_df = pd.concat([result_df, pd.DataFrame([total_row])], ignore_index=True)

    return result_df, tanggal_kemarin, tanggal_hari_ini

def render_complex_table_html(df, tanggal_kemarin, tanggal_hari_ini):
    """Render table with complex multi-row header in HTML"""

    # Build HTML row by row
    rows_html = ""
    for _, row in df.iterrows():
        is_total = row['Kanwil'] == 'TOTAL'
        row_style = 'font-weight:bold; background:#ddebf7;' if is_total else ''

        # Helper function to format value or show '-'
        def format_value(val, decimal=2, is_percent=False):
            if pd.isna(val) or val is None:
                return '-'
            if is_percent:
                return f'• {val:,.1f}%'
            return f'• {val:,.{decimal}f}'

        rows_html += f'<tr style="{row_style}">'
        rows_html += f'<td>{row["NO"]}</td>'
        rows_html += f'<td style="text-align: left; padding-left: 10px;">{row["Kanwil"]}</td>'
        rows_html += f'<td>{format_value(row["Target Beras"])}</td>'
        rows_html += f'<td>{format_value(row["Target Gabah"])}</td>'
        rows_html += f'<td>{format_value(row["Real SD Kemarin Beras"])}</td>'
        rows_html += f'<td>{format_value(row["Real SD Kemarin Gabah"])}</td>'
        rows_html += f'<td>{format_value(row["Real Hari Ini Beras"])}</td>'
        rows_html += f'<td>{format_value(row["Real Hari Ini Gabah"])}</td>'
        rows_html += f'<td>{format_value(row["Real SD Tgl Beras"])}</td>'
        rows_html += f'<td>{format_value(row["Real SD Tgl Gabah"])}</td>'
        rows_html += f'<td>{format_value(row["Capaian Beras %"], is_percent=True)}</td>'
        rows_html += f'<td>{format_value(row["Capaian Gabah %"], is_percent=True)}</td>'
        rows_html += '</tr>'

    html = f'''
<table border="1" cellspacing="0" cellpadding="8" style="border-collapse: collapse; font-size: 12px; text-align: center; width: 100%; font-family: Arial;">
    <thead>
        <tr style="background:#1f497d; color:white; font-weight:bold;">
            <th rowspan="3" style="vertical-align: middle;">NO</th>
            <th rowspan="3" style="vertical-align: middle;">KANWIL</th>
            <th colspan="2" rowspan="2" style="vertical-align: middle;">TARGET</th>
            <th colspan="2">REALISASI SD. KEMARIN</th>
            <th colspan="2">REALISASI HARI INI</th>
            <th colspan="2">REALISASI SD. TGL</th>
            <th colspan="2" rowspan="2" style="vertical-align: middle;">CAPAIAN</th>
        </tr>
        <tr style="background:#1f497d; color:white; font-weight:bold;">
            <th colspan="2">{tanggal_kemarin.strftime("%d %B %Y")}</th>
            <th colspan="2">{tanggal_hari_ini.strftime("%d %B %Y")}</th>
            <th colspan="2">{tanggal_hari_ini.strftime("%d %B %Y")}</th>
        </tr>
        <tr style="background:#4bacc6; color:white; font-weight:bold;">
            <th>BERAS</th><th>GABAH</th>
            <th>BERAS</th><th>GABAH</th>
            <th>BERAS</th><th>GABAH</th>
            <th>BERAS</th><th>GABAH</th>
            <th>% BERAS</th><th>% GABAH</th>
        </tr>
    </thead>
    <tbody>
        {rows_html}
    </tbody>
</table>
'''

    return html

def create_summary_excel_export(data_sentra, data_lainnya, start_date, end_date, total_sentra, total_lainnya, total_seindo, capaian_sentra, capaian_lainnya, capaian_seindo):
    """Create Excel file for summary table with same styling as HTML"""
    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Write empty dataframe to create workbook
        pd.DataFrame().to_excel(writer, sheet_name='Summary', index=False)

        # Get workbook and worksheet
        workbook = writer.book
        worksheet = workbook['Summary']

        # Define styles
        header_fill = PatternFill(start_color='f2f2f2', end_color='f2f2f2', fill_type='solid')
        sentra_fill = PatternFill(start_color='ffe599', end_color='ffe599', fill_type='solid')
        lainnya_fill = PatternFill(start_color='c9daf8', end_color='c9daf8', fill_type='solid')
        seindo_fill = PatternFill(start_color='b6d7a8', end_color='b6d7a8', fill_type='solid')
        header_font = Font(bold=True, size=11)
        bold_font = Font(bold=True, size=10)
        normal_font = Font(size=10)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write header row 1
        worksheet.cell(row=1, column=1, value='No.').fill = header_fill
        worksheet.cell(row=1, column=1).font = header_font
        worksheet.cell(row=1, column=1).alignment = center_align
        worksheet.cell(row=1, column=1).border = thin_border
        worksheet.merge_cells('A1:A2')

        worksheet.cell(row=1, column=2, value='Kanwil').fill = header_fill
        worksheet.cell(row=1, column=2).font = header_font
        worksheet.cell(row=1, column=2).alignment = center_align
        worksheet.cell(row=1, column=2).border = thin_border
        worksheet.merge_cells('B1:B2')

        worksheet.cell(row=1, column=3, value='Target Setara Beras').fill = header_fill
        worksheet.cell(row=1, column=3).font = header_font
        worksheet.cell(row=1, column=3).alignment = center_align
        worksheet.cell(row=1, column=3).border = thin_border
        worksheet.merge_cells('C1:C2')

        worksheet.cell(row=1, column=4, value=f'Realisasi Periode: {start_date.strftime("%d %b %Y")} - {end_date.strftime("%d %b %Y")}').fill = header_fill
        worksheet.cell(row=1, column=4).font = header_font
        worksheet.cell(row=1, column=4).alignment = center_align
        worksheet.cell(row=1, column=4).border = thin_border
        worksheet.merge_cells('D1:H1')

        # Write header row 2
        headers_row2 = ['Beras (a)', 'GKG (b)', 'GKP (c)', 'Setara Beras (d)', 'Capaian (%)']
        for col_idx, header in enumerate(headers_row2, start=4):
            cell = worksheet.cell(row=2, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        # Current row
        current_row = 3

        # Write Kanwil Sentra Produksi header
        cell = worksheet.cell(row=current_row, column=1, value='a) Kanwil Sentra Produksi')
        cell.fill = sentra_fill
        cell.font = bold_font
        cell.alignment = left_align
        cell.border = thin_border
        worksheet.merge_cells(f'A{current_row}:H{current_row}')
        current_row += 1

        # Write data for Kanwil Sentra Produksi
        for row_data in data_sentra:
            worksheet.cell(row=current_row, column=1, value=row_data['No']).alignment = center_align
            worksheet.cell(row=current_row, column=1).border = thin_border
            worksheet.cell(row=current_row, column=2, value=row_data['Kanwil']).alignment = left_align
            worksheet.cell(row=current_row, column=2).border = thin_border
            worksheet.cell(row=current_row, column=3, value=row_data['Target Setara Beras']).number_format = '#,##0.00'
            worksheet.cell(row=current_row, column=3).alignment = center_align
            worksheet.cell(row=current_row, column=3).border = thin_border
            worksheet.cell(row=current_row, column=4, value=row_data['Beras (a)']).number_format = '#,##0.00'
            worksheet.cell(row=current_row, column=4).alignment = center_align
            worksheet.cell(row=current_row, column=4).border = thin_border
            worksheet.cell(row=current_row, column=5, value=row_data['GKG (b)']).number_format = '#,##0.00'
            worksheet.cell(row=current_row, column=5).alignment = center_align
            worksheet.cell(row=current_row, column=5).border = thin_border
            worksheet.cell(row=current_row, column=6, value=row_data['GKP (c)']).number_format = '#,##0.00'
            worksheet.cell(row=current_row, column=6).alignment = center_align
            worksheet.cell(row=current_row, column=6).border = thin_border
            worksheet.cell(row=current_row, column=7, value=row_data['Setara Beras (d)']).number_format = '#,##0.00'
            worksheet.cell(row=current_row, column=7).alignment = center_align
            worksheet.cell(row=current_row, column=7).border = thin_border
            worksheet.cell(row=current_row, column=8, value=row_data['Capaian (%)']).number_format = '0.0"%"'
            worksheet.cell(row=current_row, column=8).alignment = center_align
            worksheet.cell(row=current_row, column=8).border = thin_border
            current_row += 1

        # Write Total Kanwil Sentra Produksi
        worksheet.cell(row=current_row, column=1, value='Total Kanwil Sentra Produksi').fill = sentra_fill
        worksheet.cell(row=current_row, column=1).font = bold_font
        worksheet.cell(row=current_row, column=1).alignment = left_align
        worksheet.cell(row=current_row, column=1).border = thin_border
        worksheet.merge_cells(f'A{current_row}:B{current_row}')
        worksheet.cell(row=current_row, column=3, value=total_sentra['Target Setara Beras']).number_format = '#,##0.00'
        worksheet.cell(row=current_row, column=3).fill = sentra_fill
        worksheet.cell(row=current_row, column=3).font = bold_font
        worksheet.cell(row=current_row, column=3).alignment = center_align
        worksheet.cell(row=current_row, column=3).border = thin_border
        worksheet.cell(row=current_row, column=4, value=total_sentra['Beras (a)']).number_format = '#,##0.00'
        worksheet.cell(row=current_row, column=4).fill = sentra_fill
        worksheet.cell(row=current_row, column=4).font = bold_font
        worksheet.cell(row=current_row, column=4).alignment = center_align
        worksheet.cell(row=current_row, column=4).border = thin_border
        worksheet.cell(row=current_row, column=5, value=total_sentra['GKG (b)']).number_format = '#,##0.00'
        worksheet.cell(row=current_row, column=5).fill = sentra_fill
        worksheet.cell(row=current_row, column=5).font = bold_font
        worksheet.cell(row=current_row, column=5).alignment = center_align
        worksheet.cell(row=current_row, column=5).border = thin_border
        worksheet.cell(row=current_row, column=6, value=total_sentra['GKP (c)']).number_format = '#,##0.00'
        worksheet.cell(row=current_row, column=6).fill = sentra_fill
        worksheet.cell(row=current_row, column=6).font = bold_font
        worksheet.cell(row=current_row, column=6).alignment = center_align
        worksheet.cell(row=current_row, column=6).border = thin_border
        worksheet.cell(row=current_row, column=7, value=total_sentra['Setara Beras (d)']).number_format = '#,##0.00'
        worksheet.cell(row=current_row, column=7).fill = sentra_fill
        worksheet.cell(row=current_row, column=7).font = bold_font
        worksheet.cell(row=current_row, column=7).alignment = center_align
        worksheet.cell(row=current_row, column=7).border = thin_border
        worksheet.cell(row=current_row, column=8, value=capaian_sentra).number_format = '0.0"%"'
        worksheet.cell(row=current_row, column=8).fill = sentra_fill
        worksheet.cell(row=current_row, column=8).font = bold_font
        worksheet.cell(row=current_row, column=8).alignment = center_align
        worksheet.cell(row=current_row, column=8).border = thin_border
        current_row += 1

        # Write Kanwil Lainnya header
        cell = worksheet.cell(row=current_row, column=1, value='b) Kanwil Lainnya')
        cell.fill = lainnya_fill
        cell.font = bold_font
        cell.alignment = left_align
        cell.border = thin_border
        worksheet.merge_cells(f'A{current_row}:H{current_row}')
        current_row += 1

        # Write data for Kanwil Lainnya
        for row_data in data_lainnya:
            worksheet.cell(row=current_row, column=1, value=row_data['No']).alignment = center_align
            worksheet.cell(row=current_row, column=1).border = thin_border
            worksheet.cell(row=current_row, column=2, value=row_data['Kanwil']).alignment = left_align
            worksheet.cell(row=current_row, column=2).border = thin_border
            worksheet.cell(row=current_row, column=3, value=row_data['Target Setara Beras']).number_format = '#,##0.00'
            worksheet.cell(row=current_row, column=3).alignment = center_align
            worksheet.cell(row=current_row, column=3).border = thin_border
            worksheet.cell(row=current_row, column=4, value=row_data['Beras (a)']).number_format = '#,##0.00'
            worksheet.cell(row=current_row, column=4).alignment = center_align
            worksheet.cell(row=current_row, column=4).border = thin_border
            worksheet.cell(row=current_row, column=5, value=row_data['GKG (b)']).number_format = '#,##0.00'
            worksheet.cell(row=current_row, column=5).alignment = center_align
            worksheet.cell(row=current_row, column=5).border = thin_border
            worksheet.cell(row=current_row, column=6, value=row_data['GKP (c)']).number_format = '#,##0.00'
            worksheet.cell(row=current_row, column=6).alignment = center_align
            worksheet.cell(row=current_row, column=6).border = thin_border
            worksheet.cell(row=current_row, column=7, value=row_data['Setara Beras (d)']).number_format = '#,##0.00'
            worksheet.cell(row=current_row, column=7).alignment = center_align
            worksheet.cell(row=current_row, column=7).border = thin_border
            worksheet.cell(row=current_row, column=8, value=row_data['Capaian (%)']).number_format = '0.0"%"'
            worksheet.cell(row=current_row, column=8).alignment = center_align
            worksheet.cell(row=current_row, column=8).border = thin_border
            current_row += 1

        # Write Total Kanwil Lainnya
        worksheet.cell(row=current_row, column=1, value='Total Kanwil Lainnya').fill = lainnya_fill
        worksheet.cell(row=current_row, column=1).font = bold_font
        worksheet.cell(row=current_row, column=1).alignment = left_align
        worksheet.cell(row=current_row, column=1).border = thin_border
        worksheet.merge_cells(f'A{current_row}:B{current_row}')
        worksheet.cell(row=current_row, column=3, value=total_lainnya['Target Setara Beras']).number_format = '#,##0.00'
        worksheet.cell(row=current_row, column=3).fill = lainnya_fill
        worksheet.cell(row=current_row, column=3).font = bold_font
        worksheet.cell(row=current_row, column=3).alignment = center_align
        worksheet.cell(row=current_row, column=3).border = thin_border
        worksheet.cell(row=current_row, column=4, value=total_lainnya['Beras (a)']).number_format = '#,##0.00'
        worksheet.cell(row=current_row, column=4).fill = lainnya_fill
        worksheet.cell(row=current_row, column=4).font = bold_font
        worksheet.cell(row=current_row, column=4).alignment = center_align
        worksheet.cell(row=current_row, column=4).border = thin_border
        worksheet.cell(row=current_row, column=5, value=total_lainnya['GKG (b)']).number_format = '#,##0.00'
        worksheet.cell(row=current_row, column=5).fill = lainnya_fill
        worksheet.cell(row=current_row, column=5).font = bold_font
        worksheet.cell(row=current_row, column=5).alignment = center_align
        worksheet.cell(row=current_row, column=5).border = thin_border
        worksheet.cell(row=current_row, column=6, value=total_lainnya['GKP (c)']).number_format = '#,##0.00'
        worksheet.cell(row=current_row, column=6).fill = lainnya_fill
        worksheet.cell(row=current_row, column=6).font = bold_font
        worksheet.cell(row=current_row, column=6).alignment = center_align
        worksheet.cell(row=current_row, column=6).border = thin_border
        worksheet.cell(row=current_row, column=7, value=total_lainnya['Setara Beras (d)']).number_format = '#,##0.00'
        worksheet.cell(row=current_row, column=7).fill = lainnya_fill
        worksheet.cell(row=current_row, column=7).font = bold_font
        worksheet.cell(row=current_row, column=7).alignment = center_align
        worksheet.cell(row=current_row, column=7).border = thin_border
        worksheet.cell(row=current_row, column=8, value=capaian_lainnya).number_format = '0.0"%"'
        worksheet.cell(row=current_row, column=8).fill = lainnya_fill
        worksheet.cell(row=current_row, column=8).font = bold_font
        worksheet.cell(row=current_row, column=8).alignment = center_align
        worksheet.cell(row=current_row, column=8).border = thin_border
        current_row += 1

        # Write TOTAL SE-INDO
        worksheet.cell(row=current_row, column=1, value='TOTAL SE-INDO').fill = seindo_fill
        worksheet.cell(row=current_row, column=1).font = bold_font
        worksheet.cell(row=current_row, column=1).alignment = left_align
        worksheet.cell(row=current_row, column=1).border = thin_border
        worksheet.merge_cells(f'A{current_row}:B{current_row}')
        worksheet.cell(row=current_row, column=3, value=total_seindo['Target Setara Beras']).number_format = '#,##0.00'
        worksheet.cell(row=current_row, column=3).fill = seindo_fill
        worksheet.cell(row=current_row, column=3).font = bold_font
        worksheet.cell(row=current_row, column=3).alignment = center_align
        worksheet.cell(row=current_row, column=3).border = thin_border
        worksheet.cell(row=current_row, column=4, value=total_seindo['Beras (a)']).number_format = '#,##0.00'
        worksheet.cell(row=current_row, column=4).fill = seindo_fill
        worksheet.cell(row=current_row, column=4).font = bold_font
        worksheet.cell(row=current_row, column=4).alignment = center_align
        worksheet.cell(row=current_row, column=4).border = thin_border
        worksheet.cell(row=current_row, column=5, value=total_seindo['GKG (b)']).number_format = '#,##0.00'
        worksheet.cell(row=current_row, column=5).fill = seindo_fill
        worksheet.cell(row=current_row, column=5).font = bold_font
        worksheet.cell(row=current_row, column=5).alignment = center_align
        worksheet.cell(row=current_row, column=5).border = thin_border
        worksheet.cell(row=current_row, column=6, value=total_seindo['GKP (c)']).number_format = '#,##0.00'
        worksheet.cell(row=current_row, column=6).fill = seindo_fill
        worksheet.cell(row=current_row, column=6).font = bold_font
        worksheet.cell(row=current_row, column=6).alignment = center_align
        worksheet.cell(row=current_row, column=6).border = thin_border
        worksheet.cell(row=current_row, column=7, value=total_seindo['Setara Beras (d)']).number_format = '#,##0.00'
        worksheet.cell(row=current_row, column=7).fill = seindo_fill
        worksheet.cell(row=current_row, column=7).font = bold_font
        worksheet.cell(row=current_row, column=7).alignment = center_align
        worksheet.cell(row=current_row, column=7).border = thin_border
        worksheet.cell(row=current_row, column=8, value=capaian_seindo).number_format = '0.0"%"'
        worksheet.cell(row=current_row, column=8).fill = seindo_fill
        worksheet.cell(row=current_row, column=8).font = bold_font
        worksheet.cell(row=current_row, column=8).alignment = center_align
        worksheet.cell(row=current_row, column=8).border = thin_border

        # Set column widths
        worksheet.column_dimensions['A'].width = 5
        worksheet.column_dimensions['B'].width = 45
        worksheet.column_dimensions['C'].width = 18
        worksheet.column_dimensions['D'].width = 12
        worksheet.column_dimensions['E'].width = 12
        worksheet.column_dimensions['F'].width = 12
        worksheet.column_dimensions['G'].width = 18
        worksheet.column_dimensions['H'].width = 14

    output.seek(0)
    return output

def create_excel_export(df, tanggal_kemarin, tanggal_hari_ini):
    """Create Excel file with complex multi-row header and styling"""
    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Write empty dataframe to create workbook
        pd.DataFrame().to_excel(writer, sheet_name='Realisasi', index=False)

        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = workbook['Realisasi']

        # Define styles
        header_fill_dark = PatternFill(start_color='1f497d', end_color='1f497d', fill_type='solid')
        header_fill_light = PatternFill(start_color='4bacc6', end_color='4bacc6', fill_type='solid')
        total_fill = PatternFill(start_color='ddebf7', end_color='ddebf7', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        normal_font = Font(size=10)
        bold_font = Font(bold=True, size=10)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write header row 1
        headers_row1 = ['NO', 'KANWIL', 'TARGET', '', 'REALISASI SD. KEMARIN', '', 'REALISASI HARI INI', '', 'REALISASI SD. TGL', '', 'CAPAIAN', '']
        for col_idx, header in enumerate(headers_row1, start=1):
            cell = worksheet.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill_dark
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        # Merge cells for row 1
        worksheet.merge_cells('A1:A3')  # NO
        worksheet.merge_cells('B1:B3')  # KANWIL
        worksheet.merge_cells('C1:D2')  # TARGET
        worksheet.merge_cells('E1:F1')  # REALISASI SD KEMARIN
        worksheet.merge_cells('G1:H1')  # REALISASI HARI INI
        worksheet.merge_cells('I1:J1')  # REALISASI SD TGL
        worksheet.merge_cells('K1:L2')  # CAPAIAN

        # Write header row 2 (dates)
        worksheet.cell(row=2, column=5, value=tanggal_kemarin.strftime("%d %B %Y")).fill = header_fill_dark
        worksheet.cell(row=2, column=5).font = header_font
        worksheet.cell(row=2, column=5).alignment = center_align
        worksheet.cell(row=2, column=5).border = thin_border
        worksheet.merge_cells('E2:F2')

        worksheet.cell(row=2, column=7, value=tanggal_hari_ini.strftime("%d %B %Y")).fill = header_fill_dark
        worksheet.cell(row=2, column=7).font = header_font
        worksheet.cell(row=2, column=7).alignment = center_align
        worksheet.cell(row=2, column=7).border = thin_border
        worksheet.merge_cells('G2:H2')

        worksheet.cell(row=2, column=9, value=tanggal_hari_ini.strftime("%d %B %Y")).fill = header_fill_dark
        worksheet.cell(row=2, column=9).font = header_font
        worksheet.cell(row=2, column=9).alignment = center_align
        worksheet.cell(row=2, column=9).border = thin_border
        worksheet.merge_cells('I2:J2')

        # Write header row 3 (sub-headers)
        headers_row3 = ['', '', 'BERAS', 'GABAH', 'BERAS', 'GABAH', 'BERAS', 'GABAH', 'BERAS', 'GABAH', '% BERAS', '% GABAH']
        for col_idx, header in enumerate(headers_row3, start=1):
            if col_idx > 2:  # Skip NO and KANWIL
                cell = worksheet.cell(row=3, column=col_idx, value=header)
                cell.fill = header_fill_light
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border

        # Write data rows
        for idx, row in df.iterrows():
            excel_row = idx + 4  # Start from row 4 (after 3 header rows)
            is_total = row['Kanwil'] == 'TOTAL'

            # Write data
            worksheet.cell(row=excel_row, column=1, value=row['NO'])
            worksheet.cell(row=excel_row, column=2, value=row['Kanwil'])
            worksheet.cell(row=excel_row, column=3, value=row['Target Beras'] if pd.notna(row['Target Beras']) else None)
            worksheet.cell(row=excel_row, column=4, value=row['Target Gabah'] if pd.notna(row['Target Gabah']) else None)
            worksheet.cell(row=excel_row, column=5, value=row['Real SD Kemarin Beras'] if pd.notna(row['Real SD Kemarin Beras']) else None)
            worksheet.cell(row=excel_row, column=6, value=row['Real SD Kemarin Gabah'] if pd.notna(row['Real SD Kemarin Gabah']) else None)
            worksheet.cell(row=excel_row, column=7, value=row['Real Hari Ini Beras'] if pd.notna(row['Real Hari Ini Beras']) else None)
            worksheet.cell(row=excel_row, column=8, value=row['Real Hari Ini Gabah'] if pd.notna(row['Real Hari Ini Gabah']) else None)
            worksheet.cell(row=excel_row, column=9, value=row['Real SD Tgl Beras'] if pd.notna(row['Real SD Tgl Beras']) else None)
            worksheet.cell(row=excel_row, column=10, value=row['Real SD Tgl Gabah'] if pd.notna(row['Real SD Tgl Gabah']) else None)
            worksheet.cell(row=excel_row, column=11, value=row['Capaian Beras %'] if pd.notna(row['Capaian Beras %']) else None)
            worksheet.cell(row=excel_row, column=12, value=row['Capaian Gabah %'] if pd.notna(row['Capaian Gabah %']) else None)

            # Apply styling
            for col_idx in range(1, 13):
                cell = worksheet.cell(row=excel_row, column=col_idx)
                cell.border = thin_border
                cell.alignment = center_align if col_idx != 2 else left_align

                if is_total:
                    cell.fill = total_fill
                    cell.font = bold_font
                else:
                    cell.font = normal_font

                # Format numbers
                if col_idx >= 3 and col_idx <= 10:
                    cell.number_format = '#,##0.00'
                elif col_idx >= 11:
                    cell.number_format = '0.0"%"'

        # Set column widths
        worksheet.column_dimensions['A'].width = 5
        worksheet.column_dimensions['B'].width = 40
        for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
            worksheet.column_dimensions[col].width = 12

    output.seek(0)
    return output

def main():
    # Header with gradient background
    st.markdown("""
    <div class="dashboard-header">
        <h1 class="dashboard-title">Dashboard Pengadaan BULOG</h1>
        <p class="dashboard-subtitle">Sistem Informasi Realisasi Pengadaan Gabah & Beras</p>
    </div>
    """, unsafe_allow_html=True)

    # ===== SIDEBAR MENU =====
    with st.sidebar:
        st.markdown("### 📋 Menu")
        menu_option = st.radio(
            "Pilih Menu:",
            options=["📊 Dashboard Realisasi", "📁 Kelola Data"],
            label_visibility="collapsed"
        )

    # ===== MENU: KELOLA DATA =====
    if menu_option == "📁 Kelola Data":
        # CSS khusus untuk page Kelola Data - text berwarna gelap
        st.markdown("""
        <style>
            /* SUPER AGGRESSIVE: Override ALL text in main area - HIGHEST PRIORITY */
            .main,
            .main *,
            .main *::before,
            .main *::after {
                color: #1f497d !important;
            }
                    .st-emotion-cache-1s4g1qq {
                    background-color: #1f497d;
                    }
            /* radio button kelola data */
            .stMainBlockContainer.block-container.st-emotion-cache-zy6yx3.e4man114
            .stElementContainer.element-container.st-emotion-cache-zh2fnc.ek2vi381 {
            background-color: #1f497d;
                padding: 15px;
                border-radius: 10px;
            }
            .stElementContainer.element-container.st-emotion-cache-zh2fnc.ek2vi381 p{
                color: white;
            }

            /* Force ALL elements */
            div.main * {
                color: #1f497d !important;
            }

            /* Streamlit markdown containers */
            .main .stMarkdown,
            .main .stMarkdown *,
            .main [data-testid="stMarkdownContainer"],
            .main [data-testid="stMarkdownContainer"] * {
                color: #1f497d !important;
            }

            /* Headers - all levels with all possible selectors */
            .main h1, .main h2, .main h3, .main h4, .main h5, .main h6,
            .main .stMarkdown h1, .main .stMarkdown h2, .main .stMarkdown h3,
            .main .stMarkdown h4, .main .stMarkdown h5, .main .stMarkdown h6 {
                color: #1f497d !important;
            }

            /* Paragraphs and divs */
            .main p, .main span, .main div {
                color: #1f497d !important;
            }

            /* Element containers */
            .main .element-container,
            .main .element-container * {
                color: #1f497d !important;
            }

            /* Block container */
            .main [data-testid="stVerticalBlock"],
            .main [data-testid="stVerticalBlock"] * {
                color: #1f497d !important;
            }

            /* Info, warning, success, error boxes - ensure dark text */
            .stAlert, .stAlert *,
            div[data-testid="stAlert"],
            div[data-testid="stAlert"] *,
            .main .stAlert,
            .main .stAlert *,
            .main div[data-testid="stNotification"],
            .main div[data-testid="stNotification"] * {
                color: #1f497d !important;
            }

            /* Specific alert types */
            .stAlert p, .stAlert span, .stAlert div,
            div[data-testid="stAlert"] p,
            div[data-testid="stAlert"] span,
            div[data-testid="stAlert"] div {
                color: #1f497d !important;
            }

            /* Expander */
            .streamlit-expanderHeader, .streamlit-expanderHeader *,
            .main [data-testid="stExpander"],
            .main [data-testid="stExpander"] summary,
            .main [data-testid="stExpander"] summary *,
            .main details summary,
            .main details summary * {
                color: #1f497d !important;
            }

            /* Labels - all types */
            .main label, .main label *,
            .main .stRadio > label,
            .main .stRadio > label *,
            .main [data-baseweb="radio"] ~ *,
            .main .row-widget label {
                color: #1f497d !important;
            }

            /* Radio buttons - ULTRA AGGRESSIVE */
            .main .stRadio,
            .main .stRadio *,
            .main .stRadio label,
            .main .stRadio label *,
            .main .stRadio label span,
            .main .stRadio label span *,
            .main .stRadio label p,
            .main .stRadio label p *,
            .main .stRadio div,
            .main .stRadio div *,
            .main [role="radiogroup"],
            .main [role="radiogroup"] *,
            .main [role="radio"],
            .main [role="radio"] *,
            .main [role="radio"] + *,
            .main [role="radio"] ~ *,
            .main [data-baseweb="radio"],
            .main [data-baseweb="radio"] *,
            .main [data-baseweb="radio"] + *,
            .main [data-baseweb="radio"] ~ * {
                color: #1f497d !important;
            }

            /* File uploader */
            .main .stFileUploader label, .main .stFileUploader * {
                color: #1f497d !important;
            }

            /* Metrics */
            .main [data-testid="stMetric"] * {
                color: #1f497d !important;
            }

            /* Text input, number input, etc */
            .main .stTextInput label,
            .main .stNumberInput label,
            .main .stSelectbox label {
                color: #1f497d !important;
            }

            /* Checkbox */
            .main .stCheckbox label {
                color: #1f497d !important;
            }

            /* Progress bar */
            .main .stProgress,
            .main .stProgress *,
            .main div[data-testid="stProgress"],
            .main div[data-testid="stProgress"] * {
                color: #1f497d !important;
            }

            /* Status container */
            .main .stStatus,
            .main .stStatus *,
            .main .stSpinner,
            .main .stSpinner * {
                color: #1f497d !important;
            }

            /* Dataframe and table */
            .main .stDataFrame,
            .main .stDataFrame *,
            .main table,
            .main table * {
                color: #1f497d !important;
            }

            /* Code blocks */
            .main .stCodeBlock,
            .main .stCodeBlock *,
            .main code {
                color: #1f497d !important;
            }

            /* JSON display */
            .main .stJson,
            .main .stJson * {
                color: #1f497d !important;
            }

            /* EXCEPTIONS - Keep these WHITE */

            /* Sidebar - tetap putih */
            [data-testid="stSidebar"],
            [data-testid="stSidebar"] *,
            [data-testid="stSidebar"] h1,
            [data-testid="stSidebar"] h2,
            [data-testid="stSidebar"] h3,
            [data-testid="stSidebar"] p,
            [data-testid="stSidebar"] span,
            [data-testid="stSidebar"] div,
            [data-testid="stSidebar"] label {
                color: white !important;
            }

            /* Dashboard header - tetap putih */
            .dashboard-header,
            .dashboard-header * {
                color: white !important;
            }

            /* Buttons - keep white text */
            .main button,
            .main button *,
            .stDownloadButton button,
            .stDownloadButton button * {
                color: white !important;
            }
            .st-emotion-cache-11ofl8m {
                    background-color: #1f497d !important;
                    }

                    .stVerticalBlock .st-emotion-cache-tn0cau .ek2vi383{
                    background-color: #1f497d !important;
                    }

            .st-emotion-cache-fis6aj.e16n7gab7 {
                background-color: #1f497d !important;
                margin: 10px;
                padding: 1rem;
                border-radius: 20px;
            }
            .stElementContainer.element-container.st-key-select_dataframe.st-emotion-cache-zh2fnc.ek2vi381 {
                background-color: #1f497d !important;
                padding: 1rem;
                border-radius: 10px;
            }
                    .stProgress {
                color: black;
            }

        </style>

        <script>
            // JavaScript untuk memaksa warna dark pada semua element
            function forceTextColor() {
                const mainArea = document.querySelector('.main');
                if (mainArea) {
                    // Get all elements in main area
                    const allElements = mainArea.querySelectorAll('*');

                    allElements.forEach(el => {
                        // Skip sidebar, buttons, and dashboard header
                        const isSidebar = el.closest('[data-testid="stSidebar"]');
                        const isButton = el.tagName === 'BUTTON' || el.closest('button');
                        const isDashboardHeader = el.classList.contains('dashboard-header') || el.closest('.dashboard-header');

                        if (!isSidebar && !isButton && !isDashboardHeader) {
                            el.style.setProperty('color', '#1f497d', 'important');
                        }
                    });
                }
            }

            // Run immediately
            forceTextColor();

            // Run after a delay to catch dynamically loaded content
            setTimeout(forceTextColor, 100);
            setTimeout(forceTextColor, 300);
            setTimeout(forceTextColor, 500);
            setTimeout(forceTextColor, 1000);

            // Set up observer for dynamic content
            const observer = new MutationObserver(forceTextColor);
            const mainArea = document.querySelector('.main');
            if (mainArea) {
                observer.observe(mainArea, {
                    childList: true,
                    subtree: true,
                    attributes: true,
                    attributeFilter: ['style', 'class']
                });
            }
        </script>
        """, unsafe_allow_html=True)

        # Inject JavaScript using components
        import streamlit.components.v1 as st_components
        st_components.html("""
        <script>
            (function() {
                function forceTextColor() {
                    const mainArea = parent.document.querySelector('.main');
                    if (mainArea) {
                        const allElements = mainArea.querySelectorAll('*');
                        allElements.forEach(el => {
                            const isSidebar = el.closest('[data-testid="stSidebar"]');
                            const isButton = el.tagName === 'BUTTON' || el.closest('button');
                            const isDashboardHeader = el.classList.contains('dashboard-header') || el.closest('.dashboard-header');

                            if (!isSidebar && !isButton && !isDashboardHeader) {
                                el.style.setProperty('color', '#1f497d', 'important');
                            }
                        });
                    }
                }

                setTimeout(forceTextColor, 100);
                setTimeout(forceTextColor, 500);
                setTimeout(forceTextColor, 1000);
                setTimeout(forceTextColor, 2000);

                const observer = new MutationObserver(forceTextColor);
                const mainArea = parent.document.querySelector('.main');
                if (mainArea) {
                    observer.observe(mainArea, {
                        childList: true,
                        subtree: true
                    });
                }
            })();
        </script>
        """, height=0)

        # Container dengan background color untuk section Kelola Data
        st.markdown("""
            <style>
            .kelola-data-box {
                background-color: #f0f4f8;
                padding: 25px;
                border-radius: 10px;
                margin-bottom: 20px;
            }

            /* Light theme untuk radio button di dalam kelola-data-box */
            .kelola-data-box [data-testid="stRadio"] label {
                background-color: white !important;
                color: #1f497d !important;
                padding: 10px 20px !important;
                border-radius: 8px !important;
                border: 2px solid #d0d0d0 !important;
                transition: all 0.3s ease !important;
            }
                    
            .st-emotion-cache-1k9kca4 {
                background-color: #1f497d !important;
            }
            
            .st-emotion-cache-zh4rd8 {
                    background-color: #1f497d !important;
                    }
                    
            .st-emotion-cache-1weic72{
                    color: black !important;
                    }
            .stElementContainer.element-container.st-key-select_table.st-emotion-cache-zh2fnc.ek2vi381 {
                background-color: #1f497d !important;
                    padding: 15px;
                    border-radius: 10px;
                }


            .kelola-data-box [data-testid="stRadio"] label:hover {
                background-color: #1f497d !important;
            }

            .kelola-data-box [data-testid="stRadio"] label[data-checked="true"] {
                background-color: #1f497d !important;
                color: white !important;
                border-color: #1f497d !important;
            }

            /* Style untuk radio circle */
            .kelola-data-box [data-testid="stRadio"] input[type="radio"] {
                accent-color: #1f497d !important;
            }

            /* Text warna gelap untuk radio options */
            .kelola-data-box [data-testid="stRadio"] p {
                color: #1f497d !important;
            }
            .st-emotion-cache-1aq61ou {
                font-size: 16px !important;
            }
            </style>
        """, unsafe_allow_html=True)
        # st.markdown('<h3 style="color: #1f497d;">📊 Pilih Tabel yang Akan Dikelola</h3>', unsafe_allow_html=True)
        st.markdown('<div class="chart-title" style="font-size: 30px; font-weight: bold;">📋 Kelola Data</div>', unsafe_allow_html=True)

        form1, form2 = st.columns(2)
        with form1:
            uploaded_file = st.file_uploader(
                "📗 Upload File Excel (Sebelum Uplod Pilih Data yang Ingin Dikelola)",
                type=['xlsx', 'xls'],
                help="Upload file Excel untuk update data ke database"
            )

            # Tombol untuk clear cache jika user ingin upload file baru
            if 'validated_file_key' in st.session_state and uploaded_file is not None:

                if st.button("🔄 Reset & Upload File Baru", use_container_width=True, help="Hapus cache file saat ini untuk upload file baru"):
                    # Clear all file-related cache
                    if 'validated_file_key' in st.session_state:
                        del st.session_state.validated_file_key
                    if 'available_sheets' in st.session_state:
                        del st.session_state.available_sheets
                    if 'loaded_sheet_key' in st.session_state:
                        del st.session_state.loaded_sheet_key
                    if 'df_new' in st.session_state:
                        del st.session_state.df_new
                    st.success("✅ Cache berhasil dihapus! Silakan upload file baru.")
                    time.sleep(1)
                    st.rerun()

        with form2:
            st.markdown('<p style="color: black; font-size: 16px; margin-bottom: -1px;">📂 Pilih Tabel</p>', unsafe_allow_html=True)
            selected_table = st.radio(
                ".",
                options=["📈 Realisasi", "🎯 Target Kanwil", "🏢 Target Kancab"],
                help="Pilih tabel mana yang akan diupdate/replace",
                horizontal=True,
                key="select_table",
                label_visibility="collapsed"
            )

            table_map = {
                "📈 Realisasi": ("realisasi", "Export"),
                "🎯 Target Kanwil": ("target_kanwil", "Target Kanwil"),
                "🏢 Target Kancab": ("target_kancab", "Target Kancab")
            }
            table_name, expected_sheet_name = table_map[selected_table]
        

        if uploaded_file is not None:
            try:
                # Generate unique key untuk file yang di-upload
                file_key = f"{uploaded_file.name}_{uploaded_file.size}_{table_name}"

                # Cek apakah file sudah pernah divalidasi
                if 'validated_file_key' not in st.session_state or st.session_state.validated_file_key != file_key:
                    # File baru atau berbeda - lakukan validasi
                    progress_bar = st.progress(0, "🔍 Memvalidasi file Excel...")

                    # Read Excel file to get sheet names
                    excel_file = pd.ExcelFile(uploaded_file, engine='openpyxl')
                    available_sheets = excel_file.sheet_names
                    progress_bar.progress(20, "📋 Mendeteksi sheet yang tersedia...")

                    # Simpan ke session state
                    st.session_state.available_sheets = available_sheets
                    st.session_state.validated_file_key = file_key

                    progress_bar.progress(40, "✅ Validasi selesai")
                    progress_bar.empty()
                else:
                    # File sudah divalidasi, ambil dari session state
                    available_sheets = st.session_state.available_sheets

                # Pilihan sheet jika ada lebih dari satu
                if len(available_sheets) > 1:
                    st.info(f"📋 File Excel memiliki {len(available_sheets)} sheet: {', '.join(available_sheets)}")
                    selected_sheet = st.selectbox(
                        "Pilih sheet yang akan diproses:",
                        options=available_sheets,
                        index=available_sheets.index(expected_sheet_name) if expected_sheet_name in available_sheets else 0,
                        key="select_sheet"
                    )
                else:
                    selected_sheet = available_sheets[0]
                    st.info(f"📋 Menggunakan sheet: **{selected_sheet}**")

                # Generate key untuk sheet yang dipilih
                sheet_key = f"{file_key}_{selected_sheet}"

                # Cek apakah data dari sheet ini sudah pernah dibaca
                if 'loaded_sheet_key' not in st.session_state or st.session_state.loaded_sheet_key != sheet_key:
                    # Sheet baru atau berbeda - baca data
                    progress_bar = st.progress(0, f"📖 Membaca data dari sheet '{selected_sheet}'...")

                    # Read with preservasi presisi numeric (seperti di migrate script)
                    dtype_map = {}
                    if table_name == "realisasi":
                        dtype_map = {
                            'Kuantum PO (Kg)': str,
                            'In / Out': str,
                            'Harga Include ppn': str,
                            'Nominal Realisasi Incl ppn': str
                        }

                    df_new = pd.read_excel(uploaded_file, sheet_name=selected_sheet, engine='openpyxl', dtype=dtype_map)

                    # Simpan ke session state
                    st.session_state.df_new = df_new
                    st.session_state.loaded_sheet_key = sheet_key

                    progress_bar.progress(100, "✅ Data berhasil dibaca")
                    progress_bar.empty()

                    st.success(f"✅ File berhasil dibaca dari sheet **'{selected_sheet}'**: **{len(df_new):,}** records")
                else:
                    # Data sudah dibaca, ambil dari session state
                    df_new = st.session_state.df_new
                    st.info(f"ℹ️ Menggunakan data dari cache: **{len(df_new):,}** records dari sheet **'{selected_sheet}'**")

                # Show preview
                st.markdown('<p style="color: #1f497d; font-weight: 600; margin-bottom: 0;">👁️ Preview Data Baru</p>', unsafe_allow_html=True)
                with st.expander("", expanded=False):
                    st.dataframe(df_new.head(15), use_container_width=True)

                # Initialize log container in session state
                if 'process_logs' not in st.session_state:
                    st.session_state.process_logs = []

                # Mode selection
                st.markdown('<h4 style="color: #1f497d;">⚙️ Mode Upload</h4>', unsafe_allow_html=True)
                upload_mode = st.radio(
                    "Pilih mode upload:",
                    options=["🔄 Append (Tambahkan data baru)", "🔁 Replace (Ganti semua data)"],
                    help="Append: Tambahkan hanya data unik ke database | Replace: Hapus semua data lama dan ganti dengan data baru",
                    horizontal=True
                )
                if upload_mode == "🔄 Append (Tambahkan data baru)":
                    st.info("""
                    **Mode Append:**
                    - Data baru akan dibandingkan dengan data di database
                    - Hanya data **unik** yang akan ditambahkan ke database secara bertahap
                    - Data existing tetap aman
                    """)

                    # Add button to start append process
                    st.markdown("---")
                    if not st.button("▶️ Mulai Proses Append", type="primary", use_container_width=True, key="start_append"):
                        st.info("👆 Klik tombol di atas untuk memulai proses append data")
                        st.stop()

                    # Process started
                    add_log("="*60, "info")
                    add_log(f"🚀 APPEND MODE STARTED - Table: {table_name}", "info")
                    add_log(f"📊 Total records from Excel: {len(df_new):,}", "info")
                    add_log("="*60, "info")

                    # Load mapping IDs
                    st.markdown("---")
                    st.markdown('<h4 style="color: #1f497d;">📋 Loading Mapping IDs</h4>', unsafe_allow_html=True)

                    kanwil_map = {}
                    kancab_map = {}

                    add_log("🔄 Loading mapping Kanwil dari database...", "info")
                    st.info("🔄 Loading mapping Kanwil dari database...")
                    kanwil_result = supabase.table('kanwil').select('*').execute()
                    for kw in kanwil_result.data:
                        kanwil_map[kw['nama_kanwil']] = kw['kanwil_id']
                    add_log(f"✅ Loaded {len(kanwil_map)} Kanwil mappings", "success")
                    st.success(f"✅ Loaded {len(kanwil_map)} Kanwil mappings")

                    add_log("🔄 Loading mapping Kancab dari database...", "info")
                    st.info("🔄 Loading mapping Kancab dari database...")
                    kancab_result = supabase.table('kancab').select('*').execute()
                    for kc in kancab_result.data:
                        kancab_map[kc['nama_kancab']] = kc['kancab_id']
                    add_log(f"✅ Loaded {len(kancab_map)} Kancab mappings", "success")
                    st.success(f"✅ Loaded {len(kancab_map)} Kancab mappings")

                    # Use NEW COMPARISON ALGORITHM for realisasi
                    if table_name == "realisasi":
                        st.markdown("---")
                        st.markdown('<h4 style="color: #1f497d;">🔄 Using NEW COMPARISON ALGORITHM (APPEND MODE)</h4>', unsafe_allow_html=True)

                        # Step 1: Migrate to realisasi_compare
                        st.info("📥 Step 1: Migrating Excel data to realisasi_compare...")

                        # Prepare kancab mapping with kanwil
                        kancab_mapping_full = {}
                        kancab_result_full = supabase.table('kancab').select('*, kanwil!inner(nama_kanwil)').execute()
                        for k in kancab_result_full.data:
                            key = (k['kanwil']['nama_kanwil'], k['nama_kancab'])
                            kancab_mapping_full[key] = k['kancab_id']

                        total_inserted, skipped_kanwil, skipped_kancab = migrate_to_realisasi_compare_streamlit(
                            supabase, df_new, kanwil_map, kancab_mapping_full, kancab_column='Entitas'
                        )

                        # Step 2: Compare using RPC
                        comparison_results = process_comparison_with_rpc_streamlit(supabase)

                        # Step 3: Migrate unique data to realisasi
                        if comparison_results:
                            num_unique = migrate_from_compare_to_realisasi_streamlit(supabase, comparison_results)

                            # Step 4: Cleanup realisasi_compare
                            add_log("🗑️ Step 4: Cleaning up realisasi_compare...", "info")
                            st.info("🗑️ Step 4: Cleaning up realisasi_compare...")
                            try:
                                truncate_table_with_reset(supabase, "realisasi_compare")
                                st.success("✅ realisasi_compare table cleaned up")
                            except Exception as e:
                                add_log(f"⚠️ Error cleaning up realisasi_compare: {e}", "warning")
                                st.warning(f"⚠️ Error cleaning up realisasi_compare: {e}")

                            # Set unique_data for display (empty since we already migrated)
                            unique_data = []
                            num_duplicates = len(df_new) - len(comparison_results)

                            # Log final summary
                            add_log("="*60, "success")
                            add_log(f"✅ APPEND PROCESS COMPLETED!", "success")
                            add_log(f"📊 Total Excel records: {len(df_new):,}", "info")
                            add_log(f"✅ Unique records added: {num_unique:,}", "success")
                            add_log(f"⚠️ Duplicate records skipped: {num_duplicates:,}", "warning")
                            add_log("="*60, "success")
                        else:
                            st.info("✅ No unique data to migrate")
                            num_unique = 0
                            num_duplicates = len(df_new)
                            unique_data = []

                            # Log final summary
                            add_log("="*60, "info")
                            add_log(f"ℹ️ APPEND PROCESS COMPLETED - No new data", "info")
                            add_log(f"📊 All {len(df_new):,} records already exist in database", "info")
                            add_log("="*60, "info")

                    else:
                        # For target tables, use NEW COMPARISON ALGORITHM (same as realisasi)
                        st.markdown("---")
                        st.markdown('<h4 style="color: #1f497d;">🔄 Using NEW COMPARISON ALGORITHM (APPEND MODE)</h4>', unsafe_allow_html=True)

                        if table_name == "target_kanwil":
                            # Step 1: Migrate to target_kanwil_compare
                            st.info("📥 Step 1: Migrating Excel data to target_kanwil_compare...")

                            total_inserted, skipped_kanwil = migrate_to_target_kanwil_compare_streamlit(
                                supabase, df_new, kanwil_map
                            )

                            # Step 2: Compare using RPC
                            comparison_results = process_comparison_target_kanwil_with_rpc_streamlit(supabase)

                            # Step 3: Migrate unique data to target_kanwil
                            if comparison_results:
                                num_unique = migrate_from_target_kanwil_compare_to_target_kanwil_streamlit(supabase, comparison_results)

                                # Step 4: Cleanup target_kanwil_compare
                                add_log("🗑️ Step 4: Cleaning up target_kanwil_compare...", "info")
                                st.info("🗑️ Step 4: Cleaning up target_kanwil_compare...")
                                try:
                                    truncate_table_with_reset(supabase, "target_kanwil_compare")
                                    st.success("✅ target_kanwil_compare table cleaned up")
                                except Exception as e:
                                    add_log(f"⚠️ Error cleaning up target_kanwil_compare: {e}", "warning")
                                    st.warning(f"⚠️ Error cleaning up target_kanwil_compare: {e}")

                                # Set unique_data for display (empty since we already migrated)
                                unique_data = []
                                num_duplicates = len(df_new) - len(comparison_results)

                                # Log final summary
                                add_log("="*60, "success")
                                add_log(f"✅ APPEND PROCESS COMPLETED!", "success")
                                add_log(f"📊 Total Excel records: {len(df_new):,}", "info")
                                add_log(f"✅ Unique records added: {num_unique:,}", "success")
                                add_log(f"⚠️ Duplicate records skipped: {num_duplicates:,}", "warning")
                                add_log("="*60, "success")
                            else:
                                st.info("✅ No unique data to migrate")
                                num_unique = 0
                                num_duplicates = len(df_new)
                                unique_data = []

                                # Log final summary
                                add_log("="*60, "info")
                                add_log(f"ℹ️ APPEND PROCESS COMPLETED - No new data", "info")
                                add_log(f"📊 All {len(df_new):,} records already exist in database", "info")
                                add_log("="*60, "info")

                        else:  # target_kancab
                            # Step 1: Migrate to target_kancab_compare
                            st.info("📥 Step 1: Migrating Excel data to target_kancab_compare...")

                            total_inserted, skipped_kancab = migrate_to_target_kancab_compare_streamlit(
                                supabase, df_new, kancab_map
                            )

                            # Step 2: Compare using RPC
                            comparison_results = process_comparison_target_kancab_with_rpc_streamlit(supabase)

                            # Step 3: Migrate unique data to target_kancab
                            if comparison_results:
                                num_unique = migrate_from_target_kancab_compare_to_target_kancab_streamlit(supabase, comparison_results)

                                # Step 4: Cleanup target_kancab_compare
                                add_log("🗑️ Step 4: Cleaning up target_kancab_compare...", "info")
                                st.info("🗑️ Step 4: Cleaning up target_kancab_compare...")
                                try:
                                    truncate_table_with_reset(supabase, "target_kancab_compare")
                                    st.success("✅ target_kancab_compare table cleaned up")
                                except Exception as e:
                                    add_log(f"⚠️ Error cleaning up target_kancab_compare: {e}", "warning")
                                    st.warning(f"⚠️ Error cleaning up target_kancab_compare: {e}")

                                # Set unique_data for display (empty since we already migrated)
                                unique_data = []
                                num_duplicates = len(df_new) - len(comparison_results)

                                # Log final summary
                                add_log("="*60, "success")
                                add_log(f"✅ APPEND PROCESS COMPLETED!", "success")
                                add_log(f"📊 Total Excel records: {len(df_new):,}", "info")
                                add_log(f"✅ Unique records added: {num_unique:,}", "success")
                                add_log(f"⚠️ Duplicate records skipped: {num_duplicates:,}", "warning")
                                add_log("="*60, "success")
                            else:
                                st.info("✅ No unique data to migrate")
                                num_unique = 0
                                num_duplicates = len(df_new)
                                unique_data = []

                                # Log final summary
                                add_log("="*60, "info")
                                add_log(f"ℹ️ APPEND PROCESS COMPLETED - No new data", "info")
                                add_log(f"📊 All {len(df_new):,} records already exist in database", "info")
                                add_log("="*60, "info")

                    # Show metrics for all tables
                    st.markdown("---")
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("📊 Total Data Baru", f"{len(df_new):,}")
                    with col2:
                        st.metric("✅ Data Unik", f"{num_unique:,}")
                    with col3:
                        st.metric("⚠️ Data Duplikat", f"{num_duplicates:,}")

                    if num_unique > 0:
                        # For all tables using NEW ALGORITHM, show summary (already migrated)
                        st.success(f"""
                        ✅ **Append berhasil!**
                        - Tabel: **{selected_table}**
                        - Data berhasil ditambahkan: **{num_unique:,}** records
                        - Data duplikat (diabaikan): **{num_duplicates:,}** records
                        """)
                        st.balloons()
                        st.info("🔄 Refresh halaman untuk melihat data terbaru")
                    else:
                        st.warning("⚠️ Tidak ada data unik untuk ditambahkan.")

                else:  # Replace mode
                    st.warning(f"""
                    **⚠️ Mode Replace:**
                    - Semua data di tabel **{table_name}** akan **DIHAPUS**
                    - Data baru akan menggantikan sepenuhnya
                    - **TIDAK BISA dibatalkan!**
                    """)

                    st.markdown("---")
                    st.markdown('<h4 style="color: #1f497d;">📊 Debug: Info Data untuk Replace</h4>', unsafe_allow_html=True)

                    col1, col2 = st.columns(2)
                    with col1:
                        try:
                            st.info(f"🔄 Checking existing data count di tabel '{table_name}'...")
                            existing_count = supabase.table(table_name).select('*', count='exact').execute()
                            st.metric("📊 Data di Database", f"{existing_count.count:,} records")
                            st.success(f"✅ Loaded count dari database: {existing_count.count:,}")
                        except Exception as e:
                            st.metric("📊 Data di Database", "Unknown")
                            st.error(f"❌ Error getting count: {str(e)}")
                    with col2:
                        st.metric("🆕 Data Baru dari Excel", f"{len(df_new):,} records")
                        st.info(f"📋 Data sudah di-load dari sheet '{selected_sheet}'")

                    # Confirmation checkbox
                    confirm_replace = st.checkbox(
                        f"⚠️ Saya mengerti bahwa ini akan menghapus SEMUA data di tabel {table_name}",
                        value=False
                    )

                    st.markdown("---")
                    if st.button(
                        "🔁 Replace Semua Data",
                        type="primary",
                        disabled=not confirm_replace,
                        use_container_width=True
                    ):
                        try:
                            import time
                            start_time = time.time()

                            # Log start
                            add_log("="*60, "warning")
                            add_log(f"🔁 REPLACE MODE STARTED - Table: {table_name}", "warning")
                            add_log(f"⚠️ ALL existing data will be DELETED!", "warning")
                            add_log(f"📊 Total records from Excel: {len(df_new):,}", "info")
                            add_log("="*60, "warning")

                            print("=" * 80)
                            print(f"REPLACE MODE - {table_name.upper()}")
                            print("=" * 80)

                            replace_progress = st.progress(0, "🔄 Memulai proses replace...")

                            # Step 1: Get mapping IDs
                            replace_progress.progress(10, "📋 Loading mapping IDs dari database...")
                            print("[STEP 1] Loading mapping IDs...")
                            add_log("📋 Step 1: Loading mapping IDs...", "info")
                            kanwil_map = {}
                            kancab_map = {}

                            kanwil_result = supabase.table('kanwil').select('*').execute()
                            for kw in kanwil_result.data:
                                kanwil_map[kw['nama_kanwil']] = kw['kanwil_id']
                            add_log(f"✅ Loaded {len(kanwil_map)} Kanwil mappings", "success")
                            st.info(f"✅ Loaded {len(kanwil_map)} Kanwil mappings dari database")
                            print(f"[STEP 1] Loaded {len(kanwil_map)} Kanwil mappings")

                            kancab_result = supabase.table('kancab').select('*').execute()
                            for kc in kancab_result.data:
                                kancab_map[kc['nama_kancab']] = kc['kancab_id']
                            add_log(f"✅ Loaded {len(kancab_map)} Kancab mappings", "success")
                            st.info(f"✅ Loaded {len(kancab_map)} Kancab mappings dari database")
                            print(f"[STEP 1] Loaded {len(kancab_map)} Kancab mappings")

                            # Step 2-4: Use NEW COMPARISON ALGORITHM for realisasi
                            if table_name == "realisasi":
                                # Use migrate_to_realisasi_direct_streamlit (same as new_comparison_algorithm.py)
                                replace_progress.progress(20, "🔧 Using NEW COMPARISON ALGORITHM (REPLACE MODE)...")
                                print(f"[STEP 2-4] Using migrate_to_realisasi_direct for {len(df_new):,} records")

                                # Prepare kancab mapping with kanwil
                                kancab_mapping_full = {}
                                kancab_result_full = supabase.table('kancab').select('*, kanwil!inner(nama_kanwil)').execute()
                                for k in kancab_result_full.data:
                                    key = (k['kanwil']['nama_kanwil'], k['nama_kancab'])
                                    kancab_mapping_full[key] = k['kancab_id']

                                # Call the new function
                                inserted_total, skipped_kanwil, skipped_kancab = migrate_to_realisasi_direct_streamlit(
                                    supabase, df_new, kanwil_map, kancab_mapping_full, kancab_column='Entitas'
                                )
                                failed_total = 0

                            elif table_name == "target_kanwil":
                                # Use migrate_to_target_kanwil_direct_streamlit (same as new_comparison_algorithm.py)
                                replace_progress.progress(20, "🔧 Using NEW COMPARISON ALGORITHM (REPLACE MODE)...")
                                print(f"[STEP 2-4] Using migrate_to_target_kanwil_direct for {len(df_new):,} records")

                                # Call the new function
                                inserted_total, skipped_kanwil = migrate_to_target_kanwil_direct_streamlit(
                                    supabase, df_new, kanwil_map
                                )
                                failed_total = 0

                            else:  # target_kancab
                                # Use migrate_to_target_kancab_direct_streamlit (same as new_comparison_algorithm.py)
                                replace_progress.progress(20, "🔧 Using NEW COMPARISON ALGORITHM (REPLACE MODE)...")
                                print(f"[STEP 2-4] Using migrate_to_target_kancab_direct for {len(df_new):,} records")

                                # Call the new function
                                inserted_total, skipped_kancab = migrate_to_target_kancab_direct_streamlit(
                                    supabase, df_new, kancab_map
                                )
                                failed_total = 0

                            replace_progress.progress(100, "✅ Replace selesai!")
                            replace_progress.empty()

                            total_elapsed = time.time() - start_time
                            print("-" * 80)
                            print(f"REPLACE COMPLETE - Total time: {total_elapsed:.2f} seconds ({total_elapsed/60:.2f} minutes)")
                            print(f"Successfully inserted: {inserted_total:,} records")
                            print(f"Failed: {failed_total:,} records")
                            print(f"Average: {inserted_total/total_elapsed:.0f} records/second")
                            print("=" * 80)

                            # Log final summary
                            add_log("="*60, "success")
                            add_log(f"✅ REPLACE PROCESS COMPLETED!", "success")
                            add_log(f"⏱️ Total time: {total_elapsed:.2f}s ({total_elapsed/60:.2f} min)", "info")
                            add_log(f"✅ Successfully inserted: {inserted_total:,} records", "success")
                            add_log(f"❌ Failed: {failed_total:,} records", "error" if failed_total > 0 else "info")
                            add_log(f"⚡ Average speed: {inserted_total/total_elapsed if total_elapsed > 0 else 0:.0f} records/second", "info")
                            add_log("="*60, "success")

                            st.success(f"""
                            ✅ **Replace berhasil!**
                            - Tabel: **{selected_table}**
                            - Data baru disimpan: **{inserted_total:,}** records
                            - Data gagal: **{failed_total:,}** records
                            - Waktu: **{total_elapsed:.2f}** detik
                            """)

                            st.balloons()
                            st.info("🔄 Refresh halaman untuk melihat data terbaru")

                        except Exception as e:
                            add_log(f"❌ FATAL ERROR during replace: {str(e)}", "error")
                            st.error(f"❌ Error saat replace data: {str(e)}")
                            with st.expander("🔍 Detail Error"):
                                st.code(str(e))
                                st.code(traceback.format_exc())

            except Exception as e:
                st.error(f"❌ Error saat membaca file Excel: {str(e)}")
                st.info("Pastikan file Excel memiliki sheet 'Export' dengan format yang benar")

        return  # Exit early, don't show dashboard

    # ===== MENU: DASHBOARD REALISASI (DEFAULT) =====
    # Data sekarang diambil langsung dari database via RPC
    # Tidak perlu load dari Excel lagi

    colA, colB, colC = st.columns(3)
    st.markdown("""
        <style>
            h4 {
                color: #1f497d !important;
                }

            /* Checkbox/Radio selected state */
            div .st-am{
                background-color: #1f497d !important;
                }

            span .st-ae{
                background-color: #1f497d !important;
                }

            /* Judul Filter Biru */
            .filter-title {
                color: #1f497d !important;
                font-weight: 600 !important;
            }

            /* Label text dalam filter */
            label, .stSelectbox label, .stMultiSelect label, .stDateInput label {
                color: #1f497d !important;
            }

            /* Placeholder text */
            ::placeholder {
                color: #9dc3e6 !important;
            }

            /* Styling Multiselect & Date Input - Light Theme */
            div[data-baseweb="select"] > div {
                border: 1px solid #1f497d !important;
                box-shadow: none !important;
                background-color: white !important;
            }

            div[data-baseweb="select"] > div:hover {
                border-color: #4bacc6 !important;
                background-color: #f8f9fa !important;
            }

            /* Text input & date input border biru - Light theme */
            input[type="text"], input[type="date"] {
                border: 1px solid #1f497d !important;
                color: #1f497d !important;
                background-color: white !important;
            }

            input[type="text"]:focus, input[type="date"]:focus {
                border-color: #4bacc6 !important;
                background-color: #f8f9fa !important;
            }

            /* Arrow icon warna biru */
            svg {
                color: #1f497d !important;
            }

            /* Text di dalam dropdown - Light theme */
            .css-1n7v3ny-option, .css-1n7v3ny-option * {
                color: #1f497d !important;
                background-color: white !important;
            }

            .css-1n7v3ny-option:hover {
                background-color: #f0f7fa !important;
            }

            /* Multiselect selected items */
            div[data-baseweb="tag"] {
                background-color: #1f497d !important;
                color: white !important;
            }

            /* Multiselect dropdown list */
            [role="listbox"] {
                background-color: white !important;
            }

            [role="option"] {
                color: #1f497d !important;
                background-color: white !important;
            }

            [role="option"]:hover {
                background-color: #f0f7fa !important;
            }
        </style>
        """, unsafe_allow_html=True)


    # Filter 1: Akun Analitik
    with colA:
        st.markdown("#### Akun Analitik")
        # Daftar akun analitik (hardcoded, bisa diambil dari database jika diperlukan)
        all_akun_analitik = ["CBP", "PSO"]

        selected_akun_analitik = st.multiselect(
            "Pilih Akun Analitik:",
            options=all_akun_analitik,
            default=["PSO"],     # default hanya PSO
            label_visibility="collapsed",
            key="filter_akun_analitik"
        )

    # Filter 2: Kanwil (untuk Line Chart & Tabel Kancab)
    with colB:
        st.markdown("#### Kanwil")
        # Daftar kanwil (fixed 26 kanwil)
        all_kanwil = [
            "01001 - KANTOR WILAYAH ACEH",
            "02001 - KANTOR WILAYAH SUMUT",
            "03001 - KANTOR WILAYAH RIAU DAN KEPRI",
            "04001 - KANTOR WILAYAH SUMBAR",
            "05001 - KANTOR WILAYAH JAMBI",
            "06001 - KANTOR WILAYAH SUMSEL",
            "07001 - KANTOR WILAYAH BENGKULU",
            "08001 - KANTOR WILAYAH LAMPUNG",
            "09001 - KANTOR WILAYAH DKI JAKARTA BANTEN",
            "10001 - KANTOR WILAYAH JABAR",
            "11001 - KANTOR WILAYAH JATENG",
            "12001 - KANTOR WILAYAH DI YOGYAKARTA",
            "13001 - KANTOR WILAYAH JATIM",
            "14001 - KANTOR WILAYAH KALBAR",
            "15001 - KANTOR WILAYAH KALTIM KALTARA",
            "16001 - KANTOR WILAYAH KALSEL",
            "17001 - KANTOR WILAYAH KALTENG",
            "18001 - KANTOR WILAYAH SULUT GORONTALO",
            "19001 - KANTOR WILAYAH SULTENG",
            "20001 - KANTOR WILAYAH SULTRA",
            "21001 - KANTOR WILAYAH SULSEL SULBAR",
            "22001 - KANTOR WILAYAH BALI",
            "23001 - KANTOR WILAYAH N.T.B",
            "24001 - KANTOR WILAYAH N.T.T",
            "25001 - KANTOR WILAYAH MALUKU MALUT",
            "26001 - KANTOR WILAYAH PAPUA PABAR"
        ]

        selected_kanwil = st.multiselect(
            "Pilih Kanwil:",
            options=all_kanwil,
            default=["13001 - KANTOR WILAYAH JATIM"],
            label_visibility="collapsed",
            key="filter_kanwil"
        )

    # Filter 3: Date Range (UNTUK SEMUA VISUALISASI)
    with colC:
        st.markdown("#### Periode")

        # Default range: awal tahun 2025 sampai hari ini
        min_date_penerimaan = datetime(2025, 1, 1).date()
        max_date_penerimaan = datetime.now().date()

        date_range = st.date_input(
            "Range Tanggal Penerimaan:",
            value=(min_date_penerimaan, max_date_penerimaan),
            min_value=min_date_penerimaan,
            max_value=max_date_penerimaan,
            label_visibility="collapsed",
            key="filter_date_range"
        )

    # Handle date_range untuk mendapatkan start_date dan end_date
    if isinstance(date_range, tuple) and len(date_range) == 2:
        start_date, end_date = date_range
    elif isinstance(date_range, tuple) and len(date_range) == 1:
        start_date = end_date = date_range[0]
    else:
        # Single date selected
        start_date = end_date = date_range

    # Display metrics - MENGGUNAKAN RPC
    col1, col2, col3, col4 = st.columns(4)

    # Ambil data metric card dari RPC
    # Prepare parameters
    # PENTING: p_nama_kanwil harus string tunggal, bukan list
    # Jika user memilih multiple kanwil, ambil yang pertama saja untuk metric card
    # Jika tidak ada yang dipilih, gunakan None (semua kanwil)
    p_nama_kanwil = selected_kanwil[0] if selected_kanwil and len(selected_kanwil) > 0 else None
    p_akun_analitik = selected_akun_analitik[0] if selected_akun_analitik and len(selected_akun_analitik) > 0 else None
    p_start_date = start_date.strftime('%Y-%m-%d') if hasattr(start_date, 'strftime') else str(start_date)
    p_end_date = end_date.strftime('%Y-%m-%d') if hasattr(end_date, 'strftime') else str(end_date)
    p_today = p_end_date

    metric_data = get_metric_card_data(p_nama_kanwil, p_akun_analitik, p_start_date, p_end_date, p_today)

    # Metric 1: Realisasi Setara Beras Hari Ini (end_date only)
    with col1:
        realisasi_hari_ini = metric_data.get('total_setara_beras_hari_ini', 0.0)
        st.metric(f"📊 Realisasi Hari Ini ({end_date})", f"{realisasi_hari_ini:,.2f} Ton")

    # Metric 2: Total Realisasi Setara Beras (start_date - end_date)
    with col2:
        total_realisasi = metric_data.get('total_setara_beras_rentang', 0.0)
        st.metric(f"📈 Total Realisasi ({start_date}-{end_date})", f"{total_realisasi:,.2f} Ton")

    # Metric 3: Target Setara Beras (Kanwil)
    with col3:
        target_setara_beras = metric_data.get('target_setara_beras', 0.0)
        kanwil_label = p_nama_kanwil if p_nama_kanwil else "Semua Kanwil"
        st.metric(f"🎯 Target {kanwil_label}", f"{target_setara_beras:,.2f} Ton")

    # Metric 4: Sisa Target Setara Beras (Kanwil)
    with col4:
        raw_sisa_target = metric_data.get('sisa_target', 0.0)

        if raw_sisa_target < 0:
            display_value = f"+{abs(raw_sisa_target):,.2f}"
        else:
            display_value = f"{raw_sisa_target:,.2f}"

        st.metric(
            f"📉 Sisa Target {kanwil_label}",
            f"{display_value} Ton"
        )

    # ===== TABEL SUMMARY (DI ATAS LINE CHART) - MENGGUNAKAN RPC =====
    st.markdown('<div class="chart-title">📋 Tabel Realisasi per-Kanwil</div>', unsafe_allow_html=True)

    # Ambil data dari RPC
    data_sentra, data_lainnya = create_summary_table_from_rpc(p_akun_analitik, p_start_date, p_end_date)

    if data_sentra or data_lainnya:
        # Render HTML table
        html_summary, total_sentra, total_lainnya, total_seindo, capaian_sentra, capaian_lainnya, capaian_seindo = render_summary_table_html(
            data_sentra, data_lainnya, start_date, end_date
        )

        # Download button for summary table
        excel_summary = create_summary_excel_export(
            data_sentra, data_lainnya, start_date, end_date,
            total_sentra, total_lainnya, total_seindo,
            capaian_sentra, capaian_lainnya, capaian_seindo
        )

        st.download_button(
            label="📥 Download Tabel Realisasi Kanwil (Excel)",
            data=excel_summary,
            file_name=f"summary_realisasi_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # Use components.html for better HTML rendering
        import streamlit.components.v1 as components
        components.html(html_summary, height=800, scrolling=True)
    else:
        st.warning("⚠️ Tidak ada data untuk ditampilkan")

    linech, barch = st.columns([2, 1])
    # ===== LINE CHART - MENGGUNAKAN RPC =====
    with linech:
        st.markdown('<div class="chart-title">📈 Tren Realisasi Kanwil</div>', unsafe_allow_html=True)

        try:
            fig = create_line_chart_from_rpc(p_nama_kanwil, p_akun_analitik, p_start_date, p_end_date, start_date, end_date)
            # Konfigurasi untuk memastikan tema light
            config = {
                'displayModeBar': True,
                'displaylogo': False,
                'modeBarButtonsToRemove': ['select2d', 'lasso2d'],
                'toImageButtonOptions': {
                    'format': 'png',
                    'filename': 'tren_realisasi',
                    'height': 600,
                    'width': 1200,
                    'scale': 2
                }
            }
            st.plotly_chart(fig, use_container_width=True, config=config, theme=None)
        except Exception as e:
            st.error(f"⚠️ Error saat membuat line chart: {str(e)}")
            st.info("Silakan selesaikan pemilihan filter untuk melihat chart")

    with barch:
        # ===== BAR CHART 7 HARI TERAKHIR - MENGGUNAKAN RPC =====
        st.markdown('<div class="chart-title">📊 Realisasi 7 Hari Terakhir</div>', unsafe_allow_html=True)

        try:
            fig_bar = create_bar_chart_7days_from_rpc(p_nama_kanwil, p_akun_analitik, p_end_date, end_date)
            # Konfigurasi untuk memastikan tema light
            config_bar = {
                'displayModeBar': True,
                'displaylogo': False,
                'modeBarButtonsToRemove': ['select2d', 'lasso2d'],
                'toImageButtonOptions': {
                    'format': 'png',
                    'filename': 'realisasi_7_hari',
                    'height': 600,
                    'width': 800,
                    'scale': 2
                }
            }
            st.plotly_chart(fig_bar, use_container_width=True, config=config_bar, theme=None)
        except Exception as e:
            st.error(f"⚠️ Error saat membuat bar chart: {str(e)}")
            st.info("Silakan selesaikan pemilihan filter untuk melihat chart")

    # ===== TABEL REALISASI DETAIL PER-KANCAB - MENGGUNAKAN RPC =====
    st.markdown('<div class="chart-title">📊 Tabel Realisasi per-Kancab</div>', unsafe_allow_html=True)

    # Debug: Print parameter sebelum memanggil function
    print(f"DEBUG MAIN - Calling create_kancab_table_from_rpc with p_nama_kanwil: {p_nama_kanwil}")
    print(f"DEBUG MAIN - selected_kanwil from filter: {selected_kanwil}")

    # Ambil data dari RPC
    kancab_df = create_kancab_table_from_rpc(p_nama_kanwil, p_akun_analitik, p_start_date, p_end_date)

    if not kancab_df.empty:
        excel_kancab = create_kancab_excel_export(kancab_df, end_date)
        st.download_button(
                label="📥 Download Tabel Realisasi Kancab (Excel)",
                data=excel_kancab,
                file_name=f"realisasi_kancab_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        # Render HTML table
        html_kancab = render_kancab_table_html(kancab_df, start_date, end_date)
        # Use components.html untuk scrolling seperti tabel kanwil
        import streamlit.components.v1 as components
        components.html(html_kancab, height=800, scrolling=True)
    else:
        if p_nama_kanwil is None:
            st.info("ℹ️ Silakan pilih Kanwil untuk melihat data Kancab")
        else:
            st.warning(f"⚠️ Tidak ada data Kancab untuk {p_nama_kanwil}")
    st.markdown("""
    <style>
    .stDownloadButton>button {
        background: linear-gradient(135deg, #1f497d 0%, #4bacc6 100%) !important;
        color: white !important;
        border-radius: 8px !important;
        padding: 8px 16px !important;
        border: none !important;
    }
    </style>
""", unsafe_allow_html=True)

if __name__ == "__main__":
    main()