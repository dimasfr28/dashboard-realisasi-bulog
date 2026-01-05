import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from datetime import datetime, timedelta
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import numpy as np
import gspread
import time

# Page configuration
st.set_page_config(
    page_title="Dashboard Pengadaan BULOG",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS - Light Blue Theme
st.markdown("""
<style>
            
    .st-emotion-cache-13k62yr {
        background-color: #f5f7fa;
    }
    /* Main Container */
    .main {
        background: linear-gradient(135deg, #f5f7fa 0%, #e8f4f8 100%);
        padding: 2rem;
    }
            
    .stAppToolbar{
        background: linear-gradient(135deg, #1f497d 0%, #4bacc6 100%);
    }
    
    /* Progress Bar */
    .stProgress > div > div > div > div {
        background: linear-gradient(90deg, #1f497d 0%, #4bacc6 100%);
        border-radius: 5px;
    }
    .stProgress > div > div > div {
        background-color: #e0e0e0; /* Light grey for the empty part */
        border-radius: 5px;
    }
    /* Style for the text of the progress bar */
    .stProgress label {
        color: #1f497d;
        font-weight: bold;
    }
    
    /* Header Title */
    .dashboard-header {
        background: linear-gradient(135deg, #1f497d 0%, #4bacc6 100%);
        padding: 2rem;
        border-radius: 15px;
        text-align: center;
        margin-bottom: 2rem;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
    }

    .dashboard-title {
        color: white;
        font-size: 2.5rem;
        font-weight: 700;
        margin: 0;
        text-shadow: 2px 2px 4px rgba(0, 0, 0, 0.2);
    }

    .dashboard-subtitle {
        color: #e8f4f8;
        font-size: 1.1rem;
        margin-top: 0.5rem;
    }

    /* Metrics Cards */
    div[data-testid="stMetric"] {
        background: white;
        padding: 1rem;
        border-radius: 10px;
        border-left: 4px solid #4bacc6;
        box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
    }

    div[data-testid="stMetric"] label {
        color: #000000 !important;
    }

    div[data-testid="stMetricValue"] {
        font-size: 1.8rem;
        color: #000000 !important;
        font-weight: 700;
    }

    div[data-testid="stMetricDelta"] {
        color: #000000 !important;
    }

    /* Sidebar */
    section[data-testid="stSidebar"] {
        background: linear-gradient(180deg, #1f497d 0%, #2c5f8d 100%);
    }


    /* Chart Container */
    .chart-container {
        background: white;
        padding: 2rem;
        border-radius: 15px;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
        margin: 2rem 0;
    }

    .chart-title {
        color: #1f497d;
        font-size: 1.5rem;
        font-weight: 700;
        margin-bottom: 1rem;
        border-bottom: 3px solid #4bacc6;
        padding-bottom: 0.5rem;
    }

    /* Table Container */
    .table-container {
        padding: 2rem;
        border-radius: 15px;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
        margin: 2rem 0;
    }

    /* Hide Streamlit branding */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}

    /* Better spacing */
    .block-container {
        padding-top: 1rem;
        padding-bottom: 1rem;
    }
    tbody{
            color: #1f497d;}
    stAppToolbar st-emotion-cache-14vh5up e1o8oa9v2 {
        background-color: #f5f7fa;
            }
</style>
""", unsafe_allow_html=True)

@st.cache_data(ttl=300)
def load_main_data():
    """Load and cache main data from local Excel file"""
    try:
        # Define file path
        file_path = 'assets/hasil_gabungan.xlsx'

        # Load main data from 'Realisasi' sheet
        df_realisasi = pd.read_excel(file_path, sheet_name='Export', engine='openpyxl')
        print("df_realisasi loaded")
        print(f"Loaded {len(df_realisasi)} rows from 'Export' sheet.")

        # Debug: Check GABAH data in Excel
        if 'Komoditi' in df_realisasi.columns and 'spesifikasi' in df_realisasi.columns:
            gabah_rows = df_realisasi[df_realisasi['Komoditi'] == 'GABAH']
            print(f"DEBUG EXCEL - Total GABAH rows: {len(gabah_rows)}")
            if len(gabah_rows) > 0:
                gabah_spek_unique = gabah_rows['spesifikasi'].unique().tolist()
                print(f"DEBUG EXCEL - GABAH spesifikasi values: {gabah_spek_unique}")
                print(f"DEBUG EXCEL - GABAH spesifikasi sample: {gabah_rows['spesifikasi'].head(20).tolist()}")

        # Convert date columns
        df_realisasi['Tanggal PO'] = pd.to_datetime(df_realisasi['Tanggal PO'], errors='coerce')
        df_realisasi['Tanggal Penerimaan'] = pd.to_datetime(df_realisasi['Tanggal Penerimaan'], errors='coerce')

        # Convert numeric columns to proper types
        numeric_columns = ['No. ID Pemasok', 'Tahun Stok', 'Kuantum PO (Kg)', 'In / Out',
                          'Harga Include ppn', 'Nominal Realisasi Incl ppn']
        for col in numeric_columns:
            if col in df_realisasi.columns:
                df_realisasi[col] = pd.to_numeric(df_realisasi[col], errors='coerce')

        # Debug: Check In / Out column
        if 'In / Out' in df_realisasi.columns:
            print(f"DEBUG EXCEL - In / Out sample values: {df_realisasi['In / Out'].head(10).tolist()}")
            print(f"DEBUG EXCEL - In / Out non-null count: {df_realisasi['In / Out'].notna().sum()}")
            print(f"DEBUG EXCEL - In / Out total sum: {df_realisasi['In / Out'].sum():.2f} Kg = {df_realisasi['In / Out'].sum() / 1000:.2f} Ton")

        # Clean invalid kanwil values
        invalid_kanwil_values = [
            'Total',
            'Applied filters:',
            'date_order_interval is on or after 01/01/2025',
            'tgl_penerimaan is on or after 01/01/2025 and is before 09/12/2025',
            'status_picking is done',
            'Exported data exceeded the allowed volume. Some data may have been omitted.',
            'p'
        ]

        # Remove rows with invalid kanwil values
        df_realisasi = df_realisasi[~df_realisasi['kanwil'].isin(invalid_kanwil_values)]

        # Also remove rows where kanwil contains "Applied filters" or "Exported data" (partial match)
        df_realisasi = df_realisasi[~df_realisasi['kanwil'].astype(str).str.contains('Applied filters', na=False)]
        df_realisasi = df_realisasi[~df_realisasi['kanwil'].astype(str).str.contains('Exported data', na=False)]
        df_realisasi = df_realisasi[~df_realisasi['kanwil'].astype(str).str.contains('date_order_interval', na=False)]
        df_realisasi = df_realisasi[~df_realisasi['kanwil'].astype(str).str.contains('tgl_penerimaan', na=False)]
        df_realisasi = df_realisasi[~df_realisasi['kanwil'].astype(str).str.contains('status_picking', na=False)]

        return df_realisasi

    except Exception as e:
        st.error(f"Error loading main data from Excel file: {str(e)}")
        st.info("Please make sure 'assets/hasil_gabungan.xlsx' is in the correct path and has the 'Export' sheet.")
        return pd.DataFrame()

@st.cache_data(ttl=300)
def load_target_kanwil():
    """Load and cache target kanwil data from local Excel file"""
    try:
        file_path = 'assets/hasil_gabungan.xlsx'
        # Load target data from 'Target Kanwil' sheet
        df_target_kanwil = pd.read_excel(file_path, sheet_name='Target Kanwil', engine='openpyxl')
        print(f"Loaded {len(df_target_kanwil)} rows from 'Target Kanwil' sheet.")
        if 'Target Setara Beras' in df_target_kanwil.columns:
            df_target_kanwil['Target Setara Beras'] = pd.to_numeric(df_target_kanwil['Target Setara Beras'], errors='coerce')
        return df_target_kanwil
    except Exception as e:
        st.error(f"Error loading Target Kanwil data from Excel file: {str(e)}")
        st.info("Please make sure 'assets/hasil_gabungan.xlsx' is in the correct path and has the 'Target Kanwil' sheet.")
        return pd.DataFrame()

@st.cache_data(ttl=300)
def load_target_kancab():
    """Load and cache target kancab data from local Excel file"""
    try:
        file_path = 'assets/hasil_gabungan.xlsx'
        # Load target data from 'Target Kancab' sheet
        df_target_kancab = pd.read_excel(file_path, sheet_name='Target Kancab', engine='openpyxl')
        if len(df_target_kancab) > 0 and 'Target Setara Beras' in df_target_kancab.columns:
            df_target_kancab['Target Setara Beras'] = pd.to_numeric(df_target_kancab['Target Setara Beras'], errors='coerce')
        return df_target_kancab
    except Exception as e:
        st.error(f"Error loading Target Kancab data from Excel file: {str(e)}")
        st.info("Please make sure 'assets/hasil_gabungan.xlsx' is in the correct path and has the 'Target Kancab' sheet.")
        return pd.DataFrame()

def load_all_data_with_progress():
    """Load all data and display a progress bar."""
    progress_bar = st.progress(0, "Memulai proses pemuatan data...")

    try:
        progress_bar.progress(10, "Memuat data utama...")
        df_realisasi = load_main_data()

        progress_bar.progress(50, "Memuat data target kanwil...")
        df_target_kanwil = load_target_kanwil()

        progress_bar.progress(75, "Memuat data target kancab...")
        df_target_kancab = load_target_kancab()

        progress_bar.progress(100, "Semua data berhasil dimuat.")
        progress_bar.empty()

        return df_realisasi, df_target_kanwil, df_target_kancab

    except Exception as e:
        st.error(f"Terjadi kesalahan saat memuat data: {str(e)}")
        progress_bar.empty()
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

def calculate_setara_beras(df):
    """
    Calculate Setara Beras from a dataframe
    Formula: d = a + 0.635*b + 0.53375*c
    where:
    a = BERAS (BERAS PREMIUM + BERAS MEDIUM)
    b = GKG (GABAH with spesifikasi containing 'GKG')
    c = GKP (GABAH with spesifikasi containing 'GKP')
    Returns value in Ton
    """
    # a) BERAS (BERAS MEDIUM + BERAS PREMIUM) - convert to Ton
    beras = df[
        df['Komoditi'].isin(['BERAS MEDIUM', 'BERAS PREMIUM'])
    ]['In / Out'].sum() / 1000

    # b) GKG - Gabah dengan spesifikasi GKG - convert to Ton
    gkg = df[
        (df['Komoditi'] == 'GABAH') &
        (df['spesifikasi'].astype(str).str.contains('GKG', case=False, na=False))
    ]['In / Out'].sum() / 1000

    # c) GKP - Gabah dengan spesifikasi GKP - convert to Ton
    gkp = df[
        (df['Komoditi'] == 'GABAH') &
        (df['spesifikasi'].astype(str).str.contains('GKP', case=False, na=False))
    ]['In / Out'].sum() / 1000

    # d) Setara Beras = a + 0.635*b + 0.53375*c
    setara_beras = beras + (0.635 * gkg) + (0.53375 * gkp)

    return setara_beras

def create_summary_table(df_filtered, df_target):
    """
    Create summary table with Kanwil Sentra Produksi and Kanwil Lainnya
    Shows: No, Kanwil, Target Setara Beras, Realisasi (Beras, GKG, GKP, Setara Beras, Capaian %)
    Formula: Setara Beras (d) = a + 0.635*b + 0.53375*c
    where a=Beras, b=GKG, c=GKP
    NOTE: df_filtered should already be filtered by periode realisasi
    """
    # List kanwil sentra produksi
    kanwil_sentra = [
        '08001 - KANTOR WILAYAH LAMPUNG',
        '21001 - KANTOR WILAYAH SULSEL SULBAR',
        '20001 - KANTOR WILAYAH SULTRA',
        '12001 - KANTOR WILAYAH DI YOGYAKARTA',
        '09001 - KANTOR WILAYAH DKI JAKARTA BANTEN',
        '23001 - KANTOR WILAYAH N.T.B',
        '13001 - KANTOR WILAYAH JATIM',
        '10001 - KANTOR WILAYAH JABAR',
        '01001 - KANTOR WILAYAH ACEH',
        '06001 - KANTOR WILAYAH SUMSEL',
        '11001 - KANTOR WILAYAH JATENG'
    ]

    # List kanwil lainnya
    kanwil_lainnya = [
        '15001 - KANTOR WILAYAH KALTIM KALTARA',
        '25001 - KANTOR WILAYAH MALUKU MALUT',
        '26001 - KANTOR WILAYAH PAPUA PABAR',
        '02001 - KANTOR WILAYAH SUMUT',
        '04001 - KANTOR WILAYAH SUMBAR',
        '17001 - KANTOR WILAYAH KALTENG',
        '16001 - KANTOR WILAYAH KALSEL',
        '14001 - KANTOR WILAYAH KALBAR',
        '18001 - KANTOR WILAYAH SULUT GORONTALO',
        '05001 - KANTOR WILAYAH JAMBI',
        '19001 - KANTOR WILAYAH SULTENG',
        '24001 - KANTOR WILAYAH N.T.T',
        '03001 - KANTOR WILAYAH RIAU DAN KEPRI',
        '22001 - KANTOR WILAYAH BALI',
        '07001 - KANTOR WILAYAH BENGKULU'
    ]

    def get_kanwil_data(kanwil):
        """Get data for a specific kanwil"""
        # df_filtered is already filtered by periode realisasi
        df_kanwil = df_filtered[df_filtered['kanwil'] == kanwil]

        # Beras (all komoditi containing 'BERAS') - convert to Ton
        beras = df_kanwil[
            df_kanwil['Komoditi'].astype(str).str.contains('BERAS', case=False, na=False)
        ]['In / Out'].sum() / 1000

        # GKG - Gabah dengan spesifikasi GKG - convert to Ton
        gkg = df_kanwil[
            (df_kanwil['Komoditi'] == 'GABAH') &
            (df_kanwil['spesifikasi'].astype(str).str.contains('GKG', case=False, na=False))
        ]['In / Out'].sum() / 1000

        # GKP - Gabah dengan spesifikasi GKP - convert to Ton
        gkp = df_kanwil[
            (df_kanwil['Komoditi'] == 'GABAH') &
            (df_kanwil['spesifikasi'].astype(str).str.contains('GKP', case=False, na=False))
        ]['In / Out'].sum() / 1000

        # Setara Beras: d = a + 0.635*b + 0.53375*c
        setara_beras = beras + (0.635 * gkg) + (0.53375 * gkp)

        # Get target from df_target - already in Ton
        target_row = df_target[df_target['kanwil'] == kanwil]
        target_setara_beras = target_row['Target Setara Beras'].values[0] if len(target_row) > 0 else 0

        # Capaian %
        capaian = (setara_beras / target_setara_beras * 100) if target_setara_beras > 0 else 0

        return {
            'Kanwil': kanwil,
            'Target Setara Beras': target_setara_beras,
            'Beras (a)': beras,
            'GKG (b)': gkg,
            'GKP (c)': gkp,
            'Setara Beras (d)': setara_beras,
            'Capaian (%)': capaian
        }

    # Build data for Kanwil Sentra Produksi
    data_sentra = []
    for kanwil in kanwil_sentra:
        row_data = get_kanwil_data(kanwil)
        data_sentra.append(row_data)

    # Sort by Capaian (%) descending
    data_sentra = sorted(data_sentra, key=lambda x: x['Capaian (%)'], reverse=True)
    # Re-assign No after sorting
    for idx, row in enumerate(data_sentra, 1):
        row['No'] = idx

    # Build data for Kanwil Lainnya
    data_lainnya = []
    for kanwil in kanwil_lainnya:
        row_data = get_kanwil_data(kanwil)
        data_lainnya.append(row_data)

    # Sort by Capaian (%) descending
    data_lainnya = sorted(data_lainnya, key=lambda x: x['Capaian (%)'], reverse=True)
    # Re-assign No after sorting
    for idx, row in enumerate(data_lainnya, 1):
        row['No'] = idx

    return data_sentra, data_lainnya

def create_line_chart(df_filtered, start_date=None, end_date=None):
    """Create line chart with 3 lines (BERAS, GKP, GKG) without value labels and without markers.

    Args:
        df_filtered: Filtered dataframe
        start_date: Start date of the period (datetime.date or string)
        end_date: End date of the period (datetime.date or string)
    """

    # Determine date range from parameters or data
    if start_date is not None and end_date is not None:
        # Convert to datetime if string
        if isinstance(start_date, str):
            start_date = pd.to_datetime(start_date).date()
        if isinstance(end_date, str):
            end_date = pd.to_datetime(end_date).date()
    elif not df_filtered.empty and 'Tanggal Penerimaan' in df_filtered.columns:
        # Use data range from filtered data
        min_date = df_filtered['Tanggal Penerimaan'].min()
        max_date = df_filtered['Tanggal Penerimaan'].max()
        start_date = min_date.date() if pd.notna(min_date) else datetime.now().date()
        end_date = max_date.date() if pd.notna(max_date) else datetime.now().date()
    else:
        # Fallback to current year
        current_year = datetime.now().year
        start_date = datetime(current_year, 1, 1).date()
        end_date = datetime(current_year, 12, 31).date()

    # Create complete date range based on selected period
    date_range = pd.date_range(
        start=start_date,
        end=end_date,
        freq='D'
    )
    complete_dates = pd.DataFrame({'Tanggal': date_range.date})

    # Filter BERAS
    df_beras = df_filtered[df_filtered['Komoditi'].isin(['BERAS MEDIUM', 'BERAS PREMIUM'])]
    daily_beras = df_beras.groupby(df_beras['Tanggal Penerimaan'].dt.date)['In / Out'].sum().reset_index()
    daily_beras.columns = ['Tanggal', 'In Out (Kg)']
    daily_beras['In Out (Ton)'] = daily_beras['In Out (Kg)'] / 1000
    # Merge with complete dates
    daily_beras = complete_dates.merge(daily_beras[['Tanggal', 'In Out (Ton)']], on='Tanggal', how='left')
    daily_beras['In Out (Ton)'] = daily_beras['In Out (Ton)'].fillna(0)
    daily_beras = daily_beras.sort_values('Tanggal')

    # Filter GKP
    df_gkp = df_filtered[
        (df_filtered['Komoditi'] == 'GABAH') &
        (df_filtered['spesifikasi'].astype(str).str.contains('GKP', case=False, na=False))
    ]
    daily_gkp = df_gkp.groupby(df_gkp['Tanggal Penerimaan'].dt.date)['In / Out'].sum().reset_index()
    daily_gkp.columns = ['Tanggal', 'In Out (Kg)']
    daily_gkp['In Out (Ton)'] = daily_gkp['In Out (Kg)'] / 1000
    # Merge with complete dates
    daily_gkp = complete_dates.merge(daily_gkp[['Tanggal', 'In Out (Ton)']], on='Tanggal', how='left')
    daily_gkp['In Out (Ton)'] = daily_gkp['In Out (Ton)'].fillna(0)
    daily_gkp = daily_gkp.sort_values('Tanggal')

    # Filter GKG
    df_gkg = df_filtered[
        (df_filtered['Komoditi'] == 'GABAH') &
        (df_filtered['spesifikasi'].astype(str).str.contains('GKG', case=False, na=False))
    ]
    daily_gkg = df_gkg.groupby(df_gkg['Tanggal Penerimaan'].dt.date)['In / Out'].sum().reset_index()
    daily_gkg.columns = ['Tanggal', 'In Out (Kg)']
    daily_gkg['In Out (Ton)'] = daily_gkg['In Out (Kg)'] / 1000
    # Merge with complete dates
    daily_gkg = complete_dates.merge(daily_gkg[['Tanggal', 'In Out (Ton)']], on='Tanggal', how='left')
    daily_gkg['In Out (Ton)'] = daily_gkg['In Out (Ton)'].fillna(0)
    daily_gkg = daily_gkg.sort_values('Tanggal')

    # Create figure
    fig = go.Figure()

    # BERAS (tanpa marker, tanpa text)
    fig.add_trace(go.Scatter(
        x=daily_beras['Tanggal'],
        y=daily_beras['In Out (Ton)'],
        mode='lines+markers',
        name='BERAS',
        line=dict(color='#1f497d', width=3, shape='spline'),
        fill='tozeroy',
        fillcolor='rgba(75, 172, 198, 0.15)',
        hovertemplate='<b>BERAS</b><br>Tanggal: %{x}<br>In/Out: %{y:,.2f} Ton<br><extra></extra>',
    ))

    # GKP
    fig.add_trace(go.Scatter(
        x=daily_gkp['Tanggal'],
        y=daily_gkp['In Out (Ton)'],
        mode='lines+markers',
        name='GKP',
        line=dict(color='#ff6b35', width=3, shape='spline'),
        fill='tozeroy',
        fillcolor='rgba(255, 107, 53, 0.15)',
        hovertemplate='<b>GKP</b><br>Tanggal: %{x}<br>In/Out: %{y:,.2f} Ton<br><extra></extra>',
    ))

    # GKG
    fig.add_trace(go.Scatter(
        x=daily_gkg['Tanggal'],
        y=daily_gkg['In Out (Ton)'],
        mode='lines+markers',
        name='GKG',
        line=dict(color='#28a745', width=3, shape='spline'),
        fill='tozeroy',
        fillcolor='rgba(92, 184, 92, 0.15)',
        hovertemplate='<b>GKG</b><br>Tanggal: %{x}<br>In/Out: %{y:,.2f} Ton<br><extra></extra>',
    ))

    # Layout
    fig.update_layout(
        title=None,
        xaxis_title=dict(text='Tanggal Penerimaan', font=dict(color='#000000', size=14)),
        yaxis_title=dict(text='In / Out (Ton)', font=dict(color='#000000', size=14)),
        height=450,
        hovermode='x unified',
        plot_bgcolor='rgba(248, 252, 255, 0.8)',
        paper_bgcolor='white',
        font=dict(family='Arial', size=12, color='#000000'),
        legend=dict(
            orientation="h",
            yanchor="bottom",
            y=1.02,
            xanchor="right",
            x=1
        ),
        xaxis=dict(
            showgrid=True,
            gridcolor='rgba(31, 73, 125, 0.1)',
            gridwidth=1,
            showline=True,
            linecolor='#000000',
            linewidth=2,
            tickangle=0,
            range=[start_date, end_date],
            tickformatstops = [
                dict(dtickrange=[None, "M1"], value="%d %b %Y"),
                dict(dtickrange=["M1", "M12"], value="%d %b %Y"),
                dict(dtickrange=["M12", None], value="%d %b %Y")
            ],
            tickfont=dict(color="#000000", size=11),
            title_font=dict(color="#000000", size=14)
        ),
        yaxis=dict(
            showgrid=True,
            gridcolor='rgba(31, 73, 125, 0.1)',
            gridwidth=1,
            tickformat=',.0f',
            showline=True,
            linecolor='#000000',
            linewidth=2,
            tickfont=dict(color="#000000", size=11),
            title_font=dict(color="#000000", size=14)
        ),
        margin=dict(t=40, b=60, l=60, r=40)
    )

    return fig

def create_bar_chart_7days(df_filtered, end_date=None):
    """
    Create bar chart for last 7 days with 3 bars: BERAS, GKP, GKB
    Based on Kanwil filter
    Args:
        df_filtered: DataFrame that has been filtered
        end_date: End date from filter (if None, use max date from data)
    """
    # Get last 7 days from the filtered data
    if df_filtered.empty:
        # Return empty figure with proper layout
        fig = go.Figure()
        fig.update_layout(
            title=None,
            xaxis_title='Tanggal',
            yaxis_title='Realisasi (Ton)',
            height=450,
            plot_bgcolor='rgba(248, 252, 255, 0.8)',
            paper_bgcolor='white'
        )
        return fig

    # Ensure Tanggal Penerimaan is datetime
    if not pd.api.types.is_datetime64_any_dtype(df_filtered['Tanggal Penerimaan']):
        df_filtered['Tanggal Penerimaan'] = pd.to_datetime(df_filtered['Tanggal Penerimaan'])

    # Get the maximum date: use end_date from filter if provided, otherwise use max from data
    if end_date is not None:
        max_date = pd.Timestamp(end_date)
    else:
        max_date = df_filtered['Tanggal Penerimaan'].max()

    # Calculate 7 days ago from max date
    start_date_7days = max_date - timedelta(days=6)  # 6 days before max_date = 7 days total

    # Filter data for last 7 days (between start_date_7days and max_date)
    df_7days = df_filtered[
        (df_filtered['Tanggal Penerimaan'] >= start_date_7days) &
        (df_filtered['Tanggal Penerimaan'] <= max_date)
    ].copy()

    # Group by date for each commodity
    # BERAS = BERAS PREMIUM + BERAS MEDIUM
    df_beras = df_7days[df_7days['Komoditi'].isin(['BERAS MEDIUM', 'BERAS PREMIUM'])]
    daily_beras = df_beras.groupby(df_beras['Tanggal Penerimaan'].dt.date)['In / Out'].sum().reset_index()
    daily_beras.columns = ['Tanggal', 'BERAS']
    daily_beras['BERAS'] = daily_beras['BERAS'] / 1000  # Convert to Ton

    # GKP = GABAH with spesifikasi containing 'GKP'
    df_gkp = df_7days[
        (df_7days['Komoditi'] == 'GABAH') &
        (df_7days['spesifikasi'].astype(str).str.contains('GKP', case=False, na=False))
    ]
    daily_gkp = df_gkp.groupby(df_gkp['Tanggal Penerimaan'].dt.date)['In / Out'].sum().reset_index()
    daily_gkp.columns = ['Tanggal', 'GKP']
    daily_gkp['GKP'] = daily_gkp['GKP'] / 1000  # Convert to Ton

    # GKB = GABAH with spesifikasi containing 'GKB' (note: prompt says GKB, not GKG)
    df_gkb = df_7days[
        (df_7days['Komoditi'] == 'GABAH') &
        (df_7days['spesifikasi'].astype(str).str.contains('GKB', case=False, na=False))
    ]
    daily_gkb = df_gkb.groupby(df_gkb['Tanggal Penerimaan'].dt.date)['In / Out'].sum().reset_index()
    daily_gkb.columns = ['Tanggal', 'GKB']
    daily_gkb['GKB'] = daily_gkb['GKB'] / 1000  # Convert to Ton

    # Create a complete date range for the last 7 days
    date_range = pd.date_range(start=start_date_7days.date(), end=max_date.date(), freq='D')
    df_dates = pd.DataFrame({'Tanggal': date_range.date})

    # Merge all data
    df_merged = df_dates.copy()
    df_merged = df_merged.merge(daily_beras, on='Tanggal', how='left')
    df_merged = df_merged.merge(daily_gkp, on='Tanggal', how='left')
    df_merged = df_merged.merge(daily_gkb, on='Tanggal', how='left')

    # Fill NaN with 0
    df_merged = df_merged.fillna(0)

    # Sort by date
    df_merged = df_merged.sort_values('Tanggal')

    # Prepare text labels for value annotations
    text_beras = [f'{val:,.1f}' if val != 0 else '' for val in df_merged['BERAS']]
    text_gkp = [f'{val:,.1f}' if val != 0 else '' for val in df_merged['GKP']]
    text_gkb = [f'{val:,.1f}' if val != 0 else '' for val in df_merged['GKB']]

    # Create figure
    fig = go.Figure()

    # Add BERAS bar
    fig.add_trace(go.Bar(
        x=df_merged['Tanggal'],
        y=df_merged['BERAS'],
        name='BERAS',
        marker=dict(color='#1f497d'),
        hovertemplate='<b>BERAS</b><br>Tanggal: %{x}<br>Realisasi: %{y:,.2f} Ton<br><extra></extra>'
    ))

    # Add GKP bar
    fig.add_trace(go.Bar(
        x=df_merged['Tanggal'],
        y=df_merged['GKP'],
        name='GKP',
        marker=dict(color='#ff6b35'),
        hovertemplate='<b>GKP</b><br>Tanggal: %{x}<br>Realisasi: %{y:,.2f} Ton<br><extra></extra>'
    ))

    # Add GKB bar
    fig.add_trace(go.Bar(
        x=df_merged['Tanggal'],
        y=df_merged['GKB'],
        name='GKB',
        marker=dict(color='#28a745'),
        hovertemplate='<b>GKB</b><br>Tanggal: %{x}<br>Realisasi: %{y:,.2f} Ton<br><extra></extra>'
    ))

    # Add value labels with yellow background for all bars
    # Bar positions: grouped bars use x-coordinate with offsets
    # For barmode='group' with 3 traces, bars are positioned with spacing
    bar_width_offset = 0.27  # Offset for grouped bars

    # BERAS annotations (first bar in each group, leftmost)
    for idx, (date, value, text) in enumerate(zip(df_merged['Tanggal'], df_merged['BERAS'], text_beras)):
        if text:  # Only add annotation if value != 0
            fig.add_annotation(
                x=idx - bar_width_offset,  # Left position in group
                y=value,
                text=text,
                showarrow=False,
                yanchor='bottom' if value >= 0 else 'top',
                yshift=5 if value >= 0 else -5,
                bgcolor='#FFD700',  # Yellow background (gold)
                bordercolor='#000000',
                borderwidth=1,
                borderpad=3,
                font=dict(
                    size=9,
                    color='#000000',
                    family='Arial'
                )
            )

    # GKP annotations (second bar in each group, middle)
    for idx, (date, value, text) in enumerate(zip(df_merged['Tanggal'], df_merged['GKP'], text_gkp)):
        if text:
            fig.add_annotation(
                x=idx,  # Center position in group
                y=value,
                text=text,
                showarrow=False,
                yanchor='bottom' if value >= 0 else 'top',
                yshift=5 if value >= 0 else -5,
                bgcolor='#FFD700',
                bordercolor='#000000',
                borderwidth=1,
                borderpad=3,
                font=dict(
                    size=9,
                    color='#000000',
                    family='Arial'
                )
            )

    # GKB annotations (third bar in each group, rightmost)
    for idx, (date, value, text) in enumerate(zip(df_merged['Tanggal'], df_merged['GKB'], text_gkb)):
        if text:
            fig.add_annotation(
                x=idx + bar_width_offset,  # Right position in group
                y=value,
                text=text,
                showarrow=False,
                yanchor='bottom' if value >= 0 else 'top',
                yshift=5 if value >= 0 else -5,
                bgcolor='#FFD700',
                bordercolor='#000000',
                borderwidth=1,
                borderpad=3,
                font=dict(
                    size=9,
                    color='#000000',
                    family='Arial'
                )
            )

    # Update layout
    fig.update_layout(
        title=None,
        xaxis_title=dict(text='Tanggal', font=dict(color='#000000', size=14)),
        yaxis_title=dict(text='Realisasi (Ton)', font=dict(color='#000000', size=14)),
        barmode='group',  # Bars side by side
        height=450,
        plot_bgcolor='rgba(248, 252, 255, 0.8)',
        paper_bgcolor='white',
        font=dict(family='Arial', size=12, color='#000000'),
        legend=dict(
            orientation="h",
            yanchor="bottom",
            y=1.02,
            xanchor="right",
            x=1
        ),
        xaxis=dict(
            showgrid=True,
            gridcolor='rgba(31, 73, 125, 0.1)',
            gridwidth=1,
            showline=True,
            linecolor='#000000',
            linewidth=2,
            tickfont=dict(color="#000000", size=11),
            title_font=dict(color="#000000", size=14),
            type='category'  # Treat dates as categories for better spacing
        ),
        yaxis=dict(
            showgrid=True,
            gridcolor='rgba(31, 73, 125, 0.1)',
            gridwidth=1,
            tickformat=',.0f',
            showline=True,
            linecolor='#000000',
            linewidth=2,
            tickfont=dict(color="#000000", size=11),
            title_font=dict(color="#000000", size=14)
        ),
        margin=dict(t=60, b=60, l=60, r=40)  # Increased top margin from 40 to 60 for value labels
    )

    return fig

def render_summary_table_html(data_sentra, data_lainnya, start_date, end_date):
    """Render summary table with Kanwil Sentra Produksi and Kanwil Lainnya in HTML"""

    def format_value(val, decimal=2):
        if pd.isna(val) or val is None or val == 0:
            return '-'
        return f'{val:,.{decimal}f}'

    # Build rows for Kanwil Sentra Produksi
    rows_sentra = ""
    total_sentra = {
        'Target Setara Beras': 0,
        'Beras (a)': 0,
        'GKG (b)': 0,
        'GKP (c)': 0,
        'Setara Beras (d)': 0
    }

    for row in data_sentra:
        rows_sentra += '<tr>'
        rows_sentra += f'<td>{row["No"]}</td>'
        rows_sentra += f'<td style="text-align: left; padding-left: 8px;">{row["Kanwil"]}</td>'
        rows_sentra += f'<td>{format_value(row["Target Setara Beras"])}</td>'
        rows_sentra += f'<td>{format_value(row["Beras (a)"])}</td>'
        rows_sentra += f'<td>{format_value(row["GKG (b)"])}</td>'
        rows_sentra += f'<td>{format_value(row["GKP (c)"])}</td>'
        rows_sentra += f'<td>{format_value(row["Setara Beras (d)"])}</td>'
        rows_sentra += f'<td>{format_value(row["Capaian (%)"], 1)}%</td>'
        rows_sentra += '</tr>'

        # Accumulate totals
        total_sentra['Target Setara Beras'] += row['Target Setara Beras']
        total_sentra['Beras (a)'] += row['Beras (a)']
        total_sentra['GKG (b)'] += row['GKG (b)']
        total_sentra['GKP (c)'] += row['GKP (c)']
        total_sentra['Setara Beras (d)'] += row['Setara Beras (d)']

    # Calculate capaian sentra
    capaian_sentra = (total_sentra['Setara Beras (d)'] / total_sentra['Target Setara Beras'] * 100) if total_sentra['Target Setara Beras'] > 0 else 0

    # Build rows for Kanwil Lainnya
    rows_lainnya = ""
    total_lainnya = {
        'Target Setara Beras': 0,
        'Beras (a)': 0,
        'GKG (b)': 0,
        'GKP (c)': 0,
        'Setara Beras (d)': 0
    }

    for row in data_lainnya:
        rows_lainnya += '<tr>'
        rows_lainnya += f'<td>{row["No"]}</td>'
        rows_lainnya += f'<td style="text-align: left; padding-left: 8px;">{row["Kanwil"]}</td>'
        rows_lainnya += f'<td>{format_value(row["Target Setara Beras"])}</td>'
        rows_lainnya += f'<td>{format_value(row["Beras (a)"])}</td>'
        rows_lainnya += f'<td>{format_value(row["GKG (b)"])}</td>'
        rows_lainnya += f'<td>{format_value(row["GKP (c)"])}</td>'
        rows_lainnya += f'<td>{format_value(row["Setara Beras (d)"])}</td>'
        rows_lainnya += f'<td>{format_value(row["Capaian (%)"], 1)}%</td>'
        rows_lainnya += '</tr>'

        # Accumulate totals
        total_lainnya['Target Setara Beras'] += row['Target Setara Beras']
        total_lainnya['Beras (a)'] += row['Beras (a)']
        total_lainnya['GKG (b)'] += row['GKG (b)']
        total_lainnya['GKP (c)'] += row['GKP (c)']
        total_lainnya['Setara Beras (d)'] += row['Setara Beras (d)']

    # Calculate capaian lainnya
    capaian_lainnya = (total_lainnya['Setara Beras (d)'] / total_lainnya['Target Setara Beras'] * 100) if total_lainnya['Target Setara Beras'] > 0 else 0

    # Calculate grand total
    total_seindo = {
        'Target Setara Beras': total_sentra['Target Setara Beras'] + total_lainnya['Target Setara Beras'],
        'Beras (a)': total_sentra['Beras (a)'] + total_lainnya['Beras (a)'],
        'GKG (b)': total_sentra['GKG (b)'] + total_lainnya['GKG (b)'],
        'GKP (c)': total_sentra['GKP (c)'] + total_lainnya['GKP (c)'],
        'Setara Beras (d)': total_sentra['Setara Beras (d)'] + total_lainnya['Setara Beras (d)']
    }
    capaian_seindo = (total_seindo['Setara Beras (d)'] / total_seindo['Target Setara Beras'] * 100) if total_seindo['Target Setara Beras'] > 0 else 0

    html = f'''
<table border="1" cellspacing="0" cellpadding="8" style="border-collapse: collapse; font-size: 12px; width: 100%; font-family: Arial;">
    <thead>
        <tr style="background: rgb(31, 73, 125); color: white; font-weight: bold;">
            <th rowspan="2">No.</th>
            <th rowspan="2">Kanwil</th>
            <th rowspan="2">Target Setara Beras</th>
            <th colspan="5">Realisasi Periode: {start_date.strftime("%d %b %Y")} - {end_date.strftime("%d %b %Y")}</th>
        </tr>
        <tr style="background: rgb(75, 172, 198); color: white; font-weight: bold;">
            <th>Beras (a)</th>
            <th>GKG (b)</th>
            <th>GKP (c)</th>
            <th>Setara Beras (d)</th>
            <th>Capaian (%)</th>
        </tr>
    </thead>
    <tbody>
        <!-- BAGIAN A: Kanwil Sentra Produksi -->
        <tr style="background:#fdc128; font-weight:bold;">
            <td colspan="8">a) Kanwil Sentra Produksi</td>
        </tr>
        {rows_sentra}
        <!-- SUBTOTAL A -->
        <tr style="background:#ffe599; font-weight:bold;">
            <td colspan="2">Total Kanwil Sentra Produksi</td>
            <td>{format_value(total_sentra['Target Setara Beras'])}</td>
            <td>{format_value(total_sentra['Beras (a)'])}</td>
            <td>{format_value(total_sentra['GKG (b)'])}</td>
            <td>{format_value(total_sentra['GKP (c)'])}</td>
            <td>{format_value(total_sentra['Setara Beras (d)'])}</td>
            <td>{format_value(capaian_sentra, 1)}%</td>
        </tr>

        <!-- BAGIAN B: Kanwil Lainnya -->
        <tr style="background:#fdc128; font-weight:bold;">
            <td colspan="8">b) Kanwil Lainnya</td>
        </tr>
        {rows_lainnya}
        <!-- SUBTOTAL B -->
        <tr style="background:#ffe599; font-weight:bold;">
            <td colspan="2">Total Kanwil Lainnya</td>
            <td>{format_value(total_lainnya['Target Setara Beras'])}</td>
            <td>{format_value(total_lainnya['Beras (a)'])}</td>
            <td>{format_value(total_lainnya['GKG (b)'])}</td>
            <td>{format_value(total_lainnya['GKP (c)'])}</td>
            <td>{format_value(total_lainnya['Setara Beras (d)'])}</td>
            <td>{format_value(capaian_lainnya, 1)}%</td>
        </tr>

        <!-- TOTAL SE-INDO -->
        <tr style="background: rgb(31, 73, 125); color: white; font-weight: bold;">
            <td colspan="2">TOTAL SE-INDO</td>
            <td>{format_value(total_seindo['Target Setara Beras'])}</td>
            <td>{format_value(total_seindo['Beras (a)'])}</td>
            <td>{format_value(total_seindo['GKG (b)'])}</td>
            <td>{format_value(total_seindo['GKP (c)'])}</td>
            <td>{format_value(total_seindo['Setara Beras (d)'])}</td>
            <td>{format_value(capaian_seindo, 1)}%</td>
        </tr>
    </tbody>
</table>
'''
    return html, total_sentra, total_lainnya, total_seindo, capaian_sentra, capaian_lainnya, capaian_seindo

def create_kancab_table(df_filtered, df_target):
    """
    Create table per-Kancab with structure:
    Columns: NO, KANCAB, TARGET SETARA BERAS, BERAS (a), GKG (b), GKP (c), SETARA BERAS (d), CAPAIAN (%)
    NOTE: df_filtered should already be filtered by periode and selected kanwil
    """
    # Get unique Kancab (Entitas) from filtered data
    kancab_list = sorted([k for k in df_filtered['Entitas'].unique() if pd.notna(k)])

    result_data = []

    for kancab in kancab_list:
        df_kancab = df_filtered[df_filtered['Entitas'] == kancab]

        # BERAS (BERAS MEDIUM + BERAS PREMIUM) - convert to Ton
        beras = df_kancab[
            df_kancab['Komoditi'].isin(['BERAS MEDIUM', 'BERAS PREMIUM'])
        ]['In / Out'].sum() / 1000

        # GKG - Gabah dengan spesifikasi GKG - convert to Ton
        gkg = df_kancab[
            (df_kancab['Komoditi'] == 'GABAH') &
            (df_kancab['spesifikasi'].astype(str).str.contains('GKG', case=False, na=False))
        ]['In / Out'].sum() / 1000

        # GKP - Gabah dengan spesifikasi GKP - convert to Ton
        gkp = df_kancab[
            (df_kancab['Komoditi'] == 'GABAH') &
            (df_kancab['spesifikasi'].astype(str).str.contains('GKP', case=False, na=False))
        ]['In / Out'].sum() / 1000

        # Setara Beras: d = a + 0.635*b + 0.53375*c
        setara_beras = beras + (0.635 * gkg) + (0.53375 * gkp)

        # Get Target from df_target (if exists)
        # Match kancab name from df_filtered['Entitas'] with df_target['kancab']
        target_setara_beras = None
        if 'kancab' in df_target.columns and 'Target Setara Beras' in df_target.columns:
            # Try exact match first
            target_row = df_target[df_target['kancab'] == kancab]

            # If no exact match, try case-insensitive match with trimmed whitespace
            if len(target_row) == 0:
                kancab_normalized = str(kancab).strip().upper()
                df_target_normalized = df_target.copy()
                df_target_normalized['kancab_norm'] = df_target_normalized['kancab'].astype(str).str.strip().str.upper()
                target_row = df_target_normalized[df_target_normalized['kancab_norm'] == kancab_normalized]

            if len(target_row) > 0:
                target_setara_beras = target_row['Target Setara Beras'].values[0]

        # If no target found, set to None (will show as "-")
        if target_setara_beras is None or pd.isna(target_setara_beras):
            target_setara_beras = None

        # Calculate Capaian %
        if target_setara_beras is not None and target_setara_beras > 0:
            capaian = (setara_beras / target_setara_beras) * 100
        else:
            capaian = None

        # Only add if there's any data
        if setara_beras > 0 or target_setara_beras is not None:
            result_data.append({
                'Kancab': kancab,
                'Target Setara Beras': target_setara_beras,
                'Beras (a)': beras,
                'GKG (b)': gkg,
                'GKP (c)': gkp,
                'Setara Beras (d)': setara_beras,
                'Capaian (%)': capaian
            })

    result_df = pd.DataFrame(result_data)

    # Sort by Capaian (%) descending, then by Setara Beras descending for ties
    if not result_df.empty:
        # Fill NaN in Capaian with -1 so they appear at the bottom when sorted
        result_df['Capaian (%)'] = result_df['Capaian (%)'].fillna(-1)
        result_df = result_df.sort_values(['Capaian (%)', 'Setara Beras (d)'], ascending=[False, False]).reset_index(drop=True)
        # Replace -1 back to None for display
        result_df['Capaian (%)'] = result_df['Capaian (%)'].replace(-1, None)
        # Add NO column
        result_df.insert(0, 'NO', range(1, len(result_df) + 1))

        # Calculate TOTAL row
        total_target = result_df['Target Setara Beras'].dropna().sum() if result_df['Target Setara Beras'].notna().any() else None
        total_beras = result_df['Beras (a)'].sum()
        total_gkg = result_df['GKG (b)'].sum()
        total_gkp = result_df['GKP (c)'].sum()
        total_setara = result_df['Setara Beras (d)'].sum()

        # Calculate total capaian
        if total_target is not None and total_target > 0:
            total_capaian = (total_setara / total_target) * 100
        else:
            total_capaian = None

        total_row = {
            'NO': '',
            'Kancab': 'TOTAL KANWIL',
            'Target Setara Beras': total_target,
            'Beras (a)': total_beras,
            'GKG (b)': total_gkg,
            'GKP (c)': total_gkp,
            'Setara Beras (d)': total_setara,
            'Capaian (%)': total_capaian
        }
        result_df = pd.concat([result_df, pd.DataFrame([total_row])], ignore_index=True)

    return result_df

def render_kancab_table_html(df, start_date, end_date):
    """Render Kancab table in HTML with new structure"""
    if df.empty:
        return "<p>Tidak ada data</p>"

    def format_value(val, decimal=2, is_percent=False):
        if pd.isna(val) or val is None:
            return '-'
        if val == 0:
            return '-'
        if is_percent:
            return f'{val:,.1f}%'
        return f'{val:,.{decimal}f}'

    # Build rows
    rows_html = ""
    for _, row in df.iterrows():
        is_total = row['Kancab'] == 'TOTAL KANWIL'
        if is_total:
            row_style = 'background: #ffe599; color: black; font-weight: bold;'
        else:
            row_style = 'color: black;'

        rows_html += f'<tr style="{row_style}">'
        rows_html += f'<td>{row["NO"] if row["NO"] != "" else ""}</td>'
        rows_html += f'<td style="text-align: left; padding-left: 8px;">{row["Kancab"]}</td>'
        rows_html += f'<td>{format_value(row["Target Setara Beras"])}</td>'
        rows_html += f'<td>{format_value(row["Beras (a)"])}</td>'
        rows_html += f'<td>{format_value(row["GKG (b)"])}</td>'
        rows_html += f'<td>{format_value(row["GKP (c)"])}</td>'
        rows_html += f'<td>{format_value(row["Setara Beras (d)"])}</td>'
        rows_html += f'<td>{format_value(row["Capaian (%)"], is_percent=True)}</td>'
        rows_html += '</tr>'

    html = f'''
<table border="1" cellspacing="0" cellpadding="8" style="border-collapse: collapse; font-size: 12px; width: 100%; font-family: Arial;">
    <thead>
        <tr style="background: rgb(31, 73, 125); color: white; font-weight: bold;">
            <th rowspan="2">No.</th>
            <th rowspan="2">Kancab</th>
            <th rowspan="2">Target Setara Beras</th>
            <th colspan="5">Realisasi S. d. {end_date.strftime("%d %b %Y")}</th>
        </tr>
        <tr style="background: rgb(75, 172, 198); color: white; font-weight: bold;">
            <th>Beras (a)</th>
            <th>GKG (b)</th>
            <th>GKP (c)</th>
            <th>Setara Beras (d)</th>
            <th>Capaian (%)</th>
        </tr>
    </thead>
    <tbody>
        {rows_html}
    </tbody>
</table>
'''
    return html

def create_kancab_excel_export(df, end_date):
    """Create Excel file for Kancab table with same styling as HTML"""
    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Write empty dataframe to create workbook
        pd.DataFrame().to_excel(writer, sheet_name='Kancab', index=False)

        # Get workbook and worksheet
        workbook = writer.book
        worksheet = workbook['Kancab']

        # Define styles
        header_fill_dark = PatternFill(start_color='1f497d', end_color='1f497d', fill_type='solid')
        header_fill_light = PatternFill(start_color='4bacc6', end_color='4bacc6', fill_type='solid')
        total_fill = PatternFill(start_color='b6d7a8', end_color='b6d7a8', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        bold_font = Font(bold=True, size=10)
        normal_font = Font(size=10)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write header row 1
        worksheet.cell(row=1, column=1, value='No.').fill = header_fill_dark
        worksheet.cell(row=1, column=1).font = header_font
        worksheet.cell(row=1, column=1).alignment = center_align
        worksheet.cell(row=1, column=1).border = thin_border
        worksheet.merge_cells('A1:A2')

        worksheet.cell(row=1, column=2, value='Kancab').fill = header_fill_dark
        worksheet.cell(row=1, column=2).font = header_font
        worksheet.cell(row=1, column=2).alignment = center_align
        worksheet.cell(row=1, column=2).border = thin_border
        worksheet.merge_cells('B1:B2')

        worksheet.cell(row=1, column=3, value='Target Setara Beras').fill = header_fill_dark
        worksheet.cell(row=1, column=3).font = header_font
        worksheet.cell(row=1, column=3).alignment = center_align
        worksheet.cell(row=1, column=3).border = thin_border
        worksheet.merge_cells('C1:C2')

        worksheet.cell(row=1, column=4, value=f'Realisasi S. d. {end_date.strftime("%d %b %Y")}').fill = header_fill_dark
        worksheet.cell(row=1, column=4).font = header_font
        worksheet.cell(row=1, column=4).alignment = center_align
        worksheet.cell(row=1, column=4).border = thin_border
        worksheet.merge_cells('D1:H1')

        # Write header row 2
        headers_row2 = ['Beras (a)', 'GKG (b)', 'GKP (c)', 'Setara Beras (d)', 'Capaian (%)']
        for col_idx, header in enumerate(headers_row2, start=4):
            cell = worksheet.cell(row=2, column=col_idx, value=header)
            cell.fill = header_fill_light
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        # Write data rows
        current_row = 3
        for _, row in df.iterrows():
            is_total = row['Kancab'] == 'TOTAL KANWIL'

            worksheet.cell(row=current_row, column=1, value=row['NO'] if row['NO'] != '' else '')
            worksheet.cell(row=current_row, column=1).alignment = center_align
            worksheet.cell(row=current_row, column=1).border = thin_border

            worksheet.cell(row=current_row, column=2, value=row['Kancab'])
            worksheet.cell(row=current_row, column=2).alignment = left_align
            worksheet.cell(row=current_row, column=2).border = thin_border

            worksheet.cell(row=current_row, column=3, value=row['Target Setara Beras'] if pd.notna(row['Target Setara Beras']) else None)
            worksheet.cell(row=current_row, column=3).number_format = '#,##0.00'
            worksheet.cell(row=current_row, column=3).alignment = center_align
            worksheet.cell(row=current_row, column=3).border = thin_border

            worksheet.cell(row=current_row, column=4, value=row['Beras (a)'] if pd.notna(row['Beras (a)']) and row['Beras (a)'] > 0 else None)
            worksheet.cell(row=current_row, column=4).number_format = '#,##0.00'
            worksheet.cell(row=current_row, column=4).alignment = center_align
            worksheet.cell(row=current_row, column=4).border = thin_border

            worksheet.cell(row=current_row, column=5, value=row['GKG (b)'] if pd.notna(row['GKG (b)']) and row['GKG (b)'] > 0 else None)
            worksheet.cell(row=current_row, column=5).number_format = '#,##0.00'
            worksheet.cell(row=current_row, column=5).alignment = center_align
            worksheet.cell(row=current_row, column=5).border = thin_border

            worksheet.cell(row=current_row, column=6, value=row['GKP (c)'] if pd.notna(row['GKP (c)']) and row['GKP (c)'] > 0 else None)
            worksheet.cell(row=current_row, column=6).number_format = '#,##0.00'
            worksheet.cell(row=current_row, column=6).alignment = center_align
            worksheet.cell(row=current_row, column=6).border = thin_border

            worksheet.cell(row=current_row, column=7, value=row['Setara Beras (d)'] if pd.notna(row['Setara Beras (d)']) and row['Setara Beras (d)'] > 0 else None)
            worksheet.cell(row=current_row, column=7).number_format = '#,##0.00'
            worksheet.cell(row=current_row, column=7).alignment = center_align
            worksheet.cell(row=current_row, column=7).border = thin_border

            worksheet.cell(row=current_row, column=8, value=row['Capaian (%)'] if pd.notna(row['Capaian (%)']) else None)
            worksheet.cell(row=current_row, column=8).number_format = '0.0"%"'
            worksheet.cell(row=current_row, column=8).alignment = center_align
            worksheet.cell(row=current_row, column=8).border = thin_border

            # Apply styling
            if is_total:
                for col_idx in range(1, 9):
                    worksheet.cell(row=current_row, column=col_idx).fill = total_fill
                    worksheet.cell(row=current_row, column=col_idx).font = bold_font
            else:
                for col_idx in range(1, 9):
                    worksheet.cell(row=current_row, column=col_idx).font = normal_font

            current_row += 1

        # Set column widths
        worksheet.column_dimensions['A'].width = 5
        worksheet.column_dimensions['B'].width = 40
        worksheet.column_dimensions['C'].width = 18
        worksheet.column_dimensions['D'].width = 12
        worksheet.column_dimensions['E'].width = 12
        worksheet.column_dimensions['F'].width = 12
        worksheet.column_dimensions['G'].width = 18
        worksheet.column_dimensions['H'].width = 14

    output.seek(0)
    return output

def create_complex_table(df_filtered, selected_date, selected_kanwil_list):
    """
    Create complex table with multi-row header
    Columns: NO, KANWIL, TARGET (BERAS, GABAH),
             REALISASI SD KEMARIN, REALISASI HARI INI, REALISASI SD TGL,
             CAPAIAN (% BERAS, % GABAH)
    """
    # Tanggal kemarin
    tanggal_kemarin = selected_date - timedelta(days=1)
    tanggal_hari_ini = selected_date

    # Use the passed list of ALL selected kanwil
    # This ensures we show all selected kanwil even if they don't have BERAS or GABAH
    kanwil_list = sorted([k for k in selected_kanwil_list if pd.notna(k)])

    result_data = []

    for kanwil in kanwil_list:
        df_kanwil = df_filtered[df_filtered['kanwil'] == kanwil]

        # TARGET BERAS (from Kuantum PO) - BERAS MEDIUM + BERAS PREMIUM
        target_beras = df_kanwil[
            df_kanwil['Komoditi'].isin(['BERAS MEDIUM', 'BERAS PREMIUM'])
        ]['Kuantum PO (Kg)'].sum() / 1000  # to Ton

        # TARGET GABAH (from Kuantum PO)
        target_gabah = df_kanwil[
            df_kanwil['Komoditi'] == 'GABAH'
        ]['Kuantum PO (Kg)'].sum() / 1000  # to Ton

        # Check if has data for BERAS or GABAH
        has_beras = target_beras > 0
        has_gabah = target_gabah > 0

        # REALISASI SD KEMARIN - BERAS
        real_sd_kemarin_beras = df_kanwil[
            (df_kanwil['Tanggal Penerimaan'].dt.date < tanggal_kemarin) &
            (df_kanwil['Komoditi'].isin(['BERAS MEDIUM', 'BERAS PREMIUM']))
        ]['In / Out'].sum() / 1000 if has_beras else None  # to Ton

        # REALISASI SD KEMARIN - GABAH
        real_sd_kemarin_gabah = df_kanwil[
            (df_kanwil['Tanggal Penerimaan'].dt.date < tanggal_kemarin) &
            (df_kanwil['Komoditi'] == 'GABAH')
        ]['In / Out'].sum() / 1000 if has_gabah else None  # to Ton

        # REALISASI HARI INI - BERAS (only on selected date)
        real_hari_ini_beras = df_kanwil[
            (df_kanwil['Tanggal Penerimaan'].dt.date == tanggal_hari_ini) &
            (df_kanwil['Komoditi'].isin(['BERAS MEDIUM', 'BERAS PREMIUM']))
        ]['In / Out'].sum() / 1000 if has_beras else None  # to Ton

        # REALISASI HARI INI - GABAH (only on selected date)
        real_hari_ini_gabah = df_kanwil[
            (df_kanwil['Tanggal Penerimaan'].dt.date == tanggal_hari_ini) &
            (df_kanwil['Komoditi'] == 'GABAH')
        ]['In / Out'].sum() / 1000 if has_gabah else None  # to Ton

        # REALISASI SD TGL - BERAS (until selected date)
        real_sd_tgl_beras = df_kanwil[
            (df_kanwil['Tanggal Penerimaan'].dt.date <= tanggal_hari_ini) &
            (df_kanwil['Komoditi'].isin(['BERAS MEDIUM', 'BERAS PREMIUM']))
        ]['In / Out'].sum() / 1000 if has_beras else None  # to Ton

        # REALISASI SD TGL - GABAH (until selected date)
        real_sd_tgl_gabah = df_kanwil[
            (df_kanwil['Tanggal Penerimaan'].dt.date <= tanggal_hari_ini) &
            (df_kanwil['Komoditi'] == 'GABAH')
        ]['In / Out'].sum() / 1000 if has_gabah else None  # to Ton

        # CAPAIAN %
        capaian_beras = (real_sd_tgl_beras / target_beras * 100) if (has_beras and target_beras > 0) else None
        capaian_gabah = (real_sd_tgl_gabah / target_gabah * 100) if (has_gabah and target_gabah > 0) else None

        result_data.append({
            'Kanwil': kanwil,
            'Target Beras': target_beras if has_beras else None,
            'Target Gabah': target_gabah if has_gabah else None,
            'Real SD Kemarin Beras': real_sd_kemarin_beras,
            'Real SD Kemarin Gabah': real_sd_kemarin_gabah,
            'Real Hari Ini Beras': real_hari_ini_beras,
            'Real Hari Ini Gabah': real_hari_ini_gabah,
            'Real SD Tgl Beras': real_sd_tgl_beras,
            'Real SD Tgl Gabah': real_sd_tgl_gabah,
            'Capaian Beras %': capaian_beras,
            'Capaian Gabah %': capaian_gabah
        })

    result_df = pd.DataFrame(result_data)

    # Sort by Target Beras descending (None values will be at the bottom)
    result_df = result_df.sort_values('Target Beras', ascending=False, na_position='last').reset_index(drop=True)

    # Add NO column
    result_df.insert(0, 'NO', range(1, len(result_df) + 1))

    # Calculate TOTAL row (sum only non-None values)
    total_target_beras = result_df['Target Beras'].dropna().sum()
    total_target_gabah = result_df['Target Gabah'].dropna().sum()
    total_real_sd_kemarin_beras = result_df['Real SD Kemarin Beras'].dropna().sum()
    total_real_sd_kemarin_gabah = result_df['Real SD Kemarin Gabah'].dropna().sum()
    total_real_hari_ini_beras = result_df['Real Hari Ini Beras'].dropna().sum()
    total_real_hari_ini_gabah = result_df['Real Hari Ini Gabah'].dropna().sum()
    total_real_sd_tgl_beras = result_df['Real SD Tgl Beras'].dropna().sum()
    total_real_sd_tgl_gabah = result_df['Real SD Tgl Gabah'].dropna().sum()

    total_row = {
        'NO': '',
        'Kanwil': 'TOTAL',
        'Target Beras': total_target_beras if total_target_beras > 0 else None,
        'Target Gabah': total_target_gabah if total_target_gabah > 0 else None,
        'Real SD Kemarin Beras': total_real_sd_kemarin_beras if total_real_sd_kemarin_beras > 0 else None,
        'Real SD Kemarin Gabah': total_real_sd_kemarin_gabah if total_real_sd_kemarin_gabah > 0 else None,
        'Real Hari Ini Beras': total_real_hari_ini_beras if total_real_hari_ini_beras > 0 else None,
        'Real Hari Ini Gabah': total_real_hari_ini_gabah if total_real_hari_ini_gabah > 0 else None,
        'Real SD Tgl Beras': total_real_sd_tgl_beras if total_real_sd_tgl_beras > 0 else None,
        'Real SD Tgl Gabah': total_real_sd_tgl_gabah if total_real_sd_tgl_gabah > 0 else None,
        'Capaian Beras %': (total_real_sd_tgl_beras / total_target_beras * 100) if total_target_beras > 0 else None,
        'Capaian Gabah %': (total_real_sd_tgl_gabah / total_target_gabah * 100) if total_target_gabah > 0 else None
    }

    result_df = pd.concat([result_df, pd.DataFrame([total_row])], ignore_index=True)

    return result_df, tanggal_kemarin, tanggal_hari_ini

def render_complex_table_html(df, tanggal_kemarin, tanggal_hari_ini):
    """Render table with complex multi-row header in HTML"""

    # Build HTML row by row
    rows_html = ""
    for _, row in df.iterrows():
        is_total = row['Kanwil'] == 'TOTAL'
        row_style = 'font-weight:bold; background:#ddebf7;' if is_total else ''

        # Helper function to format value or show '-'
        def format_value(val, decimal=2, is_percent=False):
            if pd.isna(val) or val is None:
                return '-'
            if is_percent:
                return f'• {val:,.1f}%'
            return f'• {val:,.{decimal}f}'

        rows_html += f'<tr style="{row_style}">'
        rows_html += f'<td>{row["NO"]}</td>'
        rows_html += f'<td style="text-align: left; padding-left: 10px;">{row["Kanwil"]}</td>'
        rows_html += f'<td>{format_value(row["Target Beras"])}</td>'
        rows_html += f'<td>{format_value(row["Target Gabah"])}</td>'
        rows_html += f'<td>{format_value(row["Real SD Kemarin Beras"])}</td>'
        rows_html += f'<td>{format_value(row["Real SD Kemarin Gabah"])}</td>'
        rows_html += f'<td>{format_value(row["Real Hari Ini Beras"])}</td>'
        rows_html += f'<td>{format_value(row["Real Hari Ini Gabah"])}</td>'
        rows_html += f'<td>{format_value(row["Real SD Tgl Beras"])}</td>'
        rows_html += f'<td>{format_value(row["Real SD Tgl Gabah"])}</td>'
        rows_html += f'<td>{format_value(row["Capaian Beras %"], is_percent=True)}</td>'
        rows_html += f'<td>{format_value(row["Capaian Gabah %"], is_percent=True)}</td>'
        rows_html += '</tr>'

    html = f'''
<table border="1" cellspacing="0" cellpadding="8" style="border-collapse: collapse; font-size: 12px; text-align: center; width: 100%; font-family: Arial;">
    <thead>
        <tr style="background:#1f497d; color:white; font-weight:bold;">
            <th rowspan="3" style="vertical-align: middle;">NO</th>
            <th rowspan="3" style="vertical-align: middle;">KANWIL</th>
            <th colspan="2" rowspan="2" style="vertical-align: middle;">TARGET</th>
            <th colspan="2">REALISASI SD. KEMARIN</th>
            <th colspan="2">REALISASI HARI INI</th>
            <th colspan="2">REALISASI SD. TGL</th>
            <th colspan="2" rowspan="2" style="vertical-align: middle;">CAPAIAN</th>
        </tr>
        <tr style="background:#1f497d; color:white; font-weight:bold;">
            <th colspan="2">{tanggal_kemarin.strftime("%d %B %Y")}</th>
            <th colspan="2">{tanggal_hari_ini.strftime("%d %B %Y")}</th>
            <th colspan="2">{tanggal_hari_ini.strftime("%d %B %Y")}</th>
        </tr>
        <tr style="background:#4bacc6; color:white; font-weight:bold;">
            <th>BERAS</th><th>GABAH</th>
            <th>BERAS</th><th>GABAH</th>
            <th>BERAS</th><th>GABAH</th>
            <th>BERAS</th><th>GABAH</th>
            <th>% BERAS</th><th>% GABAH</th>
        </tr>
    </thead>
    <tbody>
        {rows_html}
    </tbody>
</table>
'''

    return html

def create_summary_excel_export(data_sentra, data_lainnya, start_date, end_date, total_sentra, total_lainnya, total_seindo, capaian_sentra, capaian_lainnya, capaian_seindo):
    """Create Excel file for summary table with same styling as HTML"""
    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Write empty dataframe to create workbook
        pd.DataFrame().to_excel(writer, sheet_name='Summary', index=False)

        # Get workbook and worksheet
        workbook = writer.book
        worksheet = workbook['Summary']

        # Define styles
        header_fill = PatternFill(start_color='f2f2f2', end_color='f2f2f2', fill_type='solid')
        sentra_fill = PatternFill(start_color='ffe599', end_color='ffe599', fill_type='solid')
        lainnya_fill = PatternFill(start_color='c9daf8', end_color='c9daf8', fill_type='solid')
        seindo_fill = PatternFill(start_color='b6d7a8', end_color='b6d7a8', fill_type='solid')
        header_font = Font(bold=True, size=11)
        bold_font = Font(bold=True, size=10)
        normal_font = Font(size=10)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write header row 1
        worksheet.cell(row=1, column=1, value='No.').fill = header_fill
        worksheet.cell(row=1, column=1).font = header_font
        worksheet.cell(row=1, column=1).alignment = center_align
        worksheet.cell(row=1, column=1).border = thin_border
        worksheet.merge_cells('A1:A2')

        worksheet.cell(row=1, column=2, value='Kanwil').fill = header_fill
        worksheet.cell(row=1, column=2).font = header_font
        worksheet.cell(row=1, column=2).alignment = center_align
        worksheet.cell(row=1, column=2).border = thin_border
        worksheet.merge_cells('B1:B2')

        worksheet.cell(row=1, column=3, value='Target Setara Beras').fill = header_fill
        worksheet.cell(row=1, column=3).font = header_font
        worksheet.cell(row=1, column=3).alignment = center_align
        worksheet.cell(row=1, column=3).border = thin_border
        worksheet.merge_cells('C1:C2')

        worksheet.cell(row=1, column=4, value=f'Realisasi Periode: {start_date.strftime("%d %b %Y")} - {end_date.strftime("%d %b %Y")}').fill = header_fill
        worksheet.cell(row=1, column=4).font = header_font
        worksheet.cell(row=1, column=4).alignment = center_align
        worksheet.cell(row=1, column=4).border = thin_border
        worksheet.merge_cells('D1:H1')

        # Write header row 2
        headers_row2 = ['Beras (a)', 'GKG (b)', 'GKP (c)', 'Setara Beras (d)', 'Capaian (%)']
        for col_idx, header in enumerate(headers_row2, start=4):
            cell = worksheet.cell(row=2, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        # Current row
        current_row = 3

        # Write Kanwil Sentra Produksi header
        cell = worksheet.cell(row=current_row, column=1, value='a) Kanwil Sentra Produksi')
        cell.fill = sentra_fill
        cell.font = bold_font
        cell.alignment = left_align
        cell.border = thin_border
        worksheet.merge_cells(f'A{current_row}:H{current_row}')
        current_row += 1

        # Write data for Kanwil Sentra Produksi
        for row_data in data_sentra:
            worksheet.cell(row=current_row, column=1, value=row_data['No']).alignment = center_align
            worksheet.cell(row=current_row, column=1).border = thin_border
            worksheet.cell(row=current_row, column=2, value=row_data['Kanwil']).alignment = left_align
            worksheet.cell(row=current_row, column=2).border = thin_border
            worksheet.cell(row=current_row, column=3, value=row_data['Target Setara Beras']).number_format = '#,##0.00'
            worksheet.cell(row=current_row, column=3).alignment = center_align
            worksheet.cell(row=current_row, column=3).border = thin_border
            worksheet.cell(row=current_row, column=4, value=row_data['Beras (a)']).number_format = '#,##0.00'
            worksheet.cell(row=current_row, column=4).alignment = center_align
            worksheet.cell(row=current_row, column=4).border = thin_border
            worksheet.cell(row=current_row, column=5, value=row_data['GKG (b)']).number_format = '#,##0.00'
            worksheet.cell(row=current_row, column=5).alignment = center_align
            worksheet.cell(row=current_row, column=5).border = thin_border
            worksheet.cell(row=current_row, column=6, value=row_data['GKP (c)']).number_format = '#,##0.00'
            worksheet.cell(row=current_row, column=6).alignment = center_align
            worksheet.cell(row=current_row, column=6).border = thin_border
            worksheet.cell(row=current_row, column=7, value=row_data['Setara Beras (d)']).number_format = '#,##0.00'
            worksheet.cell(row=current_row, column=7).alignment = center_align
            worksheet.cell(row=current_row, column=7).border = thin_border
            worksheet.cell(row=current_row, column=8, value=row_data['Capaian (%)']).number_format = '0.0"%"'
            worksheet.cell(row=current_row, column=8).alignment = center_align
            worksheet.cell(row=current_row, column=8).border = thin_border
            current_row += 1

        # Write Total Kanwil Sentra Produksi
        worksheet.cell(row=current_row, column=1, value='Total Kanwil Sentra Produksi').fill = sentra_fill
        worksheet.cell(row=current_row, column=1).font = bold_font
        worksheet.cell(row=current_row, column=1).alignment = left_align
        worksheet.cell(row=current_row, column=1).border = thin_border
        worksheet.merge_cells(f'A{current_row}:B{current_row}')
        worksheet.cell(row=current_row, column=3, value=total_sentra['Target Setara Beras']).number_format = '#,##0.00'
        worksheet.cell(row=current_row, column=3).fill = sentra_fill
        worksheet.cell(row=current_row, column=3).font = bold_font
        worksheet.cell(row=current_row, column=3).alignment = center_align
        worksheet.cell(row=current_row, column=3).border = thin_border
        worksheet.cell(row=current_row, column=4, value=total_sentra['Beras (a)']).number_format = '#,##0.00'
        worksheet.cell(row=current_row, column=4).fill = sentra_fill
        worksheet.cell(row=current_row, column=4).font = bold_font
        worksheet.cell(row=current_row, column=4).alignment = center_align
        worksheet.cell(row=current_row, column=4).border = thin_border
        worksheet.cell(row=current_row, column=5, value=total_sentra['GKG (b)']).number_format = '#,##0.00'
        worksheet.cell(row=current_row, column=5).fill = sentra_fill
        worksheet.cell(row=current_row, column=5).font = bold_font
        worksheet.cell(row=current_row, column=5).alignment = center_align
        worksheet.cell(row=current_row, column=5).border = thin_border
        worksheet.cell(row=current_row, column=6, value=total_sentra['GKP (c)']).number_format = '#,##0.00'
        worksheet.cell(row=current_row, column=6).fill = sentra_fill
        worksheet.cell(row=current_row, column=6).font = bold_font
        worksheet.cell(row=current_row, column=6).alignment = center_align
        worksheet.cell(row=current_row, column=6).border = thin_border
        worksheet.cell(row=current_row, column=7, value=total_sentra['Setara Beras (d)']).number_format = '#,##0.00'
        worksheet.cell(row=current_row, column=7).fill = sentra_fill
        worksheet.cell(row=current_row, column=7).font = bold_font
        worksheet.cell(row=current_row, column=7).alignment = center_align
        worksheet.cell(row=current_row, column=7).border = thin_border
        worksheet.cell(row=current_row, column=8, value=capaian_sentra).number_format = '0.0"%"'
        worksheet.cell(row=current_row, column=8).fill = sentra_fill
        worksheet.cell(row=current_row, column=8).font = bold_font
        worksheet.cell(row=current_row, column=8).alignment = center_align
        worksheet.cell(row=current_row, column=8).border = thin_border
        current_row += 1

        # Write Kanwil Lainnya header
        cell = worksheet.cell(row=current_row, column=1, value='b) Kanwil Lainnya')
        cell.fill = lainnya_fill
        cell.font = bold_font
        cell.alignment = left_align
        cell.border = thin_border
        worksheet.merge_cells(f'A{current_row}:H{current_row}')
        current_row += 1

        # Write data for Kanwil Lainnya
        for row_data in data_lainnya:
            worksheet.cell(row=current_row, column=1, value=row_data['No']).alignment = center_align
            worksheet.cell(row=current_row, column=1).border = thin_border
            worksheet.cell(row=current_row, column=2, value=row_data['Kanwil']).alignment = left_align
            worksheet.cell(row=current_row, column=2).border = thin_border
            worksheet.cell(row=current_row, column=3, value=row_data['Target Setara Beras']).number_format = '#,##0.00'
            worksheet.cell(row=current_row, column=3).alignment = center_align
            worksheet.cell(row=current_row, column=3).border = thin_border
            worksheet.cell(row=current_row, column=4, value=row_data['Beras (a)']).number_format = '#,##0.00'
            worksheet.cell(row=current_row, column=4).alignment = center_align
            worksheet.cell(row=current_row, column=4).border = thin_border
            worksheet.cell(row=current_row, column=5, value=row_data['GKG (b)']).number_format = '#,##0.00'
            worksheet.cell(row=current_row, column=5).alignment = center_align
            worksheet.cell(row=current_row, column=5).border = thin_border
            worksheet.cell(row=current_row, column=6, value=row_data['GKP (c)']).number_format = '#,##0.00'
            worksheet.cell(row=current_row, column=6).alignment = center_align
            worksheet.cell(row=current_row, column=6).border = thin_border
            worksheet.cell(row=current_row, column=7, value=row_data['Setara Beras (d)']).number_format = '#,##0.00'
            worksheet.cell(row=current_row, column=7).alignment = center_align
            worksheet.cell(row=current_row, column=7).border = thin_border
            worksheet.cell(row=current_row, column=8, value=row_data['Capaian (%)']).number_format = '0.0"%"'
            worksheet.cell(row=current_row, column=8).alignment = center_align
            worksheet.cell(row=current_row, column=8).border = thin_border
            current_row += 1

        # Write Total Kanwil Lainnya
        worksheet.cell(row=current_row, column=1, value='Total Kanwil Lainnya').fill = lainnya_fill
        worksheet.cell(row=current_row, column=1).font = bold_font
        worksheet.cell(row=current_row, column=1).alignment = left_align
        worksheet.cell(row=current_row, column=1).border = thin_border
        worksheet.merge_cells(f'A{current_row}:B{current_row}')
        worksheet.cell(row=current_row, column=3, value=total_lainnya['Target Setara Beras']).number_format = '#,##0.00'
        worksheet.cell(row=current_row, column=3).fill = lainnya_fill
        worksheet.cell(row=current_row, column=3).font = bold_font
        worksheet.cell(row=current_row, column=3).alignment = center_align
        worksheet.cell(row=current_row, column=3).border = thin_border
        worksheet.cell(row=current_row, column=4, value=total_lainnya['Beras (a)']).number_format = '#,##0.00'
        worksheet.cell(row=current_row, column=4).fill = lainnya_fill
        worksheet.cell(row=current_row, column=4).font = bold_font
        worksheet.cell(row=current_row, column=4).alignment = center_align
        worksheet.cell(row=current_row, column=4).border = thin_border
        worksheet.cell(row=current_row, column=5, value=total_lainnya['GKG (b)']).number_format = '#,##0.00'
        worksheet.cell(row=current_row, column=5).fill = lainnya_fill
        worksheet.cell(row=current_row, column=5).font = bold_font
        worksheet.cell(row=current_row, column=5).alignment = center_align
        worksheet.cell(row=current_row, column=5).border = thin_border
        worksheet.cell(row=current_row, column=6, value=total_lainnya['GKP (c)']).number_format = '#,##0.00'
        worksheet.cell(row=current_row, column=6).fill = lainnya_fill
        worksheet.cell(row=current_row, column=6).font = bold_font
        worksheet.cell(row=current_row, column=6).alignment = center_align
        worksheet.cell(row=current_row, column=6).border = thin_border
        worksheet.cell(row=current_row, column=7, value=total_lainnya['Setara Beras (d)']).number_format = '#,##0.00'
        worksheet.cell(row=current_row, column=7).fill = lainnya_fill
        worksheet.cell(row=current_row, column=7).font = bold_font
        worksheet.cell(row=current_row, column=7).alignment = center_align
        worksheet.cell(row=current_row, column=7).border = thin_border
        worksheet.cell(row=current_row, column=8, value=capaian_lainnya).number_format = '0.0"%"'
        worksheet.cell(row=current_row, column=8).fill = lainnya_fill
        worksheet.cell(row=current_row, column=8).font = bold_font
        worksheet.cell(row=current_row, column=8).alignment = center_align
        worksheet.cell(row=current_row, column=8).border = thin_border
        current_row += 1

        # Write TOTAL SE-INDO
        worksheet.cell(row=current_row, column=1, value='TOTAL SE-INDO').fill = seindo_fill
        worksheet.cell(row=current_row, column=1).font = bold_font
        worksheet.cell(row=current_row, column=1).alignment = left_align
        worksheet.cell(row=current_row, column=1).border = thin_border
        worksheet.merge_cells(f'A{current_row}:B{current_row}')
        worksheet.cell(row=current_row, column=3, value=total_seindo['Target Setara Beras']).number_format = '#,##0.00'
        worksheet.cell(row=current_row, column=3).fill = seindo_fill
        worksheet.cell(row=current_row, column=3).font = bold_font
        worksheet.cell(row=current_row, column=3).alignment = center_align
        worksheet.cell(row=current_row, column=3).border = thin_border
        worksheet.cell(row=current_row, column=4, value=total_seindo['Beras (a)']).number_format = '#,##0.00'
        worksheet.cell(row=current_row, column=4).fill = seindo_fill
        worksheet.cell(row=current_row, column=4).font = bold_font
        worksheet.cell(row=current_row, column=4).alignment = center_align
        worksheet.cell(row=current_row, column=4).border = thin_border
        worksheet.cell(row=current_row, column=5, value=total_seindo['GKG (b)']).number_format = '#,##0.00'
        worksheet.cell(row=current_row, column=5).fill = seindo_fill
        worksheet.cell(row=current_row, column=5).font = bold_font
        worksheet.cell(row=current_row, column=5).alignment = center_align
        worksheet.cell(row=current_row, column=5).border = thin_border
        worksheet.cell(row=current_row, column=6, value=total_seindo['GKP (c)']).number_format = '#,##0.00'
        worksheet.cell(row=current_row, column=6).fill = seindo_fill
        worksheet.cell(row=current_row, column=6).font = bold_font
        worksheet.cell(row=current_row, column=6).alignment = center_align
        worksheet.cell(row=current_row, column=6).border = thin_border
        worksheet.cell(row=current_row, column=7, value=total_seindo['Setara Beras (d)']).number_format = '#,##0.00'
        worksheet.cell(row=current_row, column=7).fill = seindo_fill
        worksheet.cell(row=current_row, column=7).font = bold_font
        worksheet.cell(row=current_row, column=7).alignment = center_align
        worksheet.cell(row=current_row, column=7).border = thin_border
        worksheet.cell(row=current_row, column=8, value=capaian_seindo).number_format = '0.0"%"'
        worksheet.cell(row=current_row, column=8).fill = seindo_fill
        worksheet.cell(row=current_row, column=8).font = bold_font
        worksheet.cell(row=current_row, column=8).alignment = center_align
        worksheet.cell(row=current_row, column=8).border = thin_border

        # Set column widths
        worksheet.column_dimensions['A'].width = 5
        worksheet.column_dimensions['B'].width = 45
        worksheet.column_dimensions['C'].width = 18
        worksheet.column_dimensions['D'].width = 12
        worksheet.column_dimensions['E'].width = 12
        worksheet.column_dimensions['F'].width = 12
        worksheet.column_dimensions['G'].width = 18
        worksheet.column_dimensions['H'].width = 14

    output.seek(0)
    return output

def create_excel_export(df, tanggal_kemarin, tanggal_hari_ini):
    """Create Excel file with complex multi-row header and styling"""
    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Write empty dataframe to create workbook
        pd.DataFrame().to_excel(writer, sheet_name='Realisasi', index=False)

        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = workbook['Realisasi']

        # Define styles
        header_fill_dark = PatternFill(start_color='1f497d', end_color='1f497d', fill_type='solid')
        header_fill_light = PatternFill(start_color='4bacc6', end_color='4bacc6', fill_type='solid')
        total_fill = PatternFill(start_color='ddebf7', end_color='ddebf7', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        normal_font = Font(size=10)
        bold_font = Font(bold=True, size=10)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write header row 1
        headers_row1 = ['NO', 'KANWIL', 'TARGET', '', 'REALISASI SD. KEMARIN', '', 'REALISASI HARI INI', '', 'REALISASI SD. TGL', '', 'CAPAIAN', '']
        for col_idx, header in enumerate(headers_row1, start=1):
            cell = worksheet.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill_dark
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        # Merge cells for row 1
        worksheet.merge_cells('A1:A3')  # NO
        worksheet.merge_cells('B1:B3')  # KANWIL
        worksheet.merge_cells('C1:D2')  # TARGET
        worksheet.merge_cells('E1:F1')  # REALISASI SD KEMARIN
        worksheet.merge_cells('G1:H1')  # REALISASI HARI INI
        worksheet.merge_cells('I1:J1')  # REALISASI SD TGL
        worksheet.merge_cells('K1:L2')  # CAPAIAN

        # Write header row 2 (dates)
        worksheet.cell(row=2, column=5, value=tanggal_kemarin.strftime("%d %B %Y")).fill = header_fill_dark
        worksheet.cell(row=2, column=5).font = header_font
        worksheet.cell(row=2, column=5).alignment = center_align
        worksheet.cell(row=2, column=5).border = thin_border
        worksheet.merge_cells('E2:F2')

        worksheet.cell(row=2, column=7, value=tanggal_hari_ini.strftime("%d %B %Y")).fill = header_fill_dark
        worksheet.cell(row=2, column=7).font = header_font
        worksheet.cell(row=2, column=7).alignment = center_align
        worksheet.cell(row=2, column=7).border = thin_border
        worksheet.merge_cells('G2:H2')

        worksheet.cell(row=2, column=9, value=tanggal_hari_ini.strftime("%d %B %Y")).fill = header_fill_dark
        worksheet.cell(row=2, column=9).font = header_font
        worksheet.cell(row=2, column=9).alignment = center_align
        worksheet.cell(row=2, column=9).border = thin_border
        worksheet.merge_cells('I2:J2')

        # Write header row 3 (sub-headers)
        headers_row3 = ['', '', 'BERAS', 'GABAH', 'BERAS', 'GABAH', 'BERAS', 'GABAH', 'BERAS', 'GABAH', '% BERAS', '% GABAH']
        for col_idx, header in enumerate(headers_row3, start=1):
            if col_idx > 2:  # Skip NO and KANWIL
                cell = worksheet.cell(row=3, column=col_idx, value=header)
                cell.fill = header_fill_light
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border

        # Write data rows
        for idx, row in df.iterrows():
            excel_row = idx + 4  # Start from row 4 (after 3 header rows)
            is_total = row['Kanwil'] == 'TOTAL'

            # Write data
            worksheet.cell(row=excel_row, column=1, value=row['NO'])
            worksheet.cell(row=excel_row, column=2, value=row['Kanwil'])
            worksheet.cell(row=excel_row, column=3, value=row['Target Beras'] if pd.notna(row['Target Beras']) else None)
            worksheet.cell(row=excel_row, column=4, value=row['Target Gabah'] if pd.notna(row['Target Gabah']) else None)
            worksheet.cell(row=excel_row, column=5, value=row['Real SD Kemarin Beras'] if pd.notna(row['Real SD Kemarin Beras']) else None)
            worksheet.cell(row=excel_row, column=6, value=row['Real SD Kemarin Gabah'] if pd.notna(row['Real SD Kemarin Gabah']) else None)
            worksheet.cell(row=excel_row, column=7, value=row['Real Hari Ini Beras'] if pd.notna(row['Real Hari Ini Beras']) else None)
            worksheet.cell(row=excel_row, column=8, value=row['Real Hari Ini Gabah'] if pd.notna(row['Real Hari Ini Gabah']) else None)
            worksheet.cell(row=excel_row, column=9, value=row['Real SD Tgl Beras'] if pd.notna(row['Real SD Tgl Beras']) else None)
            worksheet.cell(row=excel_row, column=10, value=row['Real SD Tgl Gabah'] if pd.notna(row['Real SD Tgl Gabah']) else None)
            worksheet.cell(row=excel_row, column=11, value=row['Capaian Beras %'] if pd.notna(row['Capaian Beras %']) else None)
            worksheet.cell(row=excel_row, column=12, value=row['Capaian Gabah %'] if pd.notna(row['Capaian Gabah %']) else None)

            # Apply styling
            for col_idx in range(1, 13):
                cell = worksheet.cell(row=excel_row, column=col_idx)
                cell.border = thin_border
                cell.alignment = center_align if col_idx != 2 else left_align

                if is_total:
                    cell.fill = total_fill
                    cell.font = bold_font
                else:
                    cell.font = normal_font

                # Format numbers
                if col_idx >= 3 and col_idx <= 10:
                    cell.number_format = '#,##0.00'
                elif col_idx >= 11:
                    cell.number_format = '0.0"%"'

        # Set column widths
        worksheet.column_dimensions['A'].width = 5
        worksheet.column_dimensions['B'].width = 40
        for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
            worksheet.column_dimensions[col].width = 12

    output.seek(0)
    return output

def main():
    # Header with gradient background
    st.markdown("""
    <div class="dashboard-header">
        <h1 class="dashboard-title">Dashboard Pengadaan BULOG</h1>
        <p class="dashboard-subtitle">Sistem Informasi Realisasi Pengadaan Gabah & Beras</p>
    </div>
    """, unsafe_allow_html=True)

    # ===== SIDEBAR MENU =====
    with st.sidebar:
        st.markdown("### 📋 Menu")
        menu_option = st.radio(
            "Pilih Menu:",
            options=["📊 Dashboard Realisasi", "📁 Kelola Data"],
            label_visibility="collapsed"
        )

    # ===== MENU: KELOLA DATA =====
    if menu_option == "📁 Kelola Data":
        # CSS khusus untuk page Kelola Data - text berwarna gelap
        st.markdown("""
        <style>
            /* SUPER AGGRESSIVE: Override ALL text in main area - HIGHEST PRIORITY */
            .main,
            .main *,
            .main *::before,
            .main *::after {
                color: #1f497d !important;
            }

            /* Force ALL elements */
            div.main * {
                color: #1f497d !important;
            }

            /* Streamlit markdown containers */
            .main .stMarkdown,
            .main .stMarkdown *,
            .main [data-testid="stMarkdownContainer"],
            .main [data-testid="stMarkdownContainer"] * {
                color: #1f497d !important;
            }

            /* Headers - all levels with all possible selectors */
            .main h1, .main h2, .main h3, .main h4, .main h5, .main h6,
            .main .stMarkdown h1, .main .stMarkdown h2, .main .stMarkdown h3,
            .main .stMarkdown h4, .main .stMarkdown h5, .main .stMarkdown h6 {
                color: #1f497d !important;
            }

            /* Paragraphs and divs */
            .main p, .main span, .main div {
                color: #1f497d !important;
            }

            /* Element containers */
            .main .element-container,
            .main .element-container * {
                color: #1f497d !important;
            }

            /* Block container */
            .main [data-testid="stVerticalBlock"],
            .main [data-testid="stVerticalBlock"] * {
                color: #1f497d !important;
            }

            /* Info, warning, success, error boxes */
            .stAlert, .stAlert * {
                color: #1f497d !important;
            }

            /* Expander */
            .streamlit-expanderHeader, .streamlit-expanderHeader *,
            .main [data-testid="stExpander"],
            .main [data-testid="stExpander"] summary,
            .main [data-testid="stExpander"] summary *,
            .main details summary,
            .main details summary * {
                color: #1f497d !important;
            }

            /* Labels - all types */
            .main label, .main label *,
            .main .stRadio > label,
            .main .stRadio > label *,
            .main [data-baseweb="radio"] ~ *,
            .main .row-widget label {
                color: #1f497d !important;
            }

            /* Radio buttons - ULTRA AGGRESSIVE */
            .main .stRadio,
            .main .stRadio *,
            .main .stRadio label,
            .main .stRadio label *,
            .main .stRadio label span,
            .main .stRadio label span *,
            .main .stRadio label p,
            .main .stRadio label p *,
            .main .stRadio div,
            .main .stRadio div *,
            .main [role="radiogroup"],
            .main [role="radiogroup"] *,
            .main [role="radio"],
            .main [role="radio"] *,
            .main [role="radio"] + *,
            .main [role="radio"] ~ *,
            .main [data-baseweb="radio"],
            .main [data-baseweb="radio"] *,
            .main [data-baseweb="radio"] + *,
            .main [data-baseweb="radio"] ~ * {
                color: #1f497d !important;
            }

            /* File uploader */
            .main .stFileUploader label, .main .stFileUploader * {
                color: #1f497d !important;
            }

            /* Metrics */
            .main [data-testid="stMetric"] * {
                color: #1f497d !important;
            }

            /* Text input, number input, etc */
            .main .stTextInput label,
            .main .stNumberInput label,
            .main .stSelectbox label {
                color: #1f497d !important;
            }

            /* Checkbox */
            .main .stCheckbox label {
                color: #1f497d !important;
            }

            /* EXCEPTIONS - Keep these WHITE */

            /* Sidebar - tetap putih */
            [data-testid="stSidebar"],
            [data-testid="stSidebar"] *,
            [data-testid="stSidebar"] h1,
            [data-testid="stSidebar"] h2,
            [data-testid="stSidebar"] h3,
            [data-testid="stSidebar"] p,
            [data-testid="stSidebar"] span,
            [data-testid="stSidebar"] div,
            [data-testid="stSidebar"] label {
                color: white !important;
            }

            /* Dashboard header - tetap putih */
            .dashboard-header,
            .dashboard-header * {
                color: white !important;
            }

            /* Buttons - keep white text */
            .main button,
            .main button *,
            .stDownloadButton button,
            .stDownloadButton button * {
                color: white !important;
            }
            .st-emotion-cache-11ofl8m {
                    background-color: #1f497d !important;
                    }
            
                    .stVerticalBlock .st-emotion-cache-tn0cau .ek2vi383{
                    background-color: #1f497d !important;
                    }

            .st-emotion-cache-fis6aj.e16n7gab7 {
                background-color: #1f497d !important;
                margin: 10px;
                padding: 1rem;
                border-radius: 20px;
            }
            .stElementContainer.element-container.st-key-select_dataframe.st-emotion-cache-zh2fnc.ek2vi381 {
                background-color: #1f497d !important;
                padding: 1rem;
                border-radius: 10px;
            }
                    .stProgress {
                color: black;
            }

        </style>

        <script>
            // JavaScript untuk memaksa warna dark pada semua element
            function forceTextColor() {
                const mainArea = document.querySelector('.main');
                if (mainArea) {
                    // Get all elements in main area
                    const allElements = mainArea.querySelectorAll('*');

                    allElements.forEach(el => {
                        // Skip sidebar, buttons, and dashboard header
                        const isSidebar = el.closest('[data-testid="stSidebar"]');
                        const isButton = el.tagName === 'BUTTON' || el.closest('button');
                        const isDashboardHeader = el.classList.contains('dashboard-header') || el.closest('.dashboard-header');

                        if (!isSidebar && !isButton && !isDashboardHeader) {
                            el.style.setProperty('color', '#1f497d', 'important');
                        }
                    });
                }
            }

            // Run immediately
            forceTextColor();

            // Run after a delay to catch dynamically loaded content
            setTimeout(forceTextColor, 100);
            setTimeout(forceTextColor, 300);
            setTimeout(forceTextColor, 500);
            setTimeout(forceTextColor, 1000);

            // Set up observer for dynamic content
            const observer = new MutationObserver(forceTextColor);
            const mainArea = document.querySelector('.main');
            if (mainArea) {
                observer.observe(mainArea, {
                    childList: true,
                    subtree: true,
                    attributes: true,
                    attributeFilter: ['style', 'class']
                });
            }
        </script>
        """, unsafe_allow_html=True)

        # Inject JavaScript using components
        import streamlit.components.v1 as st_components
        st_components.html("""
        <script>
            (function() {
                function forceTextColor() {
                    const mainArea = parent.document.querySelector('.main');
                    if (mainArea) {
                        const allElements = mainArea.querySelectorAll('*');
                        allElements.forEach(el => {
                            const isSidebar = el.closest('[data-testid="stSidebar"]');
                            const isButton = el.tagName === 'BUTTON' || el.closest('button');
                            const isDashboardHeader = el.classList.contains('dashboard-header') || el.closest('.dashboard-header');

                            if (!isSidebar && !isButton && !isDashboardHeader) {
                                el.style.setProperty('color', '#1f497d', 'important');
                            }
                        });
                    }
                }

                setTimeout(forceTextColor, 100);
                setTimeout(forceTextColor, 500);
                setTimeout(forceTextColor, 1000);
                setTimeout(forceTextColor, 2000);

                const observer = new MutationObserver(forceTextColor);
                const mainArea = parent.document.querySelector('.main');
                if (mainArea) {
                    observer.observe(mainArea, {
                        childList: true,
                        subtree: true
                    });
                }
            })();
        </script>
        """, height=0)

        st.markdown('<h2 style="color: #1f497d;">📁 Kelola Data Excel</h2>', unsafe_allow_html=True)

        # Pilihan dataframe yang akan dikelola
        st.markdown('<h3 style="color: #1f497d;">📊 Pilih Data yang Akan Dikelola</h3>', unsafe_allow_html=True)
        selected_dataframe = st.radio(
            "Pilih dataframe:",
            options=["📈 Data Realisasi", "🎯 Target Kanwil", "🏢 Target Kancab"],
            help="Pilih data mana yang akan diupdate/replace",
            horizontal=True,
            key="select_dataframe"
        )

        # Map selection to variable names
        df_map = {
            "📈 Data Realisasi": ("df_realisasi", "Export"),
            "🎯 Target Kanwil": ("df_target_kanwil", "Target Kanwil"),
            "🏢 Target Kancab": ("df_target_kancab", "Target Kancab")
        }
        selected_df_name, selected_sheet_name = df_map[selected_dataframe]

        # Load existing data
        st.info(f"⏳ Memuat data existing untuk {selected_dataframe}...")
        df_realisasi, df_target_kanwil, df_target_kancab = load_all_data_with_progress()

        # Get the selected dataframe
        df_existing = {
            "df_realisasi": df_realisasi,
            "df_target_kanwil": df_target_kanwil,
            "df_target_kancab": df_target_kancab
        }[selected_df_name]

        st.success(f"✅ {selected_dataframe} dimuat: **{len(df_existing):,}** records")

        # File upload section
        st.markdown("---")
        st.markdown('<h3 style="color: #1f497d;">📤 Upload File Excel</h3>', unsafe_allow_html=True)
        uploaded_file = st.file_uploader(
            "Pilih file Excel (.xlsx atau .xls)",
            type=['xlsx', 'xls'],
            help="Upload file Excel untuk update data"
        )

        if uploaded_file is not None:
            try:
                st.markdown("""
                <style>
                 
                    .st-emotion-cache-11ofl8m {
                            background-color: #1f497d !important;
                            }
                    
                            .stVerticalBlock .st-emotion-cache-tn0cau .ek2vi383{
                            background-color: #1f497d !important;
                            }

                    .st-emotion-cache-fis6aj.e16n7gab7 {
                        background-color: #1f497d !important;
                        margin: 10px;
                        padding: 1rem;
                        border-radius: 20px;
                    }
                    .st-emotion-cache-13na8ym.ef5ck4x1 {
                        background-color: #1f497d !important;
                    }
                        label.st-am.st-ae.st-an.st-ai.st-g4.st-aq.st-ao.st-at 
                    p{
                        color: black !important;
                    }
                </style>
                """, unsafe_allow_html=True)
                # Validasi file Excel dan deteksi sheet
                progress_bar = st.progress(0, "🔍 Memvalidasi file Excel...")

                # Read Excel file to get sheet names
                excel_file = pd.ExcelFile(uploaded_file, engine='openpyxl')
                available_sheets = excel_file.sheet_names
                progress_bar.progress(20, "📋 Mendeteksi sheet yang tersedia...")

                # Pilihan sheet jika ada lebih dari satu
                if len(available_sheets) > 1:
                    st.info(f"📋 File Excel memiliki {len(available_sheets)} sheet: {', '.join(available_sheets)}")
                    selected_sheet = st.selectbox(
                        "Pilih sheet yang akan diproses:",
                        options=available_sheets,
                        index=available_sheets.index(selected_sheet_name) if selected_sheet_name in available_sheets else 0,
                        key="select_sheet"
                    )
                else:
                    selected_sheet = available_sheets[0]
                    st.info(f"📋 Menggunakan sheet: **{selected_sheet}**")

                progress_bar.progress(40, f"📖 Membaca data dari sheet '{selected_sheet}'...")
                df_new = pd.read_excel(uploaded_file, sheet_name=selected_sheet, engine='openpyxl')
                progress_bar.progress(60, "✅ Data berhasil dibaca")
                progress_bar.progress(100, "✅ Validasi selesai")
                progress_bar.empty()

                st.success(f"✅ File berhasil dibaca dari sheet **'{selected_sheet}'**: **{len(df_new):,}** records")

                # Show preview
                st.markdown('<p style="color: #1f497d; font-weight: 600; margin-bottom: 0;">👁️ Preview Data Baru (5 baris pertama)</p>', unsafe_allow_html=True)
                with st.expander("", expanded=False):
                    st.dataframe(df_new.head(), use_container_width=True)

                # Check columns
                st.markdown('<h4 style="color: #1f497d;">🔍 Validasi Kolom</h4>', unsafe_allow_html=True)
                missing_cols = set(df_existing.columns) - set(df_new.columns)
                extra_cols = set(df_new.columns) - set(df_existing.columns)

                if missing_cols:
                    st.warning(f"⚠️ Kolom yang hilang dari file baru: {', '.join(missing_cols)}")
                if extra_cols:
                    st.info(f"ℹ️ Kolom tambahan di file baru: {', '.join(extra_cols)}")

                if not missing_cols:
                    st.success("✅ Semua kolom cocok!")

                # Mode selection
                st.markdown('<h4 style="color: #1f497d;">⚙️ Mode Upload</h4>', unsafe_allow_html=True)
                st.markdown("""
                <style>
                    .st-emotion-cache-6px8kg 
                    .st-ae.st-af.st-ag.st-ah.st-ai.st-aj.st-ak.st-al {
                        background-color: #1f497d !important;
                        padding: 1rem;
                        border-radius: 15px;
                    }
                </style>
                """, unsafe_allow_html=True)
                upload_mode = st.radio(
                    "label",
                    options=["🔄 Append (Tambahkan data baru)", "🔁 Replace (Ganti semua data)"],
                    help="Append: Tambahkan hanya data yang belum ada | Replace: Hapus semua data lama dan ganti dengan data baru",
                    label_visibility="collapsed"
                )
                if upload_mode == "🔄 Append (Tambahkan data baru)":
                    st.info("""
                    **Mode Append:**
                    - Data baru akan dibandingkan dengan data existing
                    - Hanya baris yang **tidak duplikat** (semua kolom sama) yang akan ditambahkan
                    - Data existing tetap aman
                    """)

                    # Find duplicates
                    st.markdown('<h4 style="color: #1f497d;">🔍 Analisis Duplikasi</h4>', unsafe_allow_html=True)

                    # Ensure both dataframes have same columns for comparison
                    common_cols = list(set(df_existing.columns) & set(df_new.columns))

                    # Progress bar untuk analisis duplikasi
                    dup_progress = st.progress(0, "🔍 Memulai analisis duplikasi...")

                    # OPTIMIZED: Gunakan hash untuk perbandingan lebih cepat
                    # Create hash dari semua kolom untuk setiap baris
                    def create_row_hash(row):
                        """Create hash dari semua nilai dalam row"""
                        return hash(tuple(str(v) for v in row))

                    dup_progress.progress(20, "📊 Membuat hash untuk data existing...")
                    # Create hash untuk data existing
                    df_existing_subset = df_existing[common_cols].copy()
                    existing_hashes = set(
                        df_existing_subset.apply(create_row_hash, axis=1)
                    )

                    dup_progress.progress(50, "🆕 Membuat hash untuk data baru...")
                    # Create hash untuk data baru dan cek duplikasi
                    df_new_subset = df_new[common_cols].copy()
                    new_hashes = df_new_subset.apply(create_row_hash, axis=1)

                    dup_progress.progress(75, "🔍 Mengidentifikasi data unik...")
                    # Identify unique rows (hash tidak ada di existing)
                    unique_mask = ~new_hashes.isin(existing_hashes)

                    num_unique = unique_mask.sum()
                    num_duplicates = len(df_new) - num_unique

                    # Get unique records from original df_new
                    df_unique = df_new[unique_mask].copy()

                    dup_progress.progress(100, "✅ Analisis duplikasi selesai")
                    dup_progress.empty()

                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("📊 Total Data Baru", f"{len(df_new):,}")
                    with col2:
                        st.metric("✅ Data Unik (akan ditambahkan)", f"{num_unique:,}")
                    with col3:
                        st.metric("⚠️ Data Duplikat (akan dilewati)", f"{num_duplicates:,}")

                    if num_unique > 0:
                        with st.expander(f"👁️ Preview {num_unique} Data Unik yang Akan Ditambahkan"):
                            st.dataframe(df_unique.head(10), use_container_width=True)

                        if st.button("✅ Append Data Unik", type="primary", use_container_width=True):
                                try:
                                    # Progress bar untuk proses append
                                    append_progress = st.progress(0, "🔄 Memulai proses append...")

                                    # Combine data
                                    append_progress.progress(10, f"📊 Menggabungkan {len(df_existing):,} + {num_unique:,} records...")
                                    df_combined = pd.concat([df_existing, df_unique], ignore_index=True)

                                    # AGGRESSIVE data cleaning untuk prevent corruption
                                    append_progress.progress(30, "🧹 Membersihkan data untuk Excel...")

                                    # 1. Replace infinity values
                                    df_combined = df_combined.replace([np.inf, -np.inf], np.nan)

                                    # 2. Clean each column based on type
                                    append_progress.progress(40, "🔧 Mengonversi tipe data...")
                                    for col in df_combined.columns:
                                        # Handle numeric columns
                                        if pd.api.types.is_numeric_dtype(df_combined[col]):
                                            df_combined[col] = df_combined[col].where(pd.notna(df_combined[col]), None)
                                        # Handle datetime columns - convert to string
                                        elif pd.api.types.is_datetime64_any_dtype(df_combined[col]):
                                            df_combined[col] = df_combined[col].astype(str).replace('NaT', '')
                                        # Handle object columns - convert all to string
                                        elif df_combined[col].dtype == 'object':
                                            df_combined[col] = df_combined[col].astype(str).replace('nan', '').replace('None', '')

                                    # 3. Prepare all dataframes for saving
                                    append_progress.progress(60, "📋 Menyiapkan semua sheet...")
                                    # Update the selected dataframe
                                    if selected_df_name == "df_realisasi":
                                        df_to_save_export = df_combined
                                        df_to_save_kanwil = df_target_kanwil
                                        df_to_save_kancab = df_target_kancab
                                    elif selected_df_name == "df_target_kanwil":
                                        df_to_save_export = df_realisasi
                                        df_to_save_kanwil = df_combined
                                        df_to_save_kancab = df_target_kancab
                                    else:  # df_target_kancab
                                        df_to_save_export = df_realisasi
                                        df_to_save_kanwil = df_target_kanwil
                                        df_to_save_kancab = df_combined

                                    # Clean all dataframes
                                    for df_clean in [df_to_save_export, df_to_save_kanwil, df_to_save_kancab]:
                                        df_clean.replace([np.inf, -np.inf], np.nan, inplace=True)
                                        for col in df_clean.columns:
                                            if df_clean[col].dtype == 'object':
                                                df_clean[col] = df_clean[col].astype(str).replace('nan', '').replace('None', '')

                                    # Save to Excel
                                    output_file = "assets/hasil_gabungan.xlsx"
                                    append_progress.progress(80, f"💾 Menyimpan ke {output_file}...")

                                    with pd.ExcelWriter(output_file, engine='openpyxl', mode='w') as writer:
                                        df_to_save_export.to_excel(writer, sheet_name='Export', index=False)
                                        df_to_save_kanwil.to_excel(writer, sheet_name='Target Kanwil', index=False)
                                        df_to_save_kancab.to_excel(writer, sheet_name='Target Kancab', index=False)

                                    append_progress.progress(100, "✅ Append selesai!")
                                    append_progress.empty()

                                    st.success(f"""
                                    ✅ **Append berhasil!**
                                    - Dataframe: **{selected_dataframe}**
                                    - Total data sebelumnya: **{len(df_existing):,}** records
                                    - Data unik ditambahkan: **{num_unique:,}** records
                                    - Total data sekarang: **{len(df_combined):,}** records
                                    - File disimpan: `{output_file}`
                                    """)

                                    st.balloons()

                                    # Auto reload dengan rerun
                                    st.info("🔄 Data berhasil disimpan! Memuat ulang data...")

                                    # Clear cache agar data ter-refresh
                                    st.cache_data.clear()

                                    time.sleep(1)  # Brief pause for user to see success
                                    st.rerun()

                                except Exception as e:
                                    st.error(f"❌ Error saat menyimpan file: {str(e)}")
                                    st.error("Detail error untuk debugging:")
                                    st.code(str(e))

                                    # Show problematic data info
                                    st.warning("Mencoba analisis data...")
                                    try:
                                        st.write("Tipe data di df_combined:")
                                        st.write(df_combined.dtypes)
                                        st.write("Info nilai NaN:")
                                        st.write(df_combined.isna().sum())
                                    except:
                                        pass
                    else:
                        st.warning("⚠️ Tidak ada data unik untuk ditambahkan. Semua data sudah ada di database.")

                else:  # Replace mode
                    st.warning("""
                    **⚠️ Mode Replace:**
                    - Semua data existing akan **DIHAPUS**
                    - Data baru akan menggantikan sepenuhnya
                    - **TIDAK BISA dibatalkan!**
                    """)

                    col1, col2 = st.columns(2)
                    with col1:
                        st.metric("📊 Data Lama", f"{len(df_existing):,} records")
                    with col2:
                        st.metric("🆕 Data Baru", f"{len(df_new):,} records")

                    # Confirmation checkbox
                    confirm_replace = st.checkbox(
                        f"⚠️ Saya mengerti bahwa ini akan menghapus SEMUA data lama di {selected_dataframe}",
                        value=False
                    )

                    st.markdown("---")
                    if st.button(
                        "🔁 Replace Semua Data",
                        type="primary",
                        disabled=not confirm_replace,
                        use_container_width=True
                    ):
                            try:
                                # Progress bar untuk proses replace
                                replace_progress = st.progress(0, "🔄 Memulai proses replace...")

                                # Clean data before saving to prevent corruption
                                replace_progress.progress(20, "🧹 Membersihkan data baru...")
                                df_new_clean = df_new.copy().replace([np.inf, -np.inf], np.nan)

                                # Convert datetime columns properly
                                replace_progress.progress(40, "🔧 Mengonversi tipe data...")
                                for col in df_new_clean.columns:
                                    if df_new_clean[col].dtype == 'object':
                                        try:
                                            df_new_clean[col] = pd.to_datetime(df_new_clean[col], errors='ignore')
                                        except:
                                            pass

                                # Prepare all dataframes for saving
                                replace_progress.progress(60, "📋 Menyiapkan semua sheet...")
                                # Update the selected dataframe
                                if selected_df_name == "df_realisasi":
                                    df_to_save_export = df_new_clean
                                    df_to_save_kanwil = df_target_kanwil
                                    df_to_save_kancab = df_target_kancab
                                elif selected_df_name == "df_target_kanwil":
                                    df_to_save_export = df_realisasi
                                    df_to_save_kanwil = df_new_clean
                                    df_to_save_kancab = df_target_kancab
                                else:  # df_target_kancab
                                    df_to_save_export = df_realisasi
                                    df_to_save_kanwil = df_target_kanwil
                                    df_to_save_kancab = df_new_clean

                                # Clean all dataframes
                                for df_clean in [df_to_save_export, df_to_save_kanwil, df_to_save_kancab]:
                                    df_clean.replace([np.inf, -np.inf], np.nan, inplace=True)
                                    for col in df_clean.columns:
                                        if df_clean[col].dtype == 'object':
                                            df_clean[col] = df_clean[col].astype(str).replace('nan', '').replace('None', '')

                                # Save new data
                                output_file = "assets/hasil_gabungan.xlsx"
                                replace_progress.progress(80, f"💾 Menyimpan ke {output_file}...")

                                with pd.ExcelWriter(output_file, engine='openpyxl', mode='w') as writer:
                                    df_to_save_export.to_excel(writer, sheet_name='Export', index=False)
                                    df_to_save_kanwil.to_excel(writer, sheet_name='Target Kanwil', index=False)
                                    df_to_save_kancab.to_excel(writer, sheet_name='Target Kancab', index=False)

                                replace_progress.progress(100, "✅ Replace selesai!")
                                replace_progress.empty()

                                st.success(f"""
                                ✅ **Replace berhasil!**
                                - Dataframe: **{selected_dataframe}**
                                - Data lama dihapus: **{len(df_existing):,}** records
                                - Data baru disimpan: **{len(df_new):,}** records
                                - File disimpan: `{output_file}`
                                """)

                                st.balloons()

                                # Auto reload dengan rerun
                                st.info("🔄 Data berhasil disimpan! Memuat ulang data...")

                                # Clear cache agar data ter-refresh
                                st.cache_data.clear()

                                time.sleep(1)  # Brief pause for user to see success
                                st.rerun()

                            except Exception as e:
                                st.error(f"❌ Error saat menyimpan file: {str(e)}")
                                st.error("Detail error untuk debugging:")
                                st.code(str(e))

            except Exception as e:
                st.error(f"❌ Error saat membaca file Excel: {str(e)}")
                st.info("Pastikan file Excel memiliki sheet 'Export' dengan format yang benar")

        return  # Exit early, don't show dashboard

    # ===== MENU: DASHBOARD REALISASI (DEFAULT) =====
    # Load data with progress bar
    df_realisasi, df_target_kanwil, df_target_kancab = load_all_data_with_progress()

    colA, colB, colC = st.columns(3)
    st.markdown("""
        <style>
            h4 {
                color: #1f497d !important;
                }
            div .st-am{
                background-color: #1f497d !important;
                }

            span .st-ae{
                background-color: #1f497d !important;
                }
                
            /* Judul Filter Biru */
            .filter-title {
                color: #1f497d !important;
                font-weight: 600 !important;
            }

            /* Styling Multiselect & Date Input */
            div[data-baseweb="select"] > div {
                border: 1px solid #1f497d !important;
                box-shadow: none !important;
            }

            div[data-baseweb="select"] > div:hover {
                border-color: #1f497d !important;
            }

            /* Text input & date input border biru */
            input[type="text"], input[type="date"] {
                border: 1px solid #1f497d !important;
                color: white !important;
            }

            /* Arrow icon warna biru */
            svg {
                color: #1f497d !important;
            }

            /* Text di dalam dropdown */
            .css-1n7v3ny-option, .css-1n7v3ny-option * {
                color: white !important;
            }
        </style>
        """, unsafe_allow_html=True)


    # Filter 1: Akun Analitik
    with colA:
        st.markdown("#### Akun Analitik")
        all_akun_analitik = sorted([a for a in df_realisasi['Akun Analitik'].unique() if pd.notna(a)])

        selected_akun_analitik = st.multiselect(
            "Pilih Akun Analitik:",
            options=all_akun_analitik,
            default=["PSO"],     # default hanya PSO
            label_visibility="collapsed",
            key="filter_akun_analitik"
        )

    # Filter 2: Kanwil (untuk Line Chart & Tabel Kancab)
    with colB:
        st.markdown("#### Kanwil")
        all_kanwil = sorted([k for k in df_realisasi['kanwil'].unique() if pd.notna(k)])

        selected_kanwil = st.multiselect(
            "Pilih Kanwil:",
            options=all_kanwil,
            default=["11001 - KANTOR WILAYAH JATENG"],
            # default=["13001 - KANTOR WILAYAH JATIM"],
            label_visibility="collapsed",
            key="filter_kanwil"
        )

    # Filter 3: Date Range (UNTUK SEMUA VISUALISASI)
    with colC:
        st.markdown("#### Periode")

        min_date_penerimaan = df_realisasi['Tanggal Penerimaan'].min().date() if pd.notna(df_realisasi['Tanggal Penerimaan'].min()) else datetime(2025, 1, 1).date()
        # max_date_penerimaan = df_realisasi['Tanggal Penerimaan'].max().date() if pd.notna(df_realisasi['Tanggal Penerimaan'].max()) else datetime.now().date()
        max_date_penerimaan = datetime.now().date()

        date_range = st.date_input(
            "Range Tanggal Penerimaan:",
            value=(min_date_penerimaan, max_date_penerimaan),
            min_value=min_date_penerimaan,
            max_value=max_date_penerimaan,
            label_visibility="collapsed",
            key="filter_date_range"
        )

    # Apply filters for general data
    df_filtered = df_realisasi.copy()

    # Filter 1: Akun Analitik (berpengaruh pada SEMUA visualisasi)
    if selected_akun_analitik:
        df_filtered = df_filtered[df_filtered['Akun Analitik'].isin(selected_akun_analitik)]

    # Filter 2: Periode (berpengaruh pada SEMUA visualisasi)
    # Handle date_range properly
    if isinstance(date_range, tuple) and len(date_range) == 2:
        start_date, end_date = date_range
        df_filtered = df_filtered[
            (df_filtered['Tanggal Penerimaan'].dt.date >= start_date) &
            (df_filtered['Tanggal Penerimaan'].dt.date <= end_date)
        ]
    elif isinstance(date_range, tuple) and len(date_range) == 1:
        start_date = end_date = date_range[0]
        df_filtered = df_filtered[df_filtered['Tanggal Penerimaan'].dt.date == start_date]
    else:
        # Single date selected
        start_date = end_date = date_range
        df_filtered = df_filtered[df_filtered['Tanggal Penerimaan'].dt.date == start_date]

    # Filter for LINE CHART & TABEL KANCAB (tambahan filter Kanwil)
    df_chart = df_filtered.copy()
    df_kancab = df_filtered.copy()
    if selected_kanwil:
        df_chart = df_chart[df_chart['kanwil'].isin(selected_kanwil)]
        df_kancab = df_kancab[df_kancab['kanwil'].isin(selected_kanwil)]

    # TABEL SUMMARY (use ALL kanwil, sudah terfilter by Akun Analitik dan Periode)
    df_summary = df_filtered.copy()

    # Display metrics
    col1, col2, col3, col4 = st.columns(4)

    # Metric 1: Realisasi Setara Beras Hari Ini (end_date only)
    # Filter: Kanwil + Periode (hanya end_date)
    with col1:
        # Data untuk hari ini (end_date) dengan filter Kanwil
        if selected_kanwil:
            df_hari_ini = df_filtered[
                (df_filtered['Tanggal Penerimaan'].dt.date == end_date) &
                (df_filtered['kanwil'].isin(selected_kanwil))
            ]
        else:
            # Jika tidak ada kanwil yang dipilih, gunakan semua kanwil
            df_hari_ini = df_filtered[df_filtered['Tanggal Penerimaan'].dt.date == end_date]

        realisasi_hari_ini = calculate_setara_beras(df_hari_ini)
        st.metric(f"📊 Realisasi Hari Ini ({end_date})", f"{realisasi_hari_ini:,.2f} Ton")

    # Metric 2: Total Realisasi Setara Beras (start_date - end_date)
    # Filter: Kanwil + Periode (start_date to end_date)
    with col2:
        # Data untuk periode (start_date to end_date) dengan filter Kanwil
        if selected_kanwil:
            df_periode = df_filtered[df_filtered['kanwil'].isin(selected_kanwil)]
        else:
            df_periode = df_filtered

        total_realisasi = calculate_setara_beras(df_periode)
        st.metric(f"📈 Total Realisasi ({start_date}-{end_date})", f"{total_realisasi:,.2f} Ton")

    # Metric 3: Target Setara Beras (Kanwil)
    # Filter: Kanwil
    with col3:
        # Ambil target dari sheet "Target" berdasarkan filter Kanwil
        if selected_kanwil:
            target_setara_beras = df_target_kanwil[
                df_target_kanwil['kanwil'].isin(selected_kanwil)
            ]['Target Setara Beras'].sum()
        else:
            # Jika tidak ada kanwil yang dipilih, gunakan semua kanwil
            target_setara_beras = df_target_kanwil['Target Setara Beras'].sum()

        st.metric(f"🎯 Target {selected_kanwil}", f"{target_setara_beras:,.2f} Ton")

    # Metric 4: Sisa Target Setara Beras (Kanwil)
    # Filter: Kanwil
    with col4:
        # Sisa = Target - Total Realisasi
        sisa_target = target_setara_beras - total_realisasi

        # Hitung persentase capaian untuk delta
        if target_setara_beras > 0:
            persentase_capaian = (total_realisasi / target_setara_beras) * 100
        else:
            persentase_capaian = 0

        st.metric(
            f"📉 Sisa Target {selected_kanwil}",
            f"{sisa_target:,.2f} Ton"
        )

    # ===== TABEL SUMMARY (DI ATAS LINE CHART) =====
    st.markdown('<div class="chart-title">📋 Tabel Realisasi per-Kanwil</div>', unsafe_allow_html=True)

    if not df_summary.empty:
        # Create summary table data (df_summary is already filtered by Akun Analitik and Periode)
        data_sentra, data_lainnya = create_summary_table(df_summary, df_target_kanwil)

        # Render HTML table
        html_summary, total_sentra, total_lainnya, total_seindo, capaian_sentra, capaian_lainnya, capaian_seindo = render_summary_table_html(
            data_sentra, data_lainnya, start_date, end_date
        )

        # Download button for summary table
        excel_summary = create_summary_excel_export(
            data_sentra, data_lainnya, start_date, end_date,
            total_sentra, total_lainnya, total_seindo,
            capaian_sentra, capaian_lainnya, capaian_seindo
        )

        st.download_button(
            label="📥 Download Tabel Realisasi Kanwil (Excel)",
            data=excel_summary,
            file_name=f"summary_realisasi_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # Use components.html for better HTML rendering
        import streamlit.components.v1 as components
        components.html(html_summary, height=800, scrolling=True)
    else:
        st.warning("⚠️ Tidak ada data untuk ditampilkan")

    linech, barch = st.columns([2, 1])
    # ===== LINE CHART =====
    with linech:
        st.markdown('<div class="chart-title">📈 Tren Realisasi Kanwil</div>', unsafe_allow_html=True)

        try:
            if not df_chart.empty:
                fig = create_line_chart(df_chart, start_date, end_date)
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.warning("⚠️ Tidak ada data untuk ditampilkan dengan filter yang dipilih")
        except Exception as e:
            st.error(f"⚠️ Error saat membuat line chart: {str(e)}")
            st.info("Silakan selesaikan pemilihan filter untuk melihat chart")

    with barch:
        # ===== BAR CHART 7 HARI TERAKHIR =====
        st.markdown('<div class="chart-title">📊 Realisasi 7 Hari Terakhir</div>', unsafe_allow_html=True)

        try:
            if not df_chart.empty:
                fig_bar = create_bar_chart_7days(df_chart, end_date)
                st.plotly_chart(fig_bar, use_container_width=True)
            else:
                st.warning("⚠️ Tidak ada data untuk ditampilkan dengan filter yang dipilih")
        except Exception as e:
            st.error(f"⚠️ Error saat membuat bar chart: {str(e)}")
            st.info("Silakan selesaikan pemilihan filter untuk melihat chart")

    # ===== TABEL REALISASI DETAIL PER-KANCAB =====
    st.markdown('<div class="chart-title">📊 Tabel Realisasi per-Kancab</div>', unsafe_allow_html=True)

    if not df_kancab.empty:
        # Create Kancab table (df_kancab is already filtered by Akun Analitik, Periode, and Kanwil)
        kancab_df = create_kancab_table(df_kancab, df_target_kancab)
        excel_kancab = create_kancab_excel_export(kancab_df, end_date)
        st.download_button(
                label="📥 Download Tabel Realisasi Kancab (Excel)",
                data=excel_kancab,
                file_name=f"realisasi_kancab_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        if not kancab_df.empty:
            # Render HTML table
            html_kancab = render_kancab_table_html(kancab_df, start_date, end_date)
            # Use components.html untuk scrolling seperti tabel kanwil
            import streamlit.components.v1 as components
            components.html(html_kancab, height=800, scrolling=True)

            # Download button (Excel with same style as HTML)
            
        else:
            st.warning("⚠️ Tidak ada data Kancab untuk ditampilkan")
    else:
        st.warning("⚠️ Tidak ada data untuk ditampilkan")
    st.markdown("""
    <style>
    .stDownloadButton>button {
        background: linear-gradient(135deg, #1f497d 0%, #4bacc6 100%) !important;
        color: white !important;
        border-radius: 8px !important;
        padding: 8px 16px !important;
        border: none !important;
    }
    </style>
""", unsafe_allow_html=True)

if __name__ == "__main__":
    main()